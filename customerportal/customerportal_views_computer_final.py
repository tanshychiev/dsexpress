from datetime import date, datetime, time, timedelta
from decimal import Decimal

import io
import os
import re
import requests

from django.conf import settings
from django.contrib import messages
from django.core.paginator import Paginator
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from openpyxl import Workbook, load_workbook

from masterdata.models import Shipper
from orders.activity import add_order_activity
from orders.audit import add_audit_log
from orders.models import AuditLog, ImportBatch, Order, OrderActivity
from orders.pricing import apply_pricing
from inventory.models import InventorySellerSetting
from inventory.services import (
    get_seller_current_stock,
    get_seller_inventory_setting,
)

from provinceops.models import ProvinceBatch, ProvinceBatchItem
from provincecod.models import ProvinceCODBatch, ProvinceCODItem
from provincecod.services import money
from reports.excel import export_delivery_report_xlsx
from reports.services import (
    get_done_queryset,
    get_pending_queryset,
    group_by_seller,
    classify_row,
    calc_totals,
    get_shipper_name,
    report_money,
)

from .models import SellerUploadBatch, SellerUploadRow
from .views import get_user_seller


ZERO = Decimal("0.00")


ORDER_DAY_START_TIME = time(20, 30)
ORDER_DAY_SHIFT = timedelta(hours=3, minutes=30)
DELIVERING_START_TIME = time(14, 30)


def _uses_830_order_day(seller):
    # Only Starek uses DS order-day cutoff:
    # previous day 8:30 PM through selected date 8:29 PM.
    seller_name = (getattr(seller, "name", "") or "").strip().lower()
    return seller_name == "starek"


def _localize_datetime(value):
    if settings.USE_TZ:
        return timezone.make_aware(value, timezone.get_current_timezone())
    return value


def _report_today(seller=None):
    today = timezone.localdate()

    if _uses_830_order_day(seller):
        now = timezone.localtime(timezone.now())
        return (now + ORDER_DAY_SHIFT).date()

    return today


def _created_range_datetimes(start_date, end_date, seller=None):
    if _uses_830_order_day(seller):
        # Starek order day example:
        # 04/07/2026 = 03/07/2026 20:30 through 04/07/2026 20:29:59
        start_dt = datetime.combine(
            start_date - timedelta(days=1),
            ORDER_DAY_START_TIME,
        )
        end_dt = datetime.combine(
            end_date,
            ORDER_DAY_START_TIME,
        )
    else:
        # Normal shops use normal calendar day.
        start_dt = datetime.combine(start_date, time.min)
        end_dt = datetime.combine(end_date + timedelta(days=1), time.min)

    return _localize_datetime(start_dt), _localize_datetime(end_dt)


# =========================================================
# COMMON HELPERS
# =========================================================

def _parse_date(value):
    try:
        return date.fromisoformat((value or "").strip())
    except (TypeError, ValueError):
        return None


def _safe_rate(part, total):
    if not total:
        return 0

    try:
        value = round((part * 100.0) / total, 2)
    except (TypeError, ValueError, ZeroDivisionError):
        return 0

    return min(max(value, 0), 100)


def _get_report_range(request, seller=None, default_period="last_30_days"):
    today = _report_today(seller)
    period = (request.GET.get("period") or default_period).strip()

    if period == "today":
        start_date = today
        end_date = today

    elif period == "yesterday":
        start_date = today - timedelta(days=1)
        end_date = start_date

    elif period == "last_3_days":
        start_date = today - timedelta(days=2)
        end_date = today

    elif period == "last_7_days":
        start_date = today - timedelta(days=6)
        end_date = today

    elif period == "last_30_days":
        start_date = today - timedelta(days=29)
        end_date = today

    elif period == "this_month":
        start_date = today.replace(day=1)
        end_date = today

    elif period == "last_month":
        first_day_this_month = today.replace(day=1)
        end_date = first_day_this_month - timedelta(days=1)
        start_date = end_date.replace(day=1)

    elif period == "custom":
        start_date = _parse_date(request.GET.get("from"))
        end_date = _parse_date(request.GET.get("to"))

        if not start_date or not end_date:
            period = "last_30_days"
            start_date = today - timedelta(days=29)
            end_date = today
        else:
            if start_date > end_date:
                start_date, end_date = end_date, start_date

            start_date = min(start_date, today)
            end_date = min(end_date, today)

            if start_date > end_date:
                start_date = end_date

            if (end_date - start_date).days > 365:
                start_date = end_date - timedelta(days=365)

    else:
        period = "last_30_days"
        start_date = today - timedelta(days=29)
        end_date = today

    if start_date == end_date:
        period_label = start_date.strftime("%d %B %Y")
    else:
        period_label = (
            f"{start_date.strftime('%d %b %Y')} - "
            f"{end_date.strftime('%d %b %Y')}"
        )

    return period, start_date, end_date, period_label


def _get_logged_in_seller(request):
    seller = get_user_seller(request.user)

    if seller is None:
        logout(request)
        return None

    return seller


def _base_seller_orders(seller):
    return Order.objects.filter(
        seller=seller,
        is_deleted=False,
    )


def _pending_statuses():
    return {
        Order.STATUS_CREATED,
        Order.STATUS_INBOUND,
        Order.STATUS_OUT_FOR_DELIVERY,
        Order.STATUS_PROVINCE_ASSIGNED,
        Order.STATUS_RETURN_ASSIGNED,
        getattr(Order, "STATUS_RETURNING", "RETURNING"),
        getattr(Order, "STATUS_PROCESSING", "PROCESSING"),
    }


def _order_delivery_fee(order):
    return getattr(order, "delivery_fee", ZERO) or ZERO


def _order_additional_fee(order):
    return getattr(order, "additional_fee", ZERO) or ZERO


def _order_province_fee(order):
    return getattr(order, "province_fee", ZERO) or ZERO


def _order_total_fee(order):
    return (
        _order_delivery_fee(order)
        + _order_additional_fee(order)
        + _order_province_fee(order)
    )


def _order_cod(order):
    price = getattr(order, "price", ZERO) or ZERO
    cod = getattr(order, "cod", ZERO) or ZERO
    return price if price != ZERO else cod


def _is_province_order(order):
    status = (getattr(order, "status", "") or "").strip().upper()
    province_fee = _order_province_fee(order)

    return (
        status
        == getattr(
            Order,
            "STATUS_PROVINCE_ASSIGNED",
            "PROVINCE_ASSIGNED",
        )
        or province_fee > ZERO
    )


def _province_cod_customer_result(order):
    """Return customer-facing result for orders already sent in Province COD.

    Internal Province COD marks a batch as SENT and keeps collection progress
    on ProvinceCODItem. For customer portal order status, a SENT province COD
    item should behave like Delivered, unless the item itself is marked RETURNED.
    """
    order_id = getattr(order, "id", None)

    if not order_id:
        return None

    try:
        from provincecod.models import ProvinceCODBatch, ProvinceCODItem
    except Exception:
        return None

    try:
        item = (
            ProvinceCODItem.objects
            .select_related("batch")
            .filter(order_id=order_id)
            .exclude(batch__status=ProvinceCODBatch.STATUS_CANCELLED)
            .order_by("-id")
            .first()
        )
    except Exception:
        return None

    if not item:
        return None

    batch_status = (getattr(getattr(item, "batch", None), "status", "") or "").strip().upper()
    cod_status = (getattr(item, "cod_status", "") or "").strip().upper()

    if cod_status == getattr(ProvinceCODItem, "STATUS_RETURNED", "RETURNED"):
        return "returned"

    if batch_status == getattr(ProvinceCODBatch, "STATUS_SENT", "SENT"):
        return "delivered"

    return None


def _is_customer_delivered_order(order):
    status = (getattr(order, "status", "") or "").strip().upper()

    if status in {
        getattr(Order, "STATUS_DELIVERED", "DELIVERED"),
        "DONE",
        "SENT",
    }:
        return True

    return _province_cod_customer_result(order) == "delivered"


def _is_customer_returned_order(order):
    status = (getattr(order, "status", "") or "").strip().upper()

    if status == getattr(Order, "STATUS_RETURNED", "RETURNED"):
        return True

    return _province_cod_customer_result(order) == "returned"


def _get_created_date(order):
    created_at = getattr(order, "created_at", None)

    if not created_at:
        return None

    try:
        if timezone.is_aware(created_at):
            created_at = timezone.localtime(created_at)

        if _uses_830_order_day(getattr(order, "seller", None)):
            return (created_at + ORDER_DAY_SHIFT).date()

        return created_at.date()
    except (AttributeError, TypeError, ValueError):
        return None



def _find_order_inventory_product(order):
    """
    Display helper for customer order list.
    Old orders may store SKU in product_desc. Newer orders may store product name.
    We try SKU first, then product name, so the customer can see both SKU and readable product name.
    """
    try:
        from inventory.models import StockProduct
        from inventory.services import match_product
    except Exception:
        StockProduct = None
        match_product = None

    seller = getattr(order, "seller", None)
    text = (getattr(order, "product_desc", "") or "").strip()

    if not seller or not text:
        return None

    product = None

    if StockProduct:
        product = (
            StockProduct.objects
            .filter(seller=seller, is_active=True, sku__iexact=text)
            .first()
        )

        if not product:
            product = (
                StockProduct.objects
                .filter(seller=seller, is_active=True, name__iexact=text)
                .first()
            )

    if not product and match_product:
        try:
            product = match_product(seller, text)
        except Exception:
            product = None

    return product



def _order_created_local_datetime(order):
    created_at = getattr(order, "created_at", None)

    if not created_at:
        return None

    try:
        if settings.USE_TZ:
            if timezone.is_naive(created_at):
                created_at = timezone.make_aware(
                    created_at,
                    timezone.get_current_timezone(),
                )
            return timezone.localtime(created_at)
        return created_at
    except Exception:
        return created_at


def _order_report_day_date(order):
    created_at = _order_created_local_datetime(order)

    if not created_at:
        return timezone.localdate()

    seller = getattr(order, "seller", None)

    if _uses_830_order_day(seller):
        return (created_at + ORDER_DAY_SHIFT).date()

    return created_at.date()


def _created_order_delivering_cutoff(order):
    """Return the next 2:30 PM delivery cutoff after the order was created.

    Customer display rule:
    - Created before / at 2:30 PM -> becomes Delivering at 2:30 PM the same day.
    - Created after 2:30 PM -> stays Created until 2:30 PM the next day.

    Examples:
    - 01-Jul 8:00 PM -> 02-Jul 2:30 PM
    - 02-Jul 9:00 AM -> 02-Jul 2:30 PM
    - 02-Jul 2:50 PM -> 03-Jul 2:30 PM
    """
    created_at = _order_created_local_datetime(order)

    if not created_at:
        return None

    try:
        cutoff_date = created_at.date()

        # If the order is created after 2:30 PM, the delivery cutoff is
        # 2:30 PM on the next day. At exactly 2:30 PM, it can show Delivering.
        if created_at.time() > DELIVERING_START_TIME:
            cutoff_date += timedelta(days=1)

        cutoff = datetime.combine(cutoff_date, DELIVERING_START_TIME)
        return _localize_datetime(cutoff)
    except Exception:
        return None


def _is_created_order_delivering_after_230(order):
    status = (getattr(order, "status", "") or "").strip().upper()

    if status != getattr(Order, "STATUS_CREATED", "CREATED"):
        return False

    if getattr(order, "delivery_shipper_id", None) or _order_has_assign_activity(order):
        return True

    cutoff = _created_order_delivering_cutoff(order)

    if not cutoff:
        return False

    try:
        return timezone.now() >= cutoff
    except Exception:
        return False


def _order_has_assign_activity(order):
    try:
        return order.activities.filter(
            action__in=[
                OrderActivity.ACTION_ASSIGN,
                OrderActivity.ACTION_OUT_FOR_DELIVERY,
            ]
        ).exists()
    except Exception:
        return False


def _last_inbound_activity_at(order):
    try:
        activity = (
            order.activities
            .filter(action=OrderActivity.ACTION_INBOUND)
            .order_by("-created_at")
            .first()
        )
        return getattr(activity, "created_at", None)
    except Exception:
        return None


def _is_unassigned_pending_over_72h(order):
    """Customer display rule.

    If an order has stayed without assignment/progress for 72 hours, show it as Pending.
    This applies to CREATED and INBOUND orders. The 72-hour pending check is counted
    from the real created/inbound time, not from the 2:30 PM display cutoff.
    """
    status = (getattr(order, "status", "") or "").strip().upper()

    pending_check_statuses = {
        getattr(Order, "STATUS_CREATED", "CREATED"),
        getattr(Order, "STATUS_INBOUND", "INBOUND"),
    }

    if status not in pending_check_statuses:
        return False

    if getattr(order, "delivery_shipper_id", None):
        return False

    if _order_has_assign_activity(order):
        return False

    if status == getattr(Order, "STATUS_INBOUND", "INBOUND"):
        start_at = _last_inbound_activity_at(order) or getattr(order, "created_at", None)
    else:
        start_at = getattr(order, "created_at", None)

    if not start_at:
        return False

    try:
        if timezone.is_naive(start_at) and settings.USE_TZ:
            start_at = timezone.make_aware(start_at, timezone.get_current_timezone())
        return timezone.now() - start_at >= timedelta(hours=72)
    except Exception:
        return False

def _get_computer_status(order):
    status = (getattr(order, "status", "") or "").strip().upper()

    if _is_customer_returned_order(order):
        return "returned", "Returned"

    if _is_customer_delivered_order(order):
        return "delivered", "Delivered"

    if status == getattr(Order, "STATUS_VOID", "VOID"):
        return "void", "Void"

    if status == getattr(
        Order,
        "STATUS_PROVINCE_ASSIGNED",
        "PROVINCE_ASSIGNED",
    ):
        return "province", "Province COD"

    if status in {
        getattr(Order, "STATUS_RETURN_ASSIGNED", "RETURN_ASSIGNED"),
        getattr(Order, "STATUS_RETURNING", "RETURNING"),
    }:
        return "returning", "Returning"

    if _is_unassigned_pending_over_72h(order):
        return "pending", "Pending"

    if status == getattr(Order, "STATUS_CREATED", "CREATED"):
        if _is_created_order_delivering_after_230(order):
            return "delivering", "Delivering"
        return "created", "Created"

    if status in {
        getattr(Order, "STATUS_INBOUND", "INBOUND"),
        getattr(Order, "STATUS_OUT_FOR_DELIVERY", "OUT_FOR_DELIVERY"),
        getattr(Order, "STATUS_PROCESSING", "PROCESSING"),
    }:
        return "delivering", "Delivering"

    return "pending", "Pending"


def _decorate_order(order):
    status_key, status_label = _get_computer_status(order)

    order.computer_status_key = status_key
    order.computer_status_label = status_label
    order.computer_cod = _order_cod(order)
    order.computer_delivery_fee = _order_delivery_fee(order)
    order.computer_additional_fee = _order_additional_fee(order)
    order.computer_province_fee = _order_province_fee(order)
    order.computer_total_fee = _order_total_fee(order)
    order.computer_fee = order.computer_total_fee
    order.computer_net = order.computer_cod - order.computer_total_fee

    # Safe display fields for templates. Do not use missing model fields
    # like order_code in template, because some Order rows only have tracking_no.
    order.computer_tracking_no = (
        getattr(order, "tracking_no", "")
        or getattr(order, "code", "")
        or str(order)
        or "-"
    )
    order.computer_receiver_name = (
        getattr(order, "receiver_name", "")
        or getattr(order, "customer_name", "")
        or "-"
    )
    order.computer_receiver_phone = (
        getattr(order, "receiver_phone", "")
        or getattr(order, "phone", "")
        or "-"
    )
    order.computer_seller_name = (
        getattr(order, "seller_name", "")
        or getattr(getattr(order, "seller", None), "name", "")
        or "-"
    )

    product = _find_order_inventory_product(order)
    raw_product = (getattr(order, "product_desc", "") or "").strip()

    if product:
        order.computer_product_sku = getattr(product, "sku", "") or "-"
        order.computer_product_name = getattr(product, "name", "") or raw_product or "-"
    else:
        # If no inventory product is found, keep the original product text visible.
        order.computer_product_sku = raw_product or "-"
        order.computer_product_name = raw_product or "-"

    reason = (getattr(order, "reason", "") or "").strip()
    if not reason and order.computer_status_key == "pending" and _is_unassigned_pending_over_72h(order):
        reason = "No assignment over 72h"
    if not reason and order.computer_status_key == "delivered" and _province_cod_customer_result(order) == "delivered":
        reason = "Province COD sent"
    order.computer_reason = reason or "-"

    return order


def _common_context(seller, period, start_date, end_date, period_label):
    uses_830_order_day = _uses_830_order_day(seller)

    return {
        "seller": seller,
        "period": period,
        "date_from": start_date.isoformat(),
        "date_to": end_date.isoformat(),
        "period_label": period_label,
        "uses_830_order_day": uses_830_order_day,
        "order_day_note": (
            "Starek order day: previous day 8:30 PM to selected date 8:29 PM."
            if uses_830_order_day
            else "Calendar day: 12:00 AM to 11:59 PM."
        ),
        "today_iso": _report_today(seller).isoformat(),
    }


# =========================================================
# COMPUTER DASHBOARD
# =========================================================

@login_required(login_url="portal:computer_login")
def computer_dashboard(request):
    seller = _get_logged_in_seller(request)

    if seller is None:
        return redirect("portal:computer_login")

    period, start_date, end_date, period_label = _get_report_range(request, seller)
    seller_orders = _base_seller_orders(seller)
    pending_statuses = _pending_statuses()

    # IMPORTANT:
    # Dashboard daily summary is a created-date cohort report.
    # Delivered / Returned / Pending / Province COD are counted from orders that were
    # CREATED on that row date. Do not group delivered or returned by done_at.
    # Only Starek uses the 8:30 PM custom order-day cutoff. Other shops use calendar day.
    created_start_dt, created_end_dt = _created_range_datetimes(
        start_date,
        end_date,
        seller,
    )

    created_orders = list(
        seller_orders.filter(
            created_at__gte=created_start_dt,
            created_at__lt=created_end_dt,
        ).order_by("created_at", "id")
    )

    rows_by_date = {}
    current_date = start_date

    while current_date <= end_date:
        rows_by_date[current_date] = {
            "date": current_date,
            "new_orders": 0,
            "delivered": 0,
            "pending": 0,
            "returned": 0,
            "province": 0,
            "province_cod": ZERO,
            "cod": ZERO,
            "fees": ZERO,
            "net": ZERO,
            "delivery_rate": 0,
            "return_rate": 0,
        }
        current_date += timedelta(days=1)

    total_pending = 0
    total_province = 0
    total_province_cod = ZERO
    total_delivered = 0
    total_returned = 0
    total_cod = ZERO
    total_fees = ZERO

    for order in created_orders:
        created_date = _get_created_date(order)

        if created_date not in rows_by_date:
            continue

        row = rows_by_date[created_date]
        status = (getattr(order, "status", "") or "").strip().upper()

        row["new_orders"] += 1

        if _is_customer_delivered_order(order):
            cod_value = _order_cod(order)
            fee_value = _order_total_fee(order)

            row["delivered"] += 1
            row["cod"] += cod_value
            row["fees"] += fee_value

            total_delivered += 1
            total_cod += cod_value
            total_fees += fee_value

        elif _is_customer_returned_order(order):
            row["returned"] += 1
            total_returned += 1

        elif status in pending_statuses:
            row["pending"] += 1
            total_pending += 1

        if _is_province_order(order):
            province_cod_value = _order_cod(order)
            row["province"] += 1
            row["province_cod"] += province_cod_value
            total_province += 1
            total_province_cod += province_cod_value

    daily_rows = []

    for row_date in sorted(rows_by_date.keys(), reverse=True):
        row = rows_by_date[row_date]
        row["delivery_rate"] = _safe_rate(
            row["delivered"],
            row["new_orders"],
        )
        row["return_rate"] = _safe_rate(
            row["returned"],
            row["new_orders"],
        )
        row["net"] = row["cod"] - row["fees"]
        daily_rows.append(row)

    total_orders = len(created_orders)
    net_balance = total_cod - total_fees

    # Average row for the daily summary table.
    # Count columns are average per day in the selected period.
    # Rate columns use the whole selected-period result.
    # Money columns stay as totals, not averages.
    report_days = max(len(daily_rows), 1)
    daily_average_row = {
        "label": "Average / Total",
        "new_orders": round(total_orders / report_days, 2),
        "delivered": round(total_delivered / report_days, 2),
        "pending": round(total_pending / report_days, 2),
        "returned": round(total_returned / report_days, 2),
        "province_cod": total_province_cod,
        "delivery_rate": _safe_rate(total_delivered, total_orders),
        "return_rate": _safe_rate(total_returned, total_orders),
        "cod": total_cod,
        "fees": total_fees,
        "net": net_balance,
    }

    # Recent orders and inventory snapshot are not shown on the dashboard now.
    recent_orders = []

    cancelled_orders = sum(
        1
        for order in created_orders
        if order.status == Order.STATUS_VOID
    )
    cancellation_rate = _safe_rate(cancelled_orders, total_orders)

    inventory_setting = get_seller_inventory_setting(seller)

    if (
        inventory_setting.stock_mode == InventorySellerSetting.NO_STOCK
        or not inventory_setting.show_stock_in_portal
    ):
        dashboard_stock_rows = []
        automatic_stock_status = "Hidden"
    else:
        dashboard_stock_rows = get_seller_current_stock(seller)
        automatic_stock_status = "Live"

    dashboard_inventory_products = len(dashboard_stock_rows)
    dashboard_available_stock = sum(
        int(row.get("available_qty", 0) or 0)
        for row in dashboard_stock_rows
    )
    dashboard_reserved_stock = sum(
        int(row.get("reserved_qty", 0) or 0)
        for row in dashboard_stock_rows
    )

    context = _common_context(
        seller,
        period,
        start_date,
        end_date,
        period_label,
    )

    context.update(
        {
            "total_orders": total_orders,
            "total_delivered": total_delivered,
            "total_pending": total_pending,
            "total_returned": total_returned,
            "total_province": total_province,
            "total_province_cod": total_province_cod,
            "total_cod": total_cod,
            "total_fees": total_fees,
            "net_balance": net_balance,
            "delivery_rate": _safe_rate(total_delivered, total_orders),
            "return_rate": _safe_rate(total_returned, total_orders),
            "daily_rows": daily_rows,
            "daily_average_row": daily_average_row,
            "recent_orders": recent_orders,
            "cancelled_orders": cancelled_orders,
            "cancellation_rate": cancellation_rate,
            "automatic_stock_status": automatic_stock_status,
            "dashboard_inventory_products": dashboard_inventory_products,
            "dashboard_available_stock": dashboard_available_stock,
            "dashboard_reserved_stock": dashboard_reserved_stock,
        }
    )

    return render(
        request,
        "customerportal/computer/dashboard.html",
        context,
    )


# =========================================================
# COMPUTER ORDERS
# =========================================================

@login_required(login_url="portal:computer_login")
def computer_orders(request):
    seller = _get_logged_in_seller(request)

    if seller is None:
        return redirect("portal:computer_login")

    period, start_date, end_date, period_label = _get_report_range(
        request,
        seller,
        default_period="last_3_days",
    )
    show_order_results = bool(request.GET)
    q = (request.GET.get("q") or "").strip()
    status_filter = (request.GET.get("status") or "ALL").strip().upper()

    created_start_dt, created_end_dt = _created_range_datetimes(
        start_date,
        end_date,
        seller,
    )

    period_orders_qs = (
        _base_seller_orders(seller)
        .select_related("seller")
        .filter(
            created_at__gte=created_start_dt,
            created_at__lt=created_end_dt,
        )
    )

    summary_orders = list(period_orders_qs.order_by("-created_at", "-id"))
    for order in summary_orders:
        _decorate_order(order)

    summary_total = len(summary_orders)
    summary_created = sum(1 for order in summary_orders if order.computer_status_key == "created")
    summary_delivering = sum(1 for order in summary_orders if order.computer_status_key == "delivering")
    summary_delivered = sum(1 for order in summary_orders if order.computer_status_key == "delivered")
    summary_pending = sum(1 for order in summary_orders if order.computer_status_key == "pending")
    summary_returned = sum(1 for order in summary_orders if order.computer_status_key == "returned")

    orders = []

    if show_order_results:
        orders_qs = period_orders_qs

        if q:
            orders_qs = orders_qs.filter(
                Q(tracking_no__icontains=q)
                | Q(seller_name__icontains=q)
                | Q(seller_order_code__icontains=q)
                | Q(receiver_name__icontains=q)
                | Q(receiver_phone__icontains=q)
                | Q(receiver_address__icontains=q)
                | Q(product_desc__icontains=q)
                | Q(reason__icontains=q)
            )

        orders = list(orders_qs.order_by("-created_at", "-id"))

        for order in orders:
            _decorate_order(order)

        if status_filter in {"CREATED", "DELIVERING", "PENDING", "DELIVERED", "RETURNED", "PROVINCE", "RETURNING", "VOID"}:
            status_key_map = {
                "CREATED": "created",
                "DELIVERING": "delivering",
                "PENDING": "pending",
                "DELIVERED": "delivered",
                "RETURNED": "returned",
                "PROVINCE": "province",
                "RETURNING": "returning",
                "VOID": "void",
            }
            wanted_key = status_key_map.get(status_filter)
            orders = [order for order in orders if order.computer_status_key == wanted_key]

    context = _common_context(
        seller,
        period,
        start_date,
        end_date,
        period_label,
    )

    context.update(
        {
            "q": q,
            "status_filter": status_filter,
            "orders": orders,
            "show_order_results": show_order_results,
            "summary_total": summary_total,
            "summary_created": summary_created,
            "summary_delivering": summary_delivering,
            "summary_delivered": summary_delivered,
            "summary_pending": summary_pending,
            "summary_returned": summary_returned,
        }
    )

    return render(
        request,
        "customerportal/computer/orders.html",
        context,
    )



# =========================================================
# COMPUTER DELIVERY REPORT HELPERS
# =========================================================

def _computer_report_datetime_from_get(value, fallback):
    raw = (value or "").strip()

    if not raw:
        dt = fallback
    else:
        try:
            dt = datetime.fromisoformat(raw)
        except Exception:
            dt = fallback

    if settings.USE_TZ and timezone.is_naive(dt):
        dt = timezone.make_aware(dt, timezone.get_current_timezone())

    return dt


def _computer_report_input_value(dt):
    if not dt:
        return ""

    try:
        dt = timezone.localtime(dt)
    except Exception:
        pass

    return dt.strftime("%Y-%m-%dT%H:%M")


def _computer_report_money(value):
    try:
        return Decimal(str(value or 0))
    except Exception:
        return ZERO


def _computer_report_apply_keyword(rows, keyword):
    keyword = (keyword or "").strip().lower()
    if not keyword:
        return rows

    filtered = []
    for o in rows:
        values = [
            getattr(o, "tracking_no", ""),
            getattr(o, "seller_order_code", ""),
            getattr(o, "seller_name", ""),
            getattr(o, "receiver_name", ""),
            getattr(o, "receiver_phone", ""),
            getattr(o, "receiver_address", ""),
            getattr(o, "product_desc", ""),
            getattr(o, "reason", ""),
            getattr(o, "report_shipper_name", ""),
        ]

        seller_obj = getattr(o, "seller", None)
        if seller_obj:
            values.append(getattr(seller_obj, "name", ""))
            values.append(getattr(seller_obj, "code", ""))

        haystack = " ".join(str(v or "") for v in values).lower()
        if keyword in haystack:
            filtered.append(o)

    return filtered


def _computer_report_apply_status_filter(rows, status_filter):
    status_filter = (status_filter or "").strip().upper()
    if not status_filter or status_filter == "ALL":
        return rows

    filtered = []
    for o in rows:
        row_type = classify_row(o)

        if status_filter == "DONE" and row_type == "done":
            filtered.append(o)
        elif status_filter == "PENDING" and row_type == "pending":
            filtered.append(o)
        elif status_filter == "DONE_RETURN" and row_type == "done_return":
            filtered.append(o)

    return filtered


def _computer_report_enrich_rows(rows):
    """Match internal report runtime display fields without saving DB."""
    order_ids = [o.id for o in rows if getattr(o, "id", None)]

    province_date_map = {}
    if order_ids:
        province_items = (
            ProvinceBatchItem.objects
            .select_related("batch")
            .filter(
                order_id__in=order_ids,
                batch__status=ProvinceBatch.STATUS_DONE,
                batch__assigned_at__isnull=False,
            )
            .order_by("-batch__assigned_at", "-id")
        )

        for item in province_items:
            if item.order_id not in province_date_map:
                province_date_map[item.order_id] = getattr(item.batch, "assigned_at", None)

    for order in rows:
        order.report_shipper_name = get_shipper_name(order)

        money = report_money(order)
        order.report_delivery_fee = _computer_report_money(money.get("delivery_fee", 0))
        order.report_additional_fee = _computer_report_money(money.get("additional_fee", 0))
        order.report_province_fee = _computer_report_money(money.get("province_fee", 0))
        order.report_total_fee = _computer_report_money(money.get("total_fee", 0))
        order.report_cod = _computer_report_money(money.get("cod", 0))

        province_date = province_date_map.get(getattr(order, "id", None))
        original_done_at = getattr(order, "done_at", None)
        final_delivery_date = original_done_at or province_date

        order.report_pickup_date = getattr(order, "created_at", None)
        order.report_delivery_date = final_delivery_date

        if not original_done_at and province_date:
            order.done_at = province_date

        if not getattr(order, "delivery_date", None):
            order.delivery_date = final_delivery_date

        if not getattr(order, "pickup_date", None):
            order.pickup_date = getattr(order, "created_at", None)

        row_type = classify_row(order)
        if row_type == "done":
            order.portal_report_status_key = "DONE"
            order.portal_report_status_label = "Done"
        elif row_type == "done_return":
            order.portal_report_status_key = "DONE_RETURN"
            order.portal_report_status_label = "Return Done"
        else:
            order.portal_report_status_key = "PENDING"
            order.portal_report_status_label = "Pending"

        product = _find_order_inventory_product(order)
        if product:
            order.portal_product_name = getattr(product, "name", "") or getattr(order, "product_desc", "") or "-"
        else:
            order.portal_product_name = getattr(order, "product_desc", "") or "-"

    return rows


def _computer_report_build_summary(rows, seller_count=0):
    total_done = 0
    total_pending = 0
    total_return = 0

    for order in rows:
        row_type = classify_row(order)
        if row_type == "done":
            total_done += 1
        elif row_type == "pending":
            total_pending += 1
        elif row_type == "done_return":
            total_return += 1

    total_cod, total_fee, total_pay = calc_totals(rows)

    return {
        "total_sent": len(rows),
        "total_done": total_done,
        "total_pending": total_pending,
        "total_return": total_return,
        "total_cod": total_cod,
        "total_fee": total_fee,
        "total_pay": total_pay,
        "total_selected_shops": seller_count,
    }


# =========================================================
# COMPUTER DELIVERY REPORT
# =========================================================

@login_required(login_url="portal:computer_login")
def computer_delivery_report(request):
    seller = _get_logged_in_seller(request)

    if seller is None:
        return redirect("portal:computer_login")

    now = timezone.localtime(timezone.now())
    today = now.date()

    default_delivery_from = datetime.combine(today, time.min)
    default_delivery_to = datetime.combine(today, time(23, 59))
    default_pending_from = datetime.combine(date(2020, 1, 1), time.min)
    default_pending_to = datetime.combine(today, time(23, 59))

    delivery_from = _computer_report_datetime_from_get(
        request.GET.get("delivery_date_from"),
        default_delivery_from,
    )
    delivery_to = _computer_report_datetime_from_get(
        request.GET.get("delivery_date_to"),
        default_delivery_to,
    )
    pending_from = _computer_report_datetime_from_get(
        request.GET.get("pending_date_from"),
        default_pending_from,
    )
    pending_to = _computer_report_datetime_from_get(
        request.GET.get("pending_date_to"),
        default_pending_to,
    )

    keyword = (request.GET.get("search") or "").strip()
    status_filter = (request.GET.get("status_filter") or "ALL").strip().upper()
    action = (request.GET.get("action") or "").strip()
    show_results = action in {"show", "export"}

    cleaned_data = {
        "seller": seller,
        "delivery_date_from": delivery_from,
        "delivery_date_to": delivery_to,
        "pending_date_from": pending_from,
        "pending_date_to": pending_to,
        "search": keyword,
        "keyword": keyword,
        "status_filter": status_filter if status_filter != "ALL" else "",
    }

    mode = "EMPTY"
    rows = []
    grouped = {}
    seller_summaries = []
    top_summary = {
        "total_sent": 0,
        "total_done": 0,
        "total_pending": 0,
        "total_return": 0,
        "total_cod": ZERO,
        "total_fee": ZERO,
        "total_pay": ZERO,
        "total_selected_shops": 0,
    }

    if show_results:
        done_rows = get_done_queryset(Order, cleaned_data)
        pending_rows = get_pending_queryset(Order, cleaned_data)

        mode = "DONE_PENDING"
        rows = list(done_rows) + list(pending_rows)
        rows = _computer_report_enrich_rows(rows)
        rows = _computer_report_apply_keyword(rows, keyword)
        rows = _computer_report_apply_status_filter(rows, status_filter)

        grouped = group_by_seller(rows)

        filtered_rows = []
        for seller_rows in grouped.values():
            filtered_rows.extend(seller_rows)

        top_summary = _computer_report_build_summary(filtered_rows, seller_count=len(grouped))

        if action == "export" and mode != "EMPTY":
            return export_delivery_report_xlsx(
                grouped,
                "Delivery Report",
                classify_row,
                calc_totals,
                delivery_from,
                delivery_to,
                filename_prefix=f"{seller.name}_Delivery_Report",
            )

        for seller_key_name, seller_rows in grouped.items():
            total_cod, total_fee, pay = calc_totals(seller_rows)
            seller_summaries.append({
                "seller_key": seller_key_name,
                "rows": seller_rows,
                "total_cod": total_cod,
                "total_fee": total_fee,
                "pay": pay,
                "total_sent": len(seller_rows),
                "total_done": len([o for o in seller_rows if classify_row(o) == "done"]),
                "total_pending": len([o for o in seller_rows if classify_row(o) == "pending"]),
                "total_return": len([o for o in seller_rows if classify_row(o) == "done_return"]),
            })

    context = _common_context(
        seller,
        "custom",
        delivery_from.date(),
        delivery_to.date(),
        f"{delivery_from.date():%d %b %Y} - {delivery_to.date():%d %b %Y}",
    )

    context.update({
        "mode": mode,
        "show_results": show_results,
        "seller_summaries": seller_summaries,
        "delivery_from": delivery_from,
        "delivery_to": delivery_to,
        "pending_from": pending_from,
        "pending_to": pending_to,
        "delivery_from_value": _computer_report_input_value(delivery_from),
        "delivery_to_value": _computer_report_input_value(delivery_to),
        "pending_from_value": _computer_report_input_value(pending_from),
        "pending_to_value": _computer_report_input_value(pending_to),
        "search_value": keyword,
        "status_filter": status_filter,
        "top_summary": top_summary,
    })

    return render(
        request,
        "customerportal/computer/delivery_report.html",
        context,
    )


# =========================================================
# COMPUTER COD REPORT
# =========================================================

def _computer_active_province_carriers():
    return Shipper.objects.filter(
        is_active=True,
        shipper_type=Shipper.TYPE_PROVINCE,
    ).order_by("name")




def _pc_local_date(value):
    """Return local date from datetime/date safely. Runtime helper only."""
    if not value:
        return None

    try:
        if isinstance(value, datetime):
            if settings.USE_TZ and timezone.is_aware(value):
                value = timezone.localtime(value)
            return value.date()

        if isinstance(value, date):
            return value
    except Exception:
        return None

    return None


def _pc_date_in_range(value, start_date, end_date):
    value_date = _pc_local_date(value)

    if not value_date:
        return None

    if start_date and value_date < start_date:
        return None

    if end_date and value_date > end_date:
        return None

    return value_date


def _pc_empty_daily_summary_row(row_date):
    return {
        "date": row_date,
        "sent_orders": 0,
        "sent_cod": ZERO,
        "received_orders": 0,
        "settled_orders": 0,
        "returned_orders": 0,
        "done_orders": 0,
        "done_rate": 0,
    }


def _build_province_cod_daily_rows(items, start_date, end_date):
    """Build a sent-date cohort summary.

    Every item stays on the date it was originally sent. Its current COD
    result is counted on that same sent-date row, even when it was received,
    paid, settled, or returned on a later date.
    """
    if not start_date or not end_date:
        return []

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    rows_by_date = {}
    current = start_date
    max_days = 370
    days = 0

    while current <= end_date and days < max_days:
        rows_by_date[current] = _pc_empty_daily_summary_row(current)
        current += timedelta(days=1)
        days += 1

    for item in items:
        original_cod = money(getattr(item, "original_cod", ZERO))

        sent_source = (
            getattr(item, "activity_date", None)
            or getattr(item, "sent_at", None)
            or getattr(getattr(item, "batch", None), "sent_at", None)
            or getattr(getattr(item, "batch", None), "assigned_at", None)
            or getattr(getattr(item, "batch", None), "created_at", None)
        )
        sent_date = _pc_date_in_range(
            sent_source,
            start_date,
            end_date,
        )

        if sent_date not in rows_by_date:
            continue

        row = rows_by_date[sent_date]
        row["sent_orders"] += 1
        row["sent_cod"] += original_cod

        cod_status = (
            getattr(item, "cod_status", "") or ""
        ).strip().upper()

        # PAID has already passed RECEIVED, so it is included in received.
        if cod_status in {
            ProvinceCODItem.STATUS_RECEIVED,
            ProvinceCODItem.STATUS_PAID,
        }:
            row["received_orders"] += 1

        if getattr(item, "seller_settled", False):
            row["settled_orders"] += 1

        if cod_status == ProvinceCODItem.STATUS_RETURNED:
            row["returned_orders"] += 1

        # Count each item once only.
        if cod_status in {
            ProvinceCODItem.STATUS_RECEIVED,
            ProvinceCODItem.STATUS_PAID,
            ProvinceCODItem.STATUS_RETURNED,
        }:
            row["done_orders"] += 1

    for row in rows_by_date.values():
        row["done_rate"] = _safe_rate(
            row["done_orders"],
            row["sent_orders"],
        )

    return [
        rows_by_date[row_date]
        for row_date in sorted(rows_by_date.keys(), reverse=True)
    ]


def _computer_province_cod_export_xlsx(rows, seller):
    wb = Workbook()
    ws = wb.active
    ws.title = "Province COD"

    headers = [
        "ID",
        "Date",
        "Batch",
        "Tracking",
        "Seller",
        "Carrier",
        "Receiver",
        "Phone",
        "Address",
        "Original COD",
        "Province Fee",
        "Status",
        "Carrier Fee",
        "Net COD",
        "Reference",
        "Settled",
        "Updated",
        "Reason / Note",
    ]
    ws.append(headers)

    for item in rows:
        order = getattr(item, "order", None)
        batch = getattr(item, "batch", None)

        reason_note = (
            getattr(item, "return_reason", "")
            or getattr(item, "note", "")
            or getattr(order, "reason", "")
            or "-"
        )

        ws.append([
            item.id,
            item.activity_date.strftime("%Y-%m-%d %H:%M") if getattr(item, "activity_date", None) else "",
            f"PVCOD-{getattr(item, 'batch_id', '')}",
            getattr(order, "tracking_no", "") if order else "",
            getattr(getattr(order, "seller", None), "name", "") if order else "",
            getattr(getattr(batch, "shipper", None), "name", "") if batch else "",
            getattr(order, "receiver_name", "") if order else "",
            getattr(order, "receiver_phone", "") if order else "",
            getattr(order, "receiver_address", "") if order else "",
            float(money(getattr(item, "original_cod", ZERO))),
            float(money(getattr(item, "province_fee", ZERO))),
            getattr(item, "display_status", "") or "PENDING",
            float(money(getattr(item, "carrier_fee", ZERO))),
            float(money(getattr(item, "net_cod", ZERO))),
            getattr(item, "carrier_reference", "") or "-",
            "YES" if getattr(item, "seller_settled", False) else "NO",
            item.updated_at.strftime("%Y-%m-%d %H:%M") if getattr(item, "updated_at", None) else "",
            reason_note,
        ])

    for col in ws.columns:
        max_len = 10
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"province-cod-{getattr(seller, 'name', 'seller')}.xlsx"
    filename = re.sub(r"[^A-Za-z0-9_.-]+", "_", filename)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required(login_url="portal:computer_login")
def computer_cod_report(request):
    seller = _get_logged_in_seller(request)

    if seller is None:
        return redirect("portal:computer_login")

    today = timezone.localdate()
    default_from = today.isoformat()
    default_to = today.isoformat()

    date_from = (request.GET.get("date_from") or default_from).strip()
    date_to = (request.GET.get("date_to") or default_to).strip()
    status = (request.GET.get("status") or "").strip().upper()
    settlement = (request.GET.get("settlement") or "").strip().upper()
    shipper_id = (request.GET.get("shipper") or "").strip()
    q = (request.GET.get("q") or "").strip()

    sort = (request.GET.get("sort") or "sent_date").strip().lower()
    direction = (request.GET.get("direction") or "desc").strip().lower()
    if direction not in {"asc", "desc"}:
        direction = "desc"

    sort_map = {
        "id": "id",
        "sent_date": "activity_date",
        "batch": "batch_id",
        "tracking": "order__tracking_no",
        "carrier": "batch__shipper__name",
        "receiver": "order__receiver_name",
        "phone": "order__receiver_phone",
        "original_cod": "original_cod",
        "province_fee": "province_fee",
        "status": "cod_status",
        "carrier_fee": "carrier_fee",
        "net_cod": "net_cod",
        "reference": "carrier_reference",
        "settled": "seller_settled",
        "updated": "updated_at",
    }

    if sort not in sort_map:
        sort = "sent_date"

    base_rows = (
        ProvinceCODItem.objects
        .select_related(
            "batch",
            "batch__shipper",
            "order",
            "order__seller",
            "received_confirmed_by",
            "paid_confirmed_by",
            "returned_confirmed_by",
            "seller_settled_by",
        )
        .exclude(batch__status=ProvinceCODBatch.STATUS_CANCELLED)
        .filter(order__seller=seller)
        .annotate(
            activity_date=Coalesce("sent_at", "batch__created_at"),
        )
    )

    if status == "PENDING":
        base_rows = base_rows.filter(cod_status="")
    elif status:
        base_rows = base_rows.filter(cod_status=status)

    if settlement == "SETTLED":
        base_rows = base_rows.filter(seller_settled=True)
    elif settlement == "UNSETTLED":
        base_rows = base_rows.filter(seller_settled=False)

    if shipper_id.isdigit():
        base_rows = base_rows.filter(batch__shipper_id=int(shipper_id))

    if q:
        base_rows = base_rows.filter(
            Q(order__tracking_no__icontains=q)
            | Q(order__receiver_name__icontains=q)
            | Q(order__receiver_phone__icontains=q)
            | Q(order__receiver_address__icontains=q)
            | Q(batch__shipper__name__icontains=q)
            | Q(carrier_reference__icontains=q)
            | Q(received_person__icontains=q)
            | Q(return_reason__icontains=q)
            | Q(note__icontains=q)
        )

    summary_start_date = _parse_date(date_from) or today
    summary_end_date = _parse_date(date_to) or today

    # The COD summary is grouped by each item's original sent date.
    # Later received, paid, settled, or returned results stay on that row.
    summary_items = list(base_rows.order_by("activity_date", "id"))

    rows = base_rows

    if date_from:
        rows = rows.filter(activity_date__date__gte=date_from)

    if date_to:
        rows = rows.filter(activity_date__date__lte=date_to)

    order_field = sort_map[sort]
    if direction == "desc":
        order_field = f"-{order_field}"

    rows = list(rows.order_by(order_field, "-id"))

    for item in rows:
        item.suggested_fee_display = item.suggested_carrier_fee()
        item.display_status = item.cod_status or "PENDING"

    for item in summary_items:
        item.suggested_fee_display = item.suggested_carrier_fee()
        item.display_status = item.cod_status or "PENDING"

    cod_daily_rows = _build_province_cod_daily_rows(
        summary_items,
        summary_start_date,
        summary_end_date,
    )

    # Average daily Done Rate.
    # Only dates with at least one sent order are counted.
    active_sent_days = [
        row
        for row in cod_daily_rows
        if row.get("sent_orders", 0) > 0
    ]

    average_done_rate = (
        round(
            sum(
                float(row.get("done_rate", 0) or 0)
                for row in active_sent_days
            ) / len(active_sent_days),
            2,
        )
        if active_sent_days
        else 0
    )

    paid_rows = [
        item
        for item in rows
        if item.cod_status == ProvinceCODItem.STATUS_PAID
    ]
    settled_rows = [
        item
        for item in rows
        if getattr(item, "seller_settled", False)
    ]

    summary_count = len(rows)
    summary_paid = len(paid_rows)
    summary_settled = len(settled_rows)

    summary = {
        "count": summary_count,
        "original_cod": sum((money(item.original_cod) for item in rows), ZERO),
        "carrier_fee": sum((money(item.carrier_fee) for item in rows), ZERO),
        "province_fee": sum((money(item.province_fee) for item in rows), ZERO),
        "net_cod": sum((money(item.net_cod) for item in rows), ZERO),
        "done_cod": sum((money(item.original_cod) for item in settled_rows), ZERO),
        "done_net_cod": sum((money(item.net_cod) for item in settled_rows), ZERO),
        "done_rate": average_done_rate,
        "pending": sum(1 for item in rows if not item.cod_status),
        "sent": sum(1 for item in rows if item.cod_status == ProvinceCODItem.STATUS_SENT),
        "received": sum(1 for item in rows if item.cod_status == ProvinceCODItem.STATUS_RECEIVED),
        "paid": summary_paid,
        "returned": sum(1 for item in rows if item.cod_status == ProvinceCODItem.STATUS_RETURNED),
        "settled": sum(1 for item in rows if item.seller_settled),
        "sent_cod": sum((money(item.original_cod) for item in rows), ZERO),
        "received_cod": sum(
            (money(item.original_cod) for item in rows if getattr(item, "received_at", None)),
            ZERO,
        ),
        "settled_cod": sum(
            (money(item.net_cod) for item in rows if getattr(item, "seller_settled", False)),
            ZERO,
        ),
        "returned_cod": sum(
            (money(item.original_cod) for item in rows if item.cod_status == ProvinceCODItem.STATUS_RETURNED),
            ZERO,
        ),
    }

    if (request.GET.get("action") or "").strip() == "export":
        return _computer_province_cod_export_xlsx(rows, seller)

    sort_urls = {}
    for key in sort_map:
        params = request.GET.copy()
        next_direction = "asc"
        if sort == key and direction == "asc":
            next_direction = "desc"
        params["sort"] = key
        params["direction"] = next_direction
        sort_urls[key] = f"?{params.urlencode()}"

    context = _common_context(
        seller,
        "custom",
        _parse_date(date_from) or today,
        _parse_date(date_to) or today,
        (
            f"{date_from} - {date_to}"
            if date_from != date_to
            else date_from
        ),
    )

    context.update(
        {
            "rows": rows,
            "summary": summary,
            "cod_daily_rows": cod_daily_rows,
            "date_from": date_from,
            "date_to": date_to,
            "status": status,
            "settlement": settlement,
            "shipper_id": shipper_id,
            "q": q,
            "sort": sort,
            "direction": direction,
            "sort_urls": sort_urls,
            "current_query": request.GET.urlencode(),
            "shippers": _computer_active_province_carriers(),
        }
    )

    return render(
        request,
        "customerportal/computer/cod_report.html",
        context,
    )

# =========================================================
# COMPUTER INVENTORY
# =========================================================

@login_required(login_url="portal:computer_login")
def computer_inventory(request):
    seller = _get_logged_in_seller(request)

    if seller is None:
        return redirect("portal:computer_login")

    setting = get_seller_inventory_setting(seller)

    if (
        setting.stock_mode == InventorySellerSetting.NO_STOCK
        or not setting.show_stock_in_portal
    ):
        stock_rows = []
        inventory_enabled = False
    else:
        stock_rows = get_seller_current_stock(seller)
        inventory_enabled = True

    q = (request.GET.get("q") or "").strip().lower()

    if q:
        stock_rows = [
            row
            for row in stock_rows
            if q in str(row.get("sku", "") or "").lower()
            or q in str(row.get("name", "") or "").lower()
            or q in str(row.get("product_type", "") or "").lower()
            or q in str(row.get("location", "") or "").lower()
        ]

    total_products = len(stock_rows)
    total_current = sum(
        int(row.get("current_qty", 0) or 0)
        for row in stock_rows
    )
    total_reserved = sum(
        int(row.get("reserved_qty", 0) or 0)
        for row in stock_rows
    )
    total_available = sum(
        int(row.get("available_qty", 0) or 0)
        for row in stock_rows
    )

    context = {
        "seller": seller,
        "setting": setting,
        "inventory_enabled": inventory_enabled,
        "stock_rows": stock_rows,
        "q": q,
        "total_products": total_products,
        "total_current": total_current,
        "total_reserved": total_reserved,
        "total_available": total_available,
    }

    return render(
        request,
        "customerportal/computer/inventory.html",
        context,
    )

# =========================================================
# COMPUTER PORTAL - CUSTOMER ORDER UPLOAD
# =========================================================

UPLOAD_HEADERS = [
    "Seller Order Code",
    "Seller",
    "Product Description",
    "Quantity",
    "COD",
    "Receiver Name",
    "Phone",
    "Address",
    "Remark",
    "SKU",
]

UPLOAD_SHEET_NAME = "DS_UPLOAD"
UPLOAD_TEMPLATE_KEY = "DS_EXPRESS_UPLOAD_TEMPLATE_V1"
UPLOAD_MAX_SIZE = 5 * 1024 * 1024
UPLOAD_MAX_ROWS = 500


def _upload_cell(value):
    return "" if value is None else str(value).strip()


def _upload_decimal(value):
    try:
        raw = str(value or "").strip()
        if raw == "":
            return Decimal("0.00")
        return Decimal(raw).quantize(Decimal("0.00"))
    except Exception:
        return Decimal("0.00")


def _upload_int(value, default=1):
    try:
        return max(int(float(str(value).strip())), 1)
    except Exception:
        return default


def _normalize_phone(value):
    raw = str(value or "").strip()
    digits = re.sub(r"\D+", "", raw)
    return digits


def _cell_has_danger(cell):
    value = cell.value

    if cell.data_type == "f":
        return True

    if isinstance(value, str):
        text = value.strip()

        if text.startswith("=") or text.startswith("@"):
            return True

        bad_words = ["<script", "</script", "drop table", "delete from", "insert into", "update "]
        low = text.lower()
        return any(x in low for x in bad_words)

    return False


def _find_upload_product_by_sku(seller, sku):
    """
    Product Description is only customer text. Inventory matching uses SKU only.
    This prevents wrong product names from changing stock matching.
    """
    try:
        from inventory.models import StockProduct
    except Exception:
        return None

    sku = (sku or "").strip()
    if not sku:
        return None

    return (
        StockProduct.objects
        .filter(seller=seller, is_active=True, sku__iexact=sku)
        .first()
    )


def _decorate_upload_rows_product_display(seller, rows):
    """
    Display helper for upload detail pages.
    Shows the readable inventory product name from SKU, while keeping the
    customer's uploaded Product Description visible for verification.
    """
    try:
        from inventory.models import StockProduct
        from inventory.services import match_product
    except Exception:
        StockProduct = None
        match_product = None

    for row in rows:
        input_description = (getattr(row, "product_name_input", "") or "").strip()
        input_sku = (getattr(row, "sku_input", "") or "").strip()
        matched_name = (getattr(row, "matched_product_name", "") or "").strip()
        matched_sku = (getattr(row, "matched_sku", "") or "").strip()
        old_product_text = (getattr(row, "product_desc", "") or "").strip()

        product = None

        if StockProduct:
            # New uploads use sku_input/matched_sku. Older uploads may have stored
            # only the SKU in product_desc, so keep that fallback.
            for val in [matched_sku, input_sku, old_product_text]:
                if not val:
                    continue
                product = (
                    StockProduct.objects
                    .filter(seller=seller, is_active=True, sku__iexact=val)
                    .first()
                )
                if product:
                    break

        if not product and match_product:
            # Fallback for very old rows where product_desc stored a product name.
            for val in [matched_name, old_product_text]:
                if not val:
                    continue
                product = match_product(seller, val)
                if product:
                    break

        if product:
            row.display_inventory_product = product.name or "-"
            row.display_inventory_sku = product.sku or "-"
        else:
            row.display_inventory_product = matched_name or "-"
            row.display_inventory_sku = matched_sku or input_sku or "-"

        # Product Description is the customer's free text. For older rows that only
        # stored SKU in product_desc, avoid showing SKU as description when possible.
        if input_description:
            row.display_product_description = input_description
        elif product and old_product_text and old_product_text.upper() == (getattr(product, "sku", "") or "").upper():
            row.display_product_description = "-"
        else:
            row.display_product_description = old_product_text or "-"

        row.display_uploaded_sku = input_sku or matched_sku or "-"


def _validate_upload_product_stock(seller, product, qty, strict_stock_used):
    try:
        from inventory.models import InventorySellerSetting
        from inventory.services import current_available_qty, get_seller_inventory_setting
    except Exception:
        return True, ""

    setting = get_seller_inventory_setting(seller)

    if setting.stock_mode == InventorySellerSetting.NO_STOCK:
        return True, ""

    if not product:
        return False, "SKU not recognized in inventory. Please check SKU."

    if setting.stock_mode == InventorySellerSetting.STRICT:
        available_qty = current_available_qty(product)
        already_used_qty = int(strict_stock_used.get(product.id, 0) or 0)
        remaining_qty = available_qty - already_used_qty

        if remaining_qty < qty:
            return False, f"Strict stock shop: not enough stock for {product.name}. Need {qty}, available {remaining_qty}."

        strict_stock_used[product.id] = already_used_qty + qty

    return True, ""


def _recalc_upload_batch(batch):
    rows = batch.rows.all()
    batch.total_rows = rows.count()
    batch.valid_rows = rows.filter(status=SellerUploadRow.STATUS_VALID).count()
    batch.error_rows = rows.filter(status=SellerUploadRow.STATUS_ERROR).count()
    batch.duplicate_rows = rows.filter(status=SellerUploadRow.STATUS_DUPLICATE).count()
    batch.save(update_fields=["total_rows", "valid_rows", "error_rows", "duplicate_rows", "updated_at"])


def _send_customer_upload_staff_telegram(batch):
    bot_token = getattr(settings, "TELEGRAM_DS_TEAM_BOT_TOKEN", "")
    chat_id = getattr(settings, "TELEGRAM_DS_TEAM_CHAT_ID", "")
    if not bot_token or not chat_id:
        return
    text = (
        "📥 New Customer Upload Waiting Approval\n\n"
        f"Shop: {batch.seller.name}\n"
        f"Batch: {batch.code}\n"
        f"Rows: {batch.total_rows}\n"
        f"Valid: {batch.valid_rows}\n"
        f"File: {batch.original_filename or '-'}\n"
        f"Upload Remark: {batch.upload_remark or '-'}"
    )
    try:
        requests.post(f"https://api.telegram.org/bot{bot_token}/sendMessage", data={"chat_id": chat_id, "text": text}, timeout=10)
    except Exception:
        pass


@login_required(login_url="portal:computer_login")
def download_customer_upload_sample(request):
    seller = _get_logged_in_seller(request)
    if seller is None:
        return redirect("portal:computer_login")

    wb = Workbook()
    ws = wb.active
    ws.title = UPLOAD_SHEET_NAME
    ws["A1"] = UPLOAD_TEMPLATE_KEY
    ws["A2"] = "Do not edit row 1. Fill data from row 4 only. Product Description is free text. SKU controls inventory matching."

    for col, header in enumerate(UPLOAD_HEADERS, start=1):
        ws.cell(row=3, column=col, value=header)

    ws.append([
        "SO-001",
        seller.name,
        "Customer product text / seller product name",
        1,
        10,
        "Customer Name",
        "012345678",
        "Phnom Penh address...",
        "",
        "260001-ITE-002",
    ])

    widths = [18, 20, 34, 12, 12, 22, 16, 34, 24, 18]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + col)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="ds_customer_upload_sample.xlsx"'
    return response


@login_required(login_url="portal:computer_login")
def computer_upload_orders(request):
    seller = _get_logged_in_seller(request)
    if seller is None:
        return redirect("portal:computer_login")

    if request.method == "POST":
        upload_remark = (request.POST.get("upload_remark") or "").strip()
        upload_file = request.FILES.get("file")
        if not upload_file:
            messages.error(request, "Please choose an Excel file.")
            return redirect("portal:computer_upload_orders")

        original_filename = upload_file.name or ""
        ext = os.path.splitext(original_filename)[1].lower()

        if ext != ".xlsx":
            batch = SellerUploadBatch.objects.create(seller=seller, uploaded_by=request.user, original_filename=original_filename, upload_remark=upload_remark, status=SellerUploadBatch.STATUS_REJECTED, reject_reason="Invalid file type. Please upload .xlsx official sample only.", rejected_by=request.user, rejected_at=timezone.now())
            messages.error(request, "Rejected: only .xlsx official sample is allowed.")
            return redirect("portal:computer_upload_order_detail", batch_id=batch.id)

        if upload_file.size > UPLOAD_MAX_SIZE:
            batch = SellerUploadBatch.objects.create(seller=seller, uploaded_by=request.user, original_filename=original_filename, upload_remark=upload_remark, status=SellerUploadBatch.STATUS_REJECTED, reject_reason="File too large. Maximum size is 5MB.", rejected_by=request.user, rejected_at=timezone.now())
            messages.error(request, "Rejected: file too large.")
            return redirect("portal:computer_upload_order_detail", batch_id=batch.id)

        batch = SellerUploadBatch.objects.create(seller=seller, uploaded_by=request.user, file=upload_file, original_filename=original_filename, upload_remark=upload_remark, status=SellerUploadBatch.STATUS_PENDING)

        try:
            batch.file.open("rb")
            wb = load_workbook(filename=batch.file, data_only=False)
        except Exception:
            batch.status = SellerUploadBatch.STATUS_REJECTED
            batch.reject_reason = "Cannot read Excel file."
            batch.rejected_by = request.user
            batch.rejected_at = timezone.now()
            batch.save()
            messages.error(request, "Rejected: cannot read Excel file.")
            return redirect("portal:computer_upload_order_detail", batch_id=batch.id)

        if UPLOAD_SHEET_NAME not in wb.sheetnames:
            batch.status = SellerUploadBatch.STATUS_REJECTED
            batch.reject_reason = "Invalid template. Sheet name must be DS_UPLOAD."
            batch.rejected_by = request.user
            batch.rejected_at = timezone.now()
            batch.save()
            messages.error(request, "Rejected: invalid DS template.")
            return redirect("portal:computer_upload_order_detail", batch_id=batch.id)

        ws = wb[UPLOAD_SHEET_NAME]
        if _upload_cell(ws["A1"].value) != UPLOAD_TEMPLATE_KEY:
            batch.status = SellerUploadBatch.STATUS_REJECTED
            batch.reject_reason = "Invalid template key. Please download the official DS sample."
            batch.rejected_by = request.user
            batch.rejected_at = timezone.now()
            batch.save()
            messages.error(request, "Rejected: please use DS official sample.")
            return redirect("portal:computer_upload_order_detail", batch_id=batch.id)

        headers = [_upload_cell(ws.cell(row=3, column=i).value) for i in range(1, len(UPLOAD_HEADERS) + 1)]
        if headers != UPLOAD_HEADERS:
            batch.status = SellerUploadBatch.STATUS_REJECTED
            batch.reject_reason = "Invalid columns. Please use the official DS sample without editing columns."
            batch.rejected_by = request.user
            batch.rejected_at = timezone.now()
            batch.save()
            messages.error(request, "Rejected: invalid columns.")
            return redirect("portal:computer_upload_order_detail", batch_id=batch.id)

        seen_codes = set()
        strict_stock_used = {}
        created_rows = 0

        for row_idx in range(4, ws.max_row + 1):
            if created_rows >= UPLOAD_MAX_ROWS:
                SellerUploadRow.objects.create(batch=batch, row_number=row_idx, status=SellerUploadRow.STATUS_ERROR, error_message=f"Upload limit is {UPLOAD_MAX_ROWS} rows.")
                continue

            row_values = [ws.cell(row=row_idx, column=col).value for col in range(1, len(UPLOAD_HEADERS) + 1)]
            if all(_upload_cell(v) == "" for v in row_values):
                continue

            created_rows += 1
            errors = []
            for col in range(1, len(UPLOAD_HEADERS) + 1):
                if _cell_has_danger(ws.cell(row=row_idx, column=col)):
                    errors.append("Dangerous formula/text detected.")
                    break

            seller_order_code = _upload_cell(row_values[0])
            seller_name = _upload_cell(row_values[1]) or (seller.name or "")
            product_description = _upload_cell(row_values[2])
            quantity = _upload_int(row_values[3], 1)
            cod = _upload_decimal(row_values[4])
            price = cod
            receiver_name = _upload_cell(row_values[5])
            phone = _normalize_phone(row_values[6])
            address = _upload_cell(row_values[7])
            remark = _upload_cell(row_values[8])
            sku_input = _upload_cell(row_values[9])

            if not seller_order_code:
                errors.append("Seller Order Code is required.")
            if not product_description:
                errors.append("Product Description is required.")
            if not sku_input:
                errors.append("SKU is required because inventory matching uses SKU.")
            if not receiver_name:
                errors.append("Receiver Name is required.")
            if not phone:
                errors.append("Phone is required.")
            elif len(phone) < 8 or len(phone) > 15:
                errors.append("Phone number is invalid.")
            if not address:
                errors.append("Address is required.")
            if cod < Decimal("0.00"):
                errors.append("COD cannot be negative.")

            normalized_code = seller_order_code.lower()
            row_status = SellerUploadRow.STATUS_VALID
            if normalized_code in seen_codes:
                row_status = SellerUploadRow.STATUS_DUPLICATE
                errors.append("Duplicate Seller Order Code in this upload.")
            elif seller_order_code and Order.objects.filter(seller=seller, seller_order_code__iexact=seller_order_code, is_deleted=False).exists():
                row_status = SellerUploadRow.STATUS_DUPLICATE
                errors.append("Seller Order Code already exists in system.")
            if seller_order_code:
                seen_codes.add(normalized_code)

            matched_product = _find_upload_product_by_sku(seller, sku_input)
            matched_product_name = getattr(matched_product, "name", "") or ""
            matched_sku = getattr(matched_product, "sku", "") or ""

            ok_stock, stock_error = _validate_upload_product_stock(seller=seller, product=matched_product, qty=quantity, strict_stock_used=strict_stock_used)
            if not ok_stock:
                errors.append(stock_error)

            if errors and row_status != SellerUploadRow.STATUS_DUPLICATE:
                row_status = SellerUploadRow.STATUS_ERROR

            SellerUploadRow.objects.create(
                batch=batch,
                row_number=row_idx,
                seller_order_code=seller_order_code,
                seller_name=seller_name,
                # Keep using existing field name for DB compatibility. It now stores Product Description.
                product_name_input=product_description,
                sku_input=sku_input,
                matched_product_name=matched_product_name,
                matched_sku=matched_sku,
                receiver_name=receiver_name,
                receiver_phone=phone,
                receiver_address=address,
                product_desc=product_description,
                quantity=quantity,
                cod=cod,
                price=price,
                remark=remark,
                status=row_status,
                error_message=" | ".join(errors),
            )

        _recalc_upload_batch(batch)

        if batch.total_rows <= 0:
            batch.status = SellerUploadBatch.STATUS_REJECTED
            batch.reject_reason = "No data rows found. Please fill data from row 4."
            batch.rejected_by = request.user
            batch.rejected_at = timezone.now()
            batch.save()
            messages.error(request, "Rejected: no data rows found.")
            return redirect("portal:computer_upload_order_detail", batch_id=batch.id)

        if batch.error_rows > 0 or batch.duplicate_rows > 0:
            batch.status = SellerUploadBatch.STATUS_REJECTED
            batch.reject_reason = f"Auto rejected because upload has {batch.error_rows} error row(s) and {batch.duplicate_rows} duplicate row(s). Please fix the file and upload again."
            batch.rejected_by = request.user
            batch.rejected_at = timezone.now()
            batch.save(update_fields=["status", "reject_reason", "rejected_by", "rejected_at", "updated_at"])
            messages.error(request, f"Upload auto rejected. Errors: {batch.error_rows}, Duplicate: {batch.duplicate_rows}. Please fix and upload again.")
            return redirect("portal:computer_upload_order_detail", batch_id=batch.id)

        _send_customer_upload_staff_telegram(batch)
        messages.success(request, f"Upload saved. Rows: {batch.total_rows}, Valid: {batch.valid_rows}. Waiting staff approval.")
        return redirect("portal:computer_upload_order_detail", batch_id=batch.id)

    today = timezone.localdate()
    default_from = today - timedelta(days=2)
    history_q = (request.GET.get("q") or "").strip()
    history_status = (request.GET.get("status") or "ALL").strip().upper()
    history_period = (request.GET.get("period") or "last_3_days").strip()
    date_from_obj = _parse_date(request.GET.get("from"))
    date_to_obj = _parse_date(request.GET.get("to"))
    batches_qs = SellerUploadBatch.objects.filter(seller=seller).order_by("-id")

    if history_period != "all":
        if not date_from_obj:
            date_from_obj = default_from
        if not date_to_obj:
            date_to_obj = today
        if date_from_obj > date_to_obj:
            date_from_obj, date_to_obj = date_to_obj, date_from_obj
        batches_qs = batches_qs.filter(created_at__date__gte=date_from_obj, created_at__date__lte=date_to_obj)

    if history_status != "ALL":
        batches_qs = batches_qs.filter(status=history_status)

    if history_q:
        batch_code_id = None
        cleaned = history_q.upper().replace("SUP-", "").strip()
        if cleaned.isdigit():
            try:
                batch_code_id = int(cleaned)
            except Exception:
                batch_code_id = None
        search_filter = (
            Q(original_filename__icontains=history_q)
            | Q(upload_remark__icontains=history_q)
            | Q(rows__seller_order_code__icontains=history_q)
            | Q(rows__seller_name__icontains=history_q)
            | Q(rows__receiver_name__icontains=history_q)
            | Q(rows__receiver_phone__icontains=history_q)
            | Q(rows__product_name_input__icontains=history_q)
            | Q(rows__sku_input__icontains=history_q)
            | Q(rows__matched_product_name__icontains=history_q)
            | Q(rows__matched_sku__icontains=history_q)
            | Q(rows__product_desc__icontains=history_q)
        )
        if batch_code_id is not None:
            search_filter = search_filter | Q(id=batch_code_id)
        batches_qs = batches_qs.filter(search_filter).distinct()

    result_count = batches_qs.count()
    paginator = Paginator(batches_qs, 50)
    page_obj = paginator.get_page(request.GET.get("page"))

    # Upload history remark is the batch-level remark entered when uploading the file.
    # Do not use row Remark from Excel here.
    batches = list(page_obj.object_list)
    for batch in batches:
        batch.display_upload_remark = (batch.upload_remark or "").strip() or "-"
        batch.display_filename = (batch.original_filename or "").strip() or "-"

    return render(request, "customerportal/computer/upload_orders.html", {"seller": seller, "batches": batches, "page_obj": page_obj, "paginator": paginator, "result_count": result_count, "history_q": history_q, "history_status": history_status, "history_period": history_period, "date_from": (date_from_obj or default_from).isoformat(), "date_to": (date_to_obj or today).isoformat()})


@login_required(login_url="portal:computer_login")
def computer_upload_order_detail(request, batch_id):
    seller = _get_logged_in_seller(request)
    if seller is None:
        return redirect("portal:computer_login")
    batch = get_object_or_404(SellerUploadBatch, id=batch_id, seller=seller)
    rows = list(batch.rows.all().order_by("row_number", "id"))
    _decorate_upload_rows_product_display(seller, rows)
    return render(request, "customerportal/computer/upload_order_detail.html", {"seller": seller, "batch": batch, "rows": rows})


@login_required(login_url="portal:computer_login")
def computer_upload_order_delete(request, batch_id):
    seller = _get_logged_in_seller(request)
    if seller is None:
        return redirect("portal:computer_login")
    batch = get_object_or_404(SellerUploadBatch, id=batch_id, seller=seller)
    if request.method != "POST":
        return redirect("portal:computer_upload_order_detail", batch_id=batch.id)
    if batch.status != SellerUploadBatch.STATUS_PENDING:
        messages.error(request, "You cannot delete this file. The file has already been prepared. Please contact support team.")
        return redirect("portal:computer_upload_order_detail", batch_id=batch.id)
    batch_code = batch.code
    batch.delete()
    messages.success(request, f"Upload {batch_code} deleted.")
    return redirect("portal:computer_upload_orders")
