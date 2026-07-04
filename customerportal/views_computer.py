from datetime import date, timedelta
from decimal import Decimal

import io
import os
import re
import requests

from django.conf import settings
from django.contrib import messages
from django.core.paginator import Paginator
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from openpyxl import Workbook, load_workbook

from orders.activity import add_order_activity
from orders.audit import add_audit_log
from orders.models import AuditLog, ImportBatch, Order, OrderActivity
from orders.pricing import apply_pricing
from inventory.models import InventorySellerSetting
from inventory.services import (
    get_seller_current_stock,
    get_seller_inventory_setting,
)

from .models import SellerUploadBatch, SellerUploadRow
from .views import get_user_seller


ZERO = Decimal("0.00")


# =========================================================
# COMMON HELPERS
# =========================================================

def _parse_date(value):
    try:
        return date.fromisoformat((value or "").strip())
    except (TypeError, ValueError):
        return None


def _safe_rate(part, total):
    if not total:
        return 0

    try:
        value = round((part * 100.0) / total, 2)
    except (TypeError, ValueError, ZeroDivisionError):
        return 0

    return min(max(value, 0), 100)


def _get_report_range(request):
    today = timezone.localdate()
    period = (request.GET.get("period") or "last_30_days").strip()

    if period == "today":
        start_date = today
        end_date = today

    elif period == "yesterday":
        start_date = today - timedelta(days=1)
        end_date = start_date

    elif period == "last_7_days":
        start_date = today - timedelta(days=6)
        end_date = today

    elif period == "last_30_days":
        start_date = today - timedelta(days=29)
        end_date = today

    elif period == "this_month":
        start_date = today.replace(day=1)
        end_date = today

    elif period == "custom":
        start_date = _parse_date(request.GET.get("from"))
        end_date = _parse_date(request.GET.get("to"))

        if not start_date or not end_date:
            period = "last_30_days"
            start_date = today - timedelta(days=29)
            end_date = today
        else:
            if start_date > end_date:
                start_date, end_date = end_date, start_date

            start_date = min(start_date, today)
            end_date = min(end_date, today)

            if start_date > end_date:
                start_date = end_date

            if (end_date - start_date).days > 365:
                start_date = end_date - timedelta(days=365)

    else:
        period = "last_30_days"
        start_date = today - timedelta(days=29)
        end_date = today

    if start_date == end_date:
        period_label = start_date.strftime("%d %B %Y")
    else:
        period_label = (
            f"{start_date.strftime('%d %b %Y')} - "
            f"{end_date.strftime('%d %b %Y')}"
        )

    return period, start_date, end_date, period_label


def _get_logged_in_seller(request):
    seller = get_user_seller(request.user)

    if seller is None:
        logout(request)
        return None

    return seller


def _base_seller_orders(seller):
    return Order.objects.filter(
        seller=seller,
        is_deleted=False,
    )


def _pending_statuses():
    return {
        Order.STATUS_CREATED,
        Order.STATUS_INBOUND,
        Order.STATUS_OUT_FOR_DELIVERY,
        Order.STATUS_PROVINCE_ASSIGNED,
        Order.STATUS_RETURN_ASSIGNED,
        getattr(Order, "STATUS_RETURNING", "RETURNING"),
        getattr(Order, "STATUS_PROCESSING", "PROCESSING"),
    }


def _order_delivery_fee(order):
    return getattr(order, "delivery_fee", ZERO) or ZERO


def _order_additional_fee(order):
    return getattr(order, "additional_fee", ZERO) or ZERO


def _order_province_fee(order):
    return getattr(order, "province_fee", ZERO) or ZERO


def _order_total_fee(order):
    return (
        _order_delivery_fee(order)
        + _order_additional_fee(order)
        + _order_province_fee(order)
    )


def _order_cod(order):
    price = getattr(order, "price", ZERO) or ZERO
    cod = getattr(order, "cod", ZERO) or ZERO
    return price if price != ZERO else cod


def _is_province_order(order):
    status = (getattr(order, "status", "") or "").strip().upper()
    province_fee = _order_province_fee(order)

    return (
        status
        == getattr(
            Order,
            "STATUS_PROVINCE_ASSIGNED",
            "PROVINCE_ASSIGNED",
        )
        or province_fee > ZERO
    )


def _get_created_date(order):
    created_at = getattr(order, "created_at", None)

    if not created_at:
        return None

    try:
        if timezone.is_aware(created_at):
            return timezone.localtime(created_at).date()
    except (TypeError, ValueError):
        pass

    try:
        return created_at.date()
    except (AttributeError, TypeError, ValueError):
        return None


def _get_computer_status(order):
    status = (getattr(order, "status", "") or "").strip().upper()

    if status == getattr(Order, "STATUS_DELIVERED", "DELIVERED"):
        return "delivered", "Delivered"

    if status == getattr(Order, "STATUS_RETURNED", "RETURNED"):
        return "returned", "Returned"

    if status == getattr(Order, "STATUS_VOID", "VOID"):
        return "void", "Void"

    if status == getattr(
        Order,
        "STATUS_PROVINCE_ASSIGNED",
        "PROVINCE_ASSIGNED",
    ):
        return "province", "Province"

    if status in {
        getattr(Order, "STATUS_RETURN_ASSIGNED", "RETURN_ASSIGNED"),
        getattr(Order, "STATUS_RETURNING", "RETURNING"),
    }:
        return "returning", "Returning"

    if status == getattr(
        Order,
        "STATUS_OUT_FOR_DELIVERY",
        "OUT_FOR_DELIVERY",
    ):
        return "pending", "Delivering"

    if status == getattr(Order, "STATUS_INBOUND", "INBOUND"):
        return "pending", "Processing"

    return "pending", "Pending"


def _decorate_order(order):
    status_key, status_label = _get_computer_status(order)

    order.computer_status_key = status_key
    order.computer_status_label = status_label
    order.computer_cod = _order_cod(order)
    order.computer_delivery_fee = _order_delivery_fee(order)
    order.computer_additional_fee = _order_additional_fee(order)
    order.computer_province_fee = _order_province_fee(order)
    order.computer_total_fee = _order_total_fee(order)
    order.computer_fee = order.computer_total_fee
    order.computer_net = order.computer_cod - order.computer_total_fee

    return order


def _common_context(seller, period, start_date, end_date, period_label):
    return {
        "seller": seller,
        "period": period,
        "date_from": start_date.isoformat(),
        "date_to": end_date.isoformat(),
        "period_label": period_label,
        "today_iso": timezone.localdate().isoformat(),
    }


# =========================================================
# COMPUTER DASHBOARD
# =========================================================

@login_required(login_url="portal:computer_login")
def computer_dashboard(request):
    seller = _get_logged_in_seller(request)

    if seller is None:
        return redirect("portal:computer_login")

    period, start_date, end_date, period_label = _get_report_range(request)
    seller_orders = _base_seller_orders(seller)

    created_orders = list(
        seller_orders.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date,
        ).order_by("created_at", "id")
    )

    delivered_orders = list(
        seller_orders.filter(
            status=Order.STATUS_DELIVERED,
            done_at__gte=start_date,
            done_at__lte=end_date,
        ).order_by("done_at", "id")
    )

    returned_orders = list(
        seller_orders.filter(
            status=Order.STATUS_RETURNED,
            done_at__gte=start_date,
            done_at__lte=end_date,
        ).order_by("done_at", "id")
    )

    pending_statuses = _pending_statuses()
    rows_by_date = {}
    current_date = start_date

    while current_date <= end_date:
        rows_by_date[current_date] = {
            "date": current_date,
            "new_orders": 0,
            "delivered": 0,
            "pending": 0,
            "returned": 0,
            "province": 0,
            "cod": ZERO,
            "fees": ZERO,
            "net": ZERO,
            "delivery_rate": 0,
            "return_rate": 0,
        }
        current_date += timedelta(days=1)

    total_pending = 0
    total_province = 0

    for order in created_orders:
        created_date = _get_created_date(order)

        if created_date not in rows_by_date:
            continue

        row = rows_by_date[created_date]
        row["new_orders"] += 1

        if order.status in pending_statuses:
            row["pending"] += 1
            total_pending += 1

        if _is_province_order(order):
            row["province"] += 1
            total_province += 1

    total_cod = ZERO
    total_fees = ZERO

    for order in delivered_orders:
        done_date = order.done_at

        if done_date not in rows_by_date:
            continue

        row = rows_by_date[done_date]
        cod_value = _order_cod(order)
        fee_value = _order_total_fee(order)

        row["delivered"] += 1
        row["cod"] += cod_value
        row["fees"] += fee_value

        total_cod += cod_value
        total_fees += fee_value

    for order in returned_orders:
        done_date = order.done_at

        if done_date in rows_by_date:
            rows_by_date[done_date]["returned"] += 1

    daily_rows = []

    for row_date in sorted(rows_by_date.keys(), reverse=True):
        row = rows_by_date[row_date]
        row["delivery_rate"] = _safe_rate(
            row["delivered"],
            row["new_orders"],
        )
        row["return_rate"] = _safe_rate(
            row["returned"],
            row["new_orders"],
        )
        row["net"] = row["cod"] - row["fees"]
        daily_rows.append(row)

    total_orders = len(created_orders)
    total_delivered = len(delivered_orders)
    total_returned = len(returned_orders)
    net_balance = total_cod - total_fees

    recent_orders = list(seller_orders.order_by("-id")[:8])

    for order in recent_orders:
        _decorate_order(order)

    cancelled_orders = sum(
        1
        for order in created_orders
        if order.status == Order.STATUS_VOID
    )
    cancellation_rate = _safe_rate(cancelled_orders, total_orders)

    inventory_setting = get_seller_inventory_setting(seller)

    if (
        inventory_setting.stock_mode == InventorySellerSetting.NO_STOCK
        or not inventory_setting.show_stock_in_portal
    ):
        dashboard_stock_rows = []
        automatic_stock_status = "Hidden"
    else:
        dashboard_stock_rows = get_seller_current_stock(seller)
        automatic_stock_status = "Live"

    dashboard_inventory_products = len(dashboard_stock_rows)
    dashboard_available_stock = sum(
        int(row.get("available_qty", 0) or 0)
        for row in dashboard_stock_rows
    )
    dashboard_reserved_stock = sum(
        int(row.get("reserved_qty", 0) or 0)
        for row in dashboard_stock_rows
    )

    context = _common_context(
        seller,
        period,
        start_date,
        end_date,
        period_label,
    )

    context.update(
        {
            "total_orders": total_orders,
            "total_delivered": total_delivered,
            "total_pending": total_pending,
            "total_returned": total_returned,
            "total_province": total_province,
            "total_cod": total_cod,
            "total_fees": total_fees,
            "net_balance": net_balance,
            "delivery_rate": _safe_rate(total_delivered, total_orders),
            "return_rate": _safe_rate(total_returned, total_orders),
            "daily_rows": daily_rows,
            "recent_orders": recent_orders,
            "cancelled_orders": cancelled_orders,
            "cancellation_rate": cancellation_rate,
            "automatic_stock_status": automatic_stock_status,
            "dashboard_inventory_products": dashboard_inventory_products,
            "dashboard_available_stock": dashboard_available_stock,
            "dashboard_reserved_stock": dashboard_reserved_stock,
        }
    )

    return render(
        request,
        "customerportal/computer/dashboard.html",
        context,
    )


# =========================================================
# COMPUTER ORDERS
# =========================================================

@login_required(login_url="portal:computer_login")
def computer_orders(request):
    seller = _get_logged_in_seller(request)

    if seller is None:
        return redirect("portal:computer_login")

    period, start_date, end_date, period_label = _get_report_range(request)
    q = (request.GET.get("q") or "").strip()
    status_filter = (request.GET.get("status") or "ALL").strip().upper()
    pending_statuses = _pending_statuses()

    period_orders = _base_seller_orders(seller).filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
    )

    summary_total = period_orders.count()
    summary_delivered = period_orders.filter(
        status=Order.STATUS_DELIVERED,
    ).count()
    summary_pending = period_orders.filter(
        status__in=pending_statuses,
    ).count()
    summary_returned = period_orders.filter(
        status=Order.STATUS_RETURNED,
    ).count()

    orders = period_orders

    if q:
        orders = orders.filter(
            Q(tracking_no__icontains=q)
            | Q(seller_order_code__icontains=q)
            | Q(receiver_name__icontains=q)
            | Q(receiver_phone__icontains=q)
            | Q(receiver_address__icontains=q)
            | Q(product_desc__icontains=q)
        )

    if status_filter == "PENDING":
        orders = orders.filter(status__in=pending_statuses)
    elif status_filter == "DELIVERED":
        orders = orders.filter(status=Order.STATUS_DELIVERED)
    elif status_filter == "RETURNED":
        orders = orders.filter(status=Order.STATUS_RETURNED)
    elif status_filter == "PROVINCE":
        orders = orders.filter(
            Q(status=Order.STATUS_PROVINCE_ASSIGNED)
            | Q(province_fee__gt=ZERO)
        )
    elif status_filter == "VOID":
        orders = orders.filter(status=Order.STATUS_VOID)

    orders = list(orders.order_by("-created_at", "-id"))

    for order in orders:
        _decorate_order(order)

    context = _common_context(
        seller,
        period,
        start_date,
        end_date,
        period_label,
    )

    context.update(
        {
            "q": q,
            "status_filter": status_filter,
            "orders": orders,
            "summary_total": summary_total,
            "summary_delivered": summary_delivered,
            "summary_pending": summary_pending,
            "summary_returned": summary_returned,
        }
    )

    return render(
        request,
        "customerportal/computer/orders.html",
        context,
    )


# =========================================================
# COMPUTER DELIVERY REPORT
# =========================================================

@login_required(login_url="portal:computer_login")
def computer_delivery_report(request):
    seller = _get_logged_in_seller(request)

    if seller is None:
        return redirect("portal:computer_login")

    period, start_date, end_date, period_label = _get_report_range(request)
    seller_orders = _base_seller_orders(seller)
    pending_statuses = _pending_statuses()

    created_orders = seller_orders.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
    )

    completed_rows = list(
        seller_orders.filter(
            status=Order.STATUS_DELIVERED,
            done_at__gte=start_date,
            done_at__lte=end_date,
        ).order_by("-done_at", "-id")
    )

    returned_rows = list(
        seller_orders.filter(
            status=Order.STATUS_RETURNED,
            done_at__gte=start_date,
            done_at__lte=end_date,
        ).order_by("-done_at", "-id")
    )

    for order in completed_rows:
        _decorate_order(order)

    for order in returned_rows:
        _decorate_order(order)

    total_created = created_orders.count()
    total_pending = created_orders.filter(
        status__in=pending_statuses,
    ).count()
    total_delivered = len(completed_rows)
    total_returned = len(returned_rows)

    context = _common_context(
        seller,
        period,
        start_date,
        end_date,
        period_label,
    )

    context.update(
        {
            "total_created": total_created,
            "total_delivered": total_delivered,
            "total_pending": total_pending,
            "total_returned": total_returned,
            "delivery_rate": _safe_rate(total_delivered, total_created),
            "return_rate": _safe_rate(total_returned, total_created),
            "completed_rows": completed_rows,
            "returned_rows": returned_rows,
        }
    )

    return render(
        request,
        "customerportal/computer/delivery_report.html",
        context,
    )


# =========================================================
# COMPUTER COD REPORT
# =========================================================

@login_required(login_url="portal:computer_login")
def computer_cod_report(request):
    seller = _get_logged_in_seller(request)

    if seller is None:
        return redirect("portal:computer_login")

    period, start_date, end_date, period_label = _get_report_range(request)

    rows = list(
        _base_seller_orders(seller).filter(
            status=Order.STATUS_DELIVERED,
            done_at__gte=start_date,
            done_at__lte=end_date,
        ).order_by("-done_at", "-id")
    )

    total_cod = ZERO
    total_delivery_fee = ZERO
    total_additional_fee = ZERO
    total_province_fee = ZERO
    total_fees = ZERO
    net_balance = ZERO

    for order in rows:
        _decorate_order(order)

        total_cod += order.computer_cod
        total_delivery_fee += order.computer_delivery_fee
        total_additional_fee += order.computer_additional_fee
        total_province_fee += order.computer_province_fee
        total_fees += order.computer_total_fee
        net_balance += order.computer_net

    context = _common_context(
        seller,
        period,
        start_date,
        end_date,
        period_label,
    )

    context.update(
        {
            "rows": rows,
            "total_orders": len(rows),
            "total_cod": total_cod,
            "total_delivery_fee": total_delivery_fee,
            "total_additional_fee": total_additional_fee,
            "total_province_fee": total_province_fee,
            "total_fees": total_fees,
            "net_balance": net_balance,
        }
    )

    return render(
        request,
        "customerportal/computer/cod_report.html",
        context,
    )

# =========================================================
# COMPUTER INVENTORY
# =========================================================

@login_required(login_url="portal:computer_login")
def computer_inventory(request):
    seller = _get_logged_in_seller(request)

    if seller is None:
        return redirect("portal:computer_login")

    setting = get_seller_inventory_setting(seller)

    if (
        setting.stock_mode == InventorySellerSetting.NO_STOCK
        or not setting.show_stock_in_portal
    ):
        stock_rows = []
        inventory_enabled = False
    else:
        stock_rows = get_seller_current_stock(seller)
        inventory_enabled = True

    q = (request.GET.get("q") or "").strip().lower()

    if q:
        stock_rows = [
            row
            for row in stock_rows
            if q in str(row.get("sku", "") or "").lower()
            or q in str(row.get("name", "") or "").lower()
            or q in str(row.get("product_type", "") or "").lower()
            or q in str(row.get("location", "") or "").lower()
        ]

    total_products = len(stock_rows)
    total_current = sum(
        int(row.get("current_qty", 0) or 0)
        for row in stock_rows
    )
    total_reserved = sum(
        int(row.get("reserved_qty", 0) or 0)
        for row in stock_rows
    )
    total_available = sum(
        int(row.get("available_qty", 0) or 0)
        for row in stock_rows
    )

    context = {
        "seller": seller,
        "setting": setting,
        "inventory_enabled": inventory_enabled,
        "stock_rows": stock_rows,
        "q": q,
        "total_products": total_products,
        "total_current": total_current,
        "total_reserved": total_reserved,
        "total_available": total_available,
    }

    return render(
        request,
        "customerportal/computer/inventory.html",
        context,
    )

# =========================================================
# COMPUTER PORTAL - CUSTOMER ORDER UPLOAD
# =========================================================

UPLOAD_HEADERS = [
    "Seller Order Code",
    "Seller",
    "Product Description",
    "Quantity",
    "COD",
    "Receiver Name",
    "Phone",
    "Address",
    "Remark",
    "SKU",
]

UPLOAD_SHEET_NAME = "DS_UPLOAD"
UPLOAD_TEMPLATE_KEY = "DS_EXPRESS_UPLOAD_TEMPLATE_V1"
UPLOAD_MAX_SIZE = 5 * 1024 * 1024
UPLOAD_MAX_ROWS = 500


def _upload_cell(value):
    return "" if value is None else str(value).strip()


def _upload_decimal(value):
    try:
        raw = str(value or "").strip()
        if raw == "":
            return Decimal("0.00")
        return Decimal(raw).quantize(Decimal("0.00"))
    except Exception:
        return Decimal("0.00")


def _upload_int(value, default=1):
    try:
        return max(int(float(str(value).strip())), 1)
    except Exception:
        return default


def _normalize_phone(value):
    raw = str(value or "").strip()
    digits = re.sub(r"\D+", "", raw)
    return digits


def _cell_has_danger(cell):
    value = cell.value

    if cell.data_type == "f":
        return True

    if isinstance(value, str):
        text = value.strip()

        if text.startswith("=") or text.startswith("@"):
            return True

        bad_words = ["<script", "</script", "drop table", "delete from", "insert into", "update "]
        low = text.lower()
        return any(x in low for x in bad_words)

    return False


def _find_upload_product_by_sku(seller, sku):
    """
    Product Description is only customer text. Inventory matching uses SKU only.
    This prevents wrong product names from changing stock matching.
    """
    try:
        from inventory.models import StockProduct
    except Exception:
        return None

    sku = (sku or "").strip()
    if not sku:
        return None

    return (
        StockProduct.objects
        .filter(seller=seller, is_active=True, sku__iexact=sku)
        .first()
    )


def _decorate_upload_rows_product_display(seller, rows):
    """
    Display helper for upload detail pages.
    Shows the readable inventory product name from SKU, while keeping the
    customer's uploaded Product Description visible for verification.
    """
    try:
        from inventory.models import StockProduct
        from inventory.services import match_product
    except Exception:
        StockProduct = None
        match_product = None

    for row in rows:
        input_description = (getattr(row, "product_name_input", "") or "").strip()
        input_sku = (getattr(row, "sku_input", "") or "").strip()
        matched_name = (getattr(row, "matched_product_name", "") or "").strip()
        matched_sku = (getattr(row, "matched_sku", "") or "").strip()
        old_product_text = (getattr(row, "product_desc", "") or "").strip()

        product = None

        if StockProduct:
            # New uploads use sku_input/matched_sku. Older uploads may have stored
            # only the SKU in product_desc, so keep that fallback.
            for val in [matched_sku, input_sku, old_product_text]:
                if not val:
                    continue
                product = (
                    StockProduct.objects
                    .filter(seller=seller, is_active=True, sku__iexact=val)
                    .first()
                )
                if product:
                    break

        if not product and match_product:
            # Fallback for very old rows where product_desc stored a product name.
            for val in [matched_name, old_product_text]:
                if not val:
                    continue
                product = match_product(seller, val)
                if product:
                    break

        if product:
            row.display_inventory_product = product.name or "-"
            row.display_inventory_sku = product.sku or "-"
        else:
            row.display_inventory_product = matched_name or "-"
            row.display_inventory_sku = matched_sku or input_sku or "-"

        # Product Description is the customer's free text. For older rows that only
        # stored SKU in product_desc, avoid showing SKU as description when possible.
        if input_description:
            row.display_product_description = input_description
        elif product and old_product_text and old_product_text.upper() == (getattr(product, "sku", "") or "").upper():
            row.display_product_description = "-"
        else:
            row.display_product_description = old_product_text or "-"

        row.display_uploaded_sku = input_sku or matched_sku or "-"


def _validate_upload_product_stock(seller, product, qty, strict_stock_used):
    try:
        from inventory.models import InventorySellerSetting
        from inventory.services import current_available_qty, get_seller_inventory_setting
    except Exception:
        return True, ""

    setting = get_seller_inventory_setting(seller)

    if setting.stock_mode == InventorySellerSetting.NO_STOCK:
        return True, ""

    if not product:
        return False, "SKU not recognized in inventory. Please check SKU."

    if setting.stock_mode == InventorySellerSetting.STRICT:
        available_qty = current_available_qty(product)
        already_used_qty = int(strict_stock_used.get(product.id, 0) or 0)
        remaining_qty = available_qty - already_used_qty

        if remaining_qty < qty:
            return False, f"Strict stock shop: not enough stock for {product.name}. Need {qty}, available {remaining_qty}."

        strict_stock_used[product.id] = already_used_qty + qty

    return True, ""


def _recalc_upload_batch(batch):
    rows = batch.rows.all()
    batch.total_rows = rows.count()
    batch.valid_rows = rows.filter(status=SellerUploadRow.STATUS_VALID).count()
    batch.error_rows = rows.filter(status=SellerUploadRow.STATUS_ERROR).count()
    batch.duplicate_rows = rows.filter(status=SellerUploadRow.STATUS_DUPLICATE).count()
    batch.save(update_fields=["total_rows", "valid_rows", "error_rows", "duplicate_rows", "updated_at"])


def _send_customer_upload_staff_telegram(batch):
    bot_token = getattr(settings, "TELEGRAM_DS_TEAM_BOT_TOKEN", "")
    chat_id = getattr(settings, "TELEGRAM_DS_TEAM_CHAT_ID", "")
    if not bot_token or not chat_id:
        return
    text = (
        "📥 New Customer Upload Waiting Approval\n\n"
        f"Shop: {batch.seller.name}\n"
        f"Batch: {batch.code}\n"
        f"Rows: {batch.total_rows}\n"
        f"Valid: {batch.valid_rows}\n"
        f"File: {batch.original_filename or '-'}\n"
        f"Upload Remark: {batch.upload_remark or '-'}"
    )
    try:
        requests.post(f"https://api.telegram.org/bot{bot_token}/sendMessage", data={"chat_id": chat_id, "text": text}, timeout=10)
    except Exception:
        pass


@login_required(login_url="portal:computer_login")
def download_customer_upload_sample(request):
    seller = _get_logged_in_seller(request)
    if seller is None:
        return redirect("portal:computer_login")

    wb = Workbook()
    ws = wb.active
    ws.title = UPLOAD_SHEET_NAME
    ws["A1"] = UPLOAD_TEMPLATE_KEY
    ws["A2"] = "Do not edit row 1. Fill data from row 4 only. Product Description is free text. SKU controls inventory matching."

    for col, header in enumerate(UPLOAD_HEADERS, start=1):
        ws.cell(row=3, column=col, value=header)

    ws.append([
        "SO-001",
        seller.name,
        "Customer product text / seller product name",
        1,
        10,
        "Customer Name",
        "012345678",
        "Phnom Penh address...",
        "",
        "260001-ITE-002",
    ])

    widths = [18, 20, 34, 12, 12, 22, 16, 34, 24, 18]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + col)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="ds_customer_upload_sample.xlsx"'
    return response


@login_required(login_url="portal:computer_login")
def computer_upload_orders(request):
    seller = _get_logged_in_seller(request)
    if seller is None:
        return redirect("portal:computer_login")

    if request.method == "POST":
        upload_remark = (request.POST.get("upload_remark") or "").strip()
        upload_file = request.FILES.get("file")
        if not upload_file:
            messages.error(request, "Please choose an Excel file.")
            return redirect("portal:computer_upload_orders")

        original_filename = upload_file.name or ""
        ext = os.path.splitext(original_filename)[1].lower()

        if ext != ".xlsx":
            batch = SellerUploadBatch.objects.create(seller=seller, uploaded_by=request.user, original_filename=original_filename, upload_remark=upload_remark, status=SellerUploadBatch.STATUS_REJECTED, reject_reason="Invalid file type. Please upload .xlsx official sample only.", rejected_by=request.user, rejected_at=timezone.now())
            messages.error(request, "Rejected: only .xlsx official sample is allowed.")
            return redirect("portal:computer_upload_order_detail", batch_id=batch.id)

        if upload_file.size > UPLOAD_MAX_SIZE:
            batch = SellerUploadBatch.objects.create(seller=seller, uploaded_by=request.user, original_filename=original_filename, upload_remark=upload_remark, status=SellerUploadBatch.STATUS_REJECTED, reject_reason="File too large. Maximum size is 5MB.", rejected_by=request.user, rejected_at=timezone.now())
            messages.error(request, "Rejected: file too large.")
            return redirect("portal:computer_upload_order_detail", batch_id=batch.id)

        batch = SellerUploadBatch.objects.create(seller=seller, uploaded_by=request.user, file=upload_file, original_filename=original_filename, upload_remark=upload_remark, status=SellerUploadBatch.STATUS_PENDING)

        try:
            batch.file.open("rb")
            wb = load_workbook(filename=batch.file, data_only=False)
        except Exception:
            batch.status = SellerUploadBatch.STATUS_REJECTED
            batch.reject_reason = "Cannot read Excel file."
            batch.rejected_by = request.user
            batch.rejected_at = timezone.now()
            batch.save()
            messages.error(request, "Rejected: cannot read Excel file.")
            return redirect("portal:computer_upload_order_detail", batch_id=batch.id)

        if UPLOAD_SHEET_NAME not in wb.sheetnames:
            batch.status = SellerUploadBatch.STATUS_REJECTED
            batch.reject_reason = "Invalid template. Sheet name must be DS_UPLOAD."
            batch.rejected_by = request.user
            batch.rejected_at = timezone.now()
            batch.save()
            messages.error(request, "Rejected: invalid DS template.")
            return redirect("portal:computer_upload_order_detail", batch_id=batch.id)

        ws = wb[UPLOAD_SHEET_NAME]
        if _upload_cell(ws["A1"].value) != UPLOAD_TEMPLATE_KEY:
            batch.status = SellerUploadBatch.STATUS_REJECTED
            batch.reject_reason = "Invalid template key. Please download the official DS sample."
            batch.rejected_by = request.user
            batch.rejected_at = timezone.now()
            batch.save()
            messages.error(request, "Rejected: please use DS official sample.")
            return redirect("portal:computer_upload_order_detail", batch_id=batch.id)

        headers = [_upload_cell(ws.cell(row=3, column=i).value) for i in range(1, len(UPLOAD_HEADERS) + 1)]
        if headers != UPLOAD_HEADERS:
            batch.status = SellerUploadBatch.STATUS_REJECTED
            batch.reject_reason = "Invalid columns. Please use the official DS sample without editing columns."
            batch.rejected_by = request.user
            batch.rejected_at = timezone.now()
            batch.save()
            messages.error(request, "Rejected: invalid columns.")
            return redirect("portal:computer_upload_order_detail", batch_id=batch.id)

        seen_codes = set()
        strict_stock_used = {}
        created_rows = 0

        for row_idx in range(4, ws.max_row + 1):
            if created_rows >= UPLOAD_MAX_ROWS:
                SellerUploadRow.objects.create(batch=batch, row_number=row_idx, status=SellerUploadRow.STATUS_ERROR, error_message=f"Upload limit is {UPLOAD_MAX_ROWS} rows.")
                continue

            row_values = [ws.cell(row=row_idx, column=col).value for col in range(1, len(UPLOAD_HEADERS) + 1)]
            if all(_upload_cell(v) == "" for v in row_values):
                continue

            created_rows += 1
            errors = []
            for col in range(1, len(UPLOAD_HEADERS) + 1):
                if _cell_has_danger(ws.cell(row=row_idx, column=col)):
                    errors.append("Dangerous formula/text detected.")
                    break

            seller_order_code = _upload_cell(row_values[0])
            seller_name = _upload_cell(row_values[1]) or (seller.name or "")
            product_description = _upload_cell(row_values[2])
            quantity = _upload_int(row_values[3], 1)
            cod = _upload_decimal(row_values[4])
            price = cod
            receiver_name = _upload_cell(row_values[5])
            phone = _normalize_phone(row_values[6])
            address = _upload_cell(row_values[7])
            remark = _upload_cell(row_values[8])
            sku_input = _upload_cell(row_values[9])

            if not seller_order_code:
                errors.append("Seller Order Code is required.")
            if not product_description:
                errors.append("Product Description is required.")
            if not sku_input:
                errors.append("SKU is required because inventory matching uses SKU.")
            if not receiver_name:
                errors.append("Receiver Name is required.")
            if not phone:
                errors.append("Phone is required.")
            elif len(phone) < 8 or len(phone) > 15:
                errors.append("Phone number is invalid.")
            if not address:
                errors.append("Address is required.")
            if cod < Decimal("0.00"):
                errors.append("COD cannot be negative.")

            normalized_code = seller_order_code.lower()
            row_status = SellerUploadRow.STATUS_VALID
            if normalized_code in seen_codes:
                row_status = SellerUploadRow.STATUS_DUPLICATE
                errors.append("Duplicate Seller Order Code in this upload.")
            elif seller_order_code and Order.objects.filter(seller=seller, seller_order_code__iexact=seller_order_code, is_deleted=False).exists():
                row_status = SellerUploadRow.STATUS_DUPLICATE
                errors.append("Seller Order Code already exists in system.")
            if seller_order_code:
                seen_codes.add(normalized_code)

            matched_product = _find_upload_product_by_sku(seller, sku_input)
            matched_product_name = getattr(matched_product, "name", "") or ""
            matched_sku = getattr(matched_product, "sku", "") or ""

            ok_stock, stock_error = _validate_upload_product_stock(seller=seller, product=matched_product, qty=quantity, strict_stock_used=strict_stock_used)
            if not ok_stock:
                errors.append(stock_error)

            if errors and row_status != SellerUploadRow.STATUS_DUPLICATE:
                row_status = SellerUploadRow.STATUS_ERROR

            SellerUploadRow.objects.create(
                batch=batch,
                row_number=row_idx,
                seller_order_code=seller_order_code,
                seller_name=seller_name,
                # Keep using existing field name for DB compatibility. It now stores Product Description.
                product_name_input=product_description,
                sku_input=sku_input,
                matched_product_name=matched_product_name,
                matched_sku=matched_sku,
                receiver_name=receiver_name,
                receiver_phone=phone,
                receiver_address=address,
                product_desc=product_description,
                quantity=quantity,
                cod=cod,
                price=price,
                remark=remark,
                status=row_status,
                error_message=" | ".join(errors),
            )

        _recalc_upload_batch(batch)

        if batch.total_rows <= 0:
            batch.status = SellerUploadBatch.STATUS_REJECTED
            batch.reject_reason = "No data rows found. Please fill data from row 4."
            batch.rejected_by = request.user
            batch.rejected_at = timezone.now()
            batch.save()
            messages.error(request, "Rejected: no data rows found.")
            return redirect("portal:computer_upload_order_detail", batch_id=batch.id)

        if batch.error_rows > 0 or batch.duplicate_rows > 0:
            batch.status = SellerUploadBatch.STATUS_REJECTED
            batch.reject_reason = f"Auto rejected because upload has {batch.error_rows} error row(s) and {batch.duplicate_rows} duplicate row(s). Please fix the file and upload again."
            batch.rejected_by = request.user
            batch.rejected_at = timezone.now()
            batch.save(update_fields=["status", "reject_reason", "rejected_by", "rejected_at", "updated_at"])
            messages.error(request, f"Upload auto rejected. Errors: {batch.error_rows}, Duplicate: {batch.duplicate_rows}. Please fix and upload again.")
            return redirect("portal:computer_upload_order_detail", batch_id=batch.id)

        _send_customer_upload_staff_telegram(batch)
        messages.success(request, f"Upload saved. Rows: {batch.total_rows}, Valid: {batch.valid_rows}. Waiting staff approval.")
        return redirect("portal:computer_upload_order_detail", batch_id=batch.id)

    today = timezone.localdate()
    default_from = today - timedelta(days=2)
    history_q = (request.GET.get("q") or "").strip()
    history_status = (request.GET.get("status") or "ALL").strip().upper()
    history_period = (request.GET.get("period") or "last_3_days").strip()
    date_from_obj = _parse_date(request.GET.get("from"))
    date_to_obj = _parse_date(request.GET.get("to"))
    batches_qs = SellerUploadBatch.objects.filter(seller=seller).order_by("-id")

    if history_period != "all":
        if not date_from_obj:
            date_from_obj = default_from
        if not date_to_obj:
            date_to_obj = today
        if date_from_obj > date_to_obj:
            date_from_obj, date_to_obj = date_to_obj, date_from_obj
        batches_qs = batches_qs.filter(created_at__date__gte=date_from_obj, created_at__date__lte=date_to_obj)

    if history_status != "ALL":
        batches_qs = batches_qs.filter(status=history_status)

    if history_q:
        batch_code_id = None
        cleaned = history_q.upper().replace("SUP-", "").strip()
        if cleaned.isdigit():
            try:
                batch_code_id = int(cleaned)
            except Exception:
                batch_code_id = None
        search_filter = (
            Q(original_filename__icontains=history_q)
            | Q(upload_remark__icontains=history_q)
            | Q(rows__seller_order_code__icontains=history_q)
            | Q(rows__seller_name__icontains=history_q)
            | Q(rows__receiver_name__icontains=history_q)
            | Q(rows__receiver_phone__icontains=history_q)
            | Q(rows__product_name_input__icontains=history_q)
            | Q(rows__sku_input__icontains=history_q)
            | Q(rows__matched_product_name__icontains=history_q)
            | Q(rows__matched_sku__icontains=history_q)
            | Q(rows__product_desc__icontains=history_q)
        )
        if batch_code_id is not None:
            search_filter = search_filter | Q(id=batch_code_id)
        batches_qs = batches_qs.filter(search_filter).distinct()

    result_count = batches_qs.count()
    paginator = Paginator(batches_qs, 50)
    page_obj = paginator.get_page(request.GET.get("page"))

    # Upload history remark is the batch-level remark entered when uploading the file.
    # Do not use row Remark from Excel here.
    batches = list(page_obj.object_list)
    for batch in batches:
        batch.display_upload_remark = (batch.upload_remark or "").strip() or "-"
        batch.display_filename = (batch.original_filename or "").strip() or "-"

    return render(request, "customerportal/computer/upload_orders.html", {"seller": seller, "batches": batches, "page_obj": page_obj, "paginator": paginator, "result_count": result_count, "history_q": history_q, "history_status": history_status, "history_period": history_period, "date_from": (date_from_obj or default_from).isoformat(), "date_to": (date_to_obj or today).isoformat()})


@login_required(login_url="portal:computer_login")
def computer_upload_order_detail(request, batch_id):
    seller = _get_logged_in_seller(request)
    if seller is None:
        return redirect("portal:computer_login")
    batch = get_object_or_404(SellerUploadBatch, id=batch_id, seller=seller)
    rows = list(batch.rows.all().order_by("row_number", "id"))
    _decorate_upload_rows_product_display(seller, rows)
    return render(request, "customerportal/computer/upload_order_detail.html", {"seller": seller, "batch": batch, "rows": rows})


@login_required(login_url="portal:computer_login")
def computer_upload_order_delete(request, batch_id):
    seller = _get_logged_in_seller(request)
    if seller is None:
        return redirect("portal:computer_login")
    batch = get_object_or_404(SellerUploadBatch, id=batch_id, seller=seller)
    if request.method != "POST":
        return redirect("portal:computer_upload_order_detail", batch_id=batch.id)
    if batch.status != SellerUploadBatch.STATUS_PENDING:
        messages.error(request, "You cannot delete this file. The file has already been prepared. Please contact support team.")
        return redirect("portal:computer_upload_order_detail", batch_id=batch.id)
    batch_code = batch.code
    batch.delete()
    messages.success(request, f"Upload {batch_code} deleted.")
    return redirect("portal:computer_upload_orders")
