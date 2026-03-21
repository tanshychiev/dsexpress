from __future__ import annotations

from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from masterdata.models import Seller


def _cell(v) -> str:
    return "" if v is None else str(v).strip()


def _to_bool(v) -> bool:
    s = _cell(v).lower()
    if s in {"0", "false", "no", "n", "inactive"}:
        return False
    return True


class Command(BaseCommand):
    help = "One-time import sellers from Excel"

    def add_arguments(self, parser):
        parser.add_argument("excel_path", type=str, help="Path to Excel file")

    def handle(self, *args, **options):
        excel_path = Path(options["excel_path"]).expanduser()

        if not excel_path.exists():
            raise CommandError(f"File not found: {excel_path}")

        try:
            wb = load_workbook(filename=excel_path, data_only=True)
            ws = wb.active
        except Exception as e:
            raise CommandError(f"Cannot open Excel file: {e}")

        headers = [_cell(c.value) for c in ws[1]]
        header_map = {h.lower(): idx for idx, h in enumerate(headers)}

        required = ["code", "name"]
        missing = [h for h in required if h not in header_map]
        if missing:
            raise CommandError(f"Missing required columns: {', '.join(missing)}")

        def get(row, col_name: str) -> str:
            idx = header_map.get(col_name.lower())
            if idx is None:
                return ""
            return _cell(row[idx].value)

        created_count = 0
        updated_count = 0
        skipped_count = 0
        errors: list[str] = []

        for row_idx in range(2, ws.max_row + 1):
            row = ws[row_idx]

            code = get(row, "Code")
            name = get(row, "Name")
            phone = get(row, "Phone")
            address = get(row, "Address")
            is_active_raw = get(row, "Is Active")

            if not any([code, name, phone, address, is_active_raw]):
                continue

            if not code:
                errors.append(f"Row {row_idx}: Code is required")
                continue

            if not name:
                errors.append(f"Row {row_idx}: Name is required")
                continue

            is_active = _to_bool(is_active_raw) if is_active_raw != "" else True

            seller = Seller.objects.filter(code=code).first()

            if seller:
                changed = False

                if seller.name != name:
                    seller.name = name
                    changed = True

                if seller.phone != phone:
                    seller.phone = phone
                    changed = True

                if seller.address != address:
                    seller.address = address
                    changed = True

                if seller.is_active != is_active:
                    seller.is_active = is_active
                    changed = True

                if changed:
                    seller.save()
                    updated_count += 1
                else:
                    skipped_count += 1
            else:
                Seller.objects.create(
                    code=code,
                    name=name,
                    phone=phone,
                    address=address,
                    is_active=is_active,
                )
                created_count += 1

        self.stdout.write(self.style.SUCCESS("Seller import finished"))
        self.stdout.write(f"Created: {created_count}")
        self.stdout.write(f"Updated: {updated_count}")
        self.stdout.write(f"Skipped: {skipped_count}")

        if errors:
            self.stdout.write(self.style.WARNING(f"Errors: {len(errors)}"))
            for err in errors:
                self.stdout.write(f" - {err}")