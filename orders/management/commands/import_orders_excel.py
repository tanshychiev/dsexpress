from __future__ import annotations

from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.utils.dateparse import parse_date, parse_datetime
from openpyxl import load_workbook

from masterdata.models import Seller, Shipper
from orders.models import Order


def to_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def to_int(value, default=0) -> int:
    if value in [None, ""]:
        return default
    try:
        return int(float(str(value).strip()))
    except Exception:
        return default


def to_decimal(value, default="0.00") -> Decimal:
    if value in [None, ""]:
        return Decimal(default)
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return Decimal(default)


def to_bool(value, default=False) -> bool:
    if value in [None, ""]:
        return default
    s = str(value).strip().lower()
    return s in ["1", "true", "yes", "y", "on"]


def to_date(value):
    if value in [None, ""]:
        return None
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day") and not hasattr(value, "hour"):
        return value
    s = str(value).strip()
    return parse_date(s)


def to_datetime(value):
    if value in [None, ""]:
        return None
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day") and hasattr(value, "hour"):
        return value
    s = str(value).strip()
    return parse_datetime(s)


class Command(BaseCommand):
    help = "Import old orders from Excel and keep same tracking/date/reason/all fields"

    def add_arguments(self, parser):
        parser.add_argument("excel_path", type=str, help="Path to Excel file")

    def handle(self, *args, **options):
        excel_path = Path(options["excel_path"])
        if not excel_path.exists():
            raise CommandError(f"File not found: {excel_path}")

        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active

        headers = [to_str(c.value) for c in ws[1]]
        header_map = {h: idx for idx, h in enumerate(headers)}

        required = ["tracking_no", "seller_code"]
        for col in required:
            if col not in header_map:
                raise CommandError(f"Missing required column: {col}")

        created_count = 0
        updated_count = 0
        skipped_count = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            def get(col_name, default=None):
                idx = header_map.get(col_name)
                if idx is None or idx >= len(row):
                    return default
                return row[idx]

            tracking_no = to_str(get("tracking_no"))
            seller_code = to_str(get("seller_code"))

            if not tracking_no:
                self.stdout.write(self.style.WARNING(f"Row {row_idx}: skipped, empty tracking_no"))
                skipped_count += 1
                continue

            if not seller_code:
                self.stdout.write(self.style.WARNING(f"Row {row_idx}: skipped, empty seller_code ({tracking_no})"))
                skipped_count += 1
                continue

            seller = Seller.objects.filter(code=seller_code).first()
            if not seller:
                self.stdout.write(self.style.WARNING(
                    f"Row {row_idx}: skipped, seller not found code={seller_code} tracking={tracking_no}"
                ))
                skipped_count += 1
                continue

            shipper_code = to_str(get("delivery_shipper_code"))
            delivery_shipper = None
            if shipper_code:
                delivery_shipper = Shipper.objects.filter(code=shipper_code).first()

            defaults = {
                "seller": seller,
                "seller_code": seller_code,
                "seller_name": to_str(get("seller_name")) or seller.name,
                "seller_order_code": to_str(get("seller_order_code")),
                "product_desc": to_str(get("product_desc")),
                "quantity": to_int(get("quantity"), 1),
                "price": to_decimal(get("price")),
                "cod": to_decimal(get("cod")),
                "delivery_fee": to_decimal(get("delivery_fee")),
                "additional_fee": to_decimal(get("additional_fee")),
                "province_fee": to_decimal(get("province_fee")),
                "receiver_name": to_str(get("receiver_name")),
                "receiver_phone": to_str(get("receiver_phone")),
                "receiver_address": to_str(get("receiver_address")),
                "remark": to_str(get("remark")),
                "reason": to_str(get("reason")),
                "status": to_str(get("status")) or Order.STATUS_CREATED,
                "delivery_shipper": delivery_shipper,
                "is_deleted": to_bool(get("is_deleted"), False),
                "is_locked": to_bool(get("is_locked"), False),
                "done_at": to_date(get("done_at")),
                "clear_delivery": to_bool(get("clear_delivery"), False),
            }

            order, created = Order.objects.update_or_create(
                tracking_no=tracking_no,
                defaults=defaults,
            )

            # keep original timestamps from Excel
            created_at = to_datetime(get("created_at"))
            updated_at = to_datetime(get("updated_at"))
            changed = []

            if created_at:
                order.created_at = created_at
                changed.append("created_at")

            if updated_at:
                order.updated_at = updated_at
                changed.append("updated_at")

            if changed:
                order.save(update_fields=changed)

            if created:
                created_count += 1
                self.stdout.write(self.style.SUCCESS(f"Created: {tracking_no}"))
            else:
                updated_count += 1
                self.stdout.write(f"Updated: {tracking_no}")

        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS(
            f"Done. Created={created_count}, Updated={updated_count}, Skipped={skipped_count}"
        ))