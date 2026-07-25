from io import BytesIO
from decimal import Decimal

from django.db.models import Q
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.utils import timezone

from .models import ProvinceCODBatch, ProvinceCODItem


ZERO = Decimal("0.00")


SORT_MAP = {
    "id": "id",
    "sent_date": "activity_date",
    "batch": "batch_id",
    "tracking": "order__tracking_no",
    "seller": "order__seller__name",
    "carrier": "batch__shipper__name",
    "receiver": "order__receiver_name",
    "phone": "order__receiver_phone",
    "original_cod": "original_cod",
    "province_fee": "province_fee",
    "status": "cod_status",
    "carrier_fee": "carrier_fee",
    "net_cod": "net_cod",
    "reference": "carrier_reference",
    "settled": "seller_settled",
    "updated": "updated_at",
}


def _money(value):
    try:
        return Decimal(str(value or ZERO)).quantize(Decimal("0.01"))
    except Exception:
        return ZERO


def _display(value, default="-"):
    text = str(value or "").strip()
    return text or default


def _date_time(value):
    if not value:
        return ""

    try:
        if timezone.is_aware(value):
            value = timezone.localtime(value)
        return value.strftime("%Y-%m-%d %H:%M")
    except Exception:
        return str(value)


def _customer_order_code(order):
    """
    Return the order code supplied by the customer/seller.

    Different DS Express versions may use a different field name, so this
    checks the common field names safely. The first non-empty value is used.
    Example value: COD P-JET-2307-01
    """
    possible_fields = (
        "order_code",
        "customer_order_code",
        "seller_order_code",
        "external_order_code",
        "code",
    )

    for field_name in possible_fields:
        value = getattr(order, field_name, None)
        if str(value or "").strip():
            return str(value).strip()

    return ""


def _real_cod(item):
    """
    The actual COD confirmed/received from the carrier.

    Province COD stores this amount in net_cod after staff enters the real
    received amount. Before confirmation it normally remains 0.00.
    """
    return _money(item.net_cod)


def _reason(item):
    """Use the issue/return reason first, then fall back to the latest note."""
    return _display(item.return_reason or item.note)


def _filtered_rows(request):
    date_from = (request.GET.get("date_from") or "").strip()
    date_to = (request.GET.get("date_to") or "").strip()
    status = (request.GET.get("status") or "").strip().upper()
    settlement = (request.GET.get("settlement") or "").strip().upper()
    seller_id = (request.GET.get("seller") or "").strip()
    shipper_id = (request.GET.get("shipper") or "").strip()
    q = (request.GET.get("q") or "").strip()

    sort = (request.GET.get("sort") or "sent_date").strip().lower()
    direction = (request.GET.get("direction") or "desc").strip().lower()

    if sort not in SORT_MAP:
        sort = "sent_date"

    if direction not in {"asc", "desc"}:
        direction = "desc"

    rows = (
        ProvinceCODItem.objects
        .select_related(
            "batch",
            "batch__shipper",
            "order",
            "order__seller",
            "received_confirmed_by",
            "paid_confirmed_by",
            "returned_confirmed_by",
            "seller_settled_by",
        )
        .exclude(batch__status=ProvinceCODBatch.STATUS_CANCELLED)
        .annotate(
            activity_date=Coalesce("sent_at", "batch__created_at"),
        )
    )

    if date_from:
        rows = rows.filter(activity_date__date__gte=date_from)

    if date_to:
        rows = rows.filter(activity_date__date__lte=date_to)

    if status == "PENDING":
        rows = rows.filter(cod_status="")
    elif status:
        rows = rows.filter(cod_status=status)

    if settlement == "SETTLED":
        rows = rows.filter(seller_settled=True)
    elif settlement == "UNSETTLED":
        rows = rows.filter(seller_settled=False)

    if seller_id.isdigit():
        rows = rows.filter(order__seller_id=int(seller_id))

    if shipper_id.isdigit():
        rows = rows.filter(batch__shipper_id=int(shipper_id))

    if q:
        rows = rows.filter(
            Q(order__tracking_no__icontains=q)
            | Q(order__receiver_name__icontains=q)
            | Q(order__receiver_phone__icontains=q)
            | Q(order__seller__name__icontains=q)
            | Q(batch__shipper__name__icontains=q)
            | Q(carrier_reference__icontains=q)
            | Q(received_person__icontains=q)
            | Q(return_reason__icontains=q)
            | Q(note__icontains=q)
        )

    order_field = SORT_MAP[sort]
    if direction == "desc":
        order_field = f"-{order_field}"

    return list(rows.order_by(order_field, "-id")), {
        "date_from": date_from,
        "date_to": date_to,
        "status": status,
        "settlement": settlement,
        "seller_id": seller_id,
        "shipper_id": shipper_id,
        "q": q,
        "sort": sort,
        "direction": direction,
    }


def export_province_cod_report_xlsx(request):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse(
            "Excel export needs openpyxl. Run: pip install openpyxl",
            content_type="text/plain; charset=utf-8",
            status=500,
        )

    rows, filters = _filtered_rows(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Province COD Report"
    ws.sheet_view.showGridLines = False

    green = "119A62"
    green_dark = "0A6D3F"
    white = "FFFFFF"
    light_green = "EAF8F0"
    light_blue = "DBEAFE"
    light_yellow = "FEF3C7"
    light_red = "FEE2E2"
    light_gray = "F1F5F9"
    border_color = "CBD5E1"

    title_fill = PatternFill("solid", fgColor=green_dark)
    header_fill = PatternFill("solid", fgColor=green)
    total_fill = PatternFill("solid", fgColor=light_green)
    thin = Side(style="thin", color=border_color)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title_font = Font(color=white, bold=True, size=18)
    white_bold = Font(color=white, bold=True)
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    # Only the requested columns are included in the downloaded Excel file.
    headers = [
        "No",
        "Sent Date",
        "DS Tracking",
        "Customer Order Code",
        "Seller",
        "Receiver",
        "Phone",
        "Address",
        "Original COD",
        "Real COD",
        "COD Status",
        "REASON",
    ]

    last_col = len(headers)
    last_col_letter = get_column_letter(last_col)

    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=last_col,
    )
    title_cell = ws.cell(
        row=1,
        column=1,
        value="DS EXPRESS - Province COD Report",
    )
    title_cell.fill = title_fill
    title_cell.font = title_font
    title_cell.alignment = center
    ws.row_dimensions[1].height = 30

    filter_parts = [
        f"Date: {filters['date_from'] or 'All'} to {filters['date_to'] or 'All'}",
        f"Status: {filters['status'] or 'All'}",
        f"Settlement: {filters['settlement'] or 'All'}",
        f"Search: {filters['q'] or '-'}",
        f"Sort: {filters['sort']} {filters['direction']}",
    ]
    ws.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=last_col,
    )
    filter_cell = ws.cell(
        row=2,
        column=1,
        value=" | ".join(filter_parts),
    )
    filter_cell.alignment = left
    filter_cell.font = Font(size=10, italic=True, color="475569")

    ws.merge_cells(
        start_row=3,
        start_column=1,
        end_row=3,
        end_column=last_col,
    )
    exported_cell = ws.cell(
        row=3,
        column=1,
        value=(
            f"Exported: {timezone.localtime().strftime('%Y-%m-%d %H:%M')}"
            f" | Orders: {len(rows)}"
        ),
    )
    exported_cell.alignment = left
    exported_cell.font = Font(size=10, color="475569")

    header_row = 5
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = white_bold
        cell.border = border
        cell.alignment = center

    widths = {
        1: 7,    # No
        2: 18,   # Sent Date
        3: 22,   # DS Tracking
        4: 24,   # Customer Order Code
        5: 20,   # Seller
        6: 22,   # Receiver
        7: 16,   # Phone
        8: 42,   # Address
        9: 15,   # Original COD
        10: 15,  # Real COD
        11: 20,  # COD Status
        12: 35,  # Reason
    }
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    data_start = header_row + 1

    status_fills = {
        "PENDING": PatternFill("solid", fgColor=light_gray),
        "SENT": PatternFill("solid", fgColor=light_blue),
        "AT_STATION": PatternFill("solid", fgColor=light_yellow),
        "OUT_FOR_DELIVERY": PatternFill("solid", fgColor=light_yellow),
        "DELIVERY_ISSUE": PatternFill("solid", fgColor=light_red),
        "RECEIVED": PatternFill("solid", fgColor=light_yellow),
        "PAID": PatternFill("solid", fgColor=light_green),
        "RETURNING": PatternFill("solid", fgColor=light_red),
        "RETURN_RECEIVED": PatternFill("solid", fgColor=light_green),
        "RETURNED": PatternFill("solid", fgColor=light_red),
    }

    for index, item in enumerate(rows, start=1):
        row_no = data_start + index - 1
        order = item.order
        seller = getattr(order, "seller", None)
        status = item.cod_status or "PENDING"

        values = [
            index,
            _date_time(item.activity_date),
            _display(getattr(order, "tracking_no", ""), ""),
            _customer_order_code(order),
            _display(getattr(seller, "name", "")),
            _display(getattr(order, "receiver_name", "")),
            _display(getattr(order, "receiver_phone", "")),
            _display(getattr(order, "receiver_address", "")),
            float(_money(item.original_cod)),
            float(_real_cod(item)),
            status,
            _reason(item),
        ]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_no, column=col, value=value)
            cell.border = border
            cell.alignment = left

        # Center: No, dates, codes, phone and status.
        for col in (1, 2, 3, 4, 7, 11):
            ws.cell(row=row_no, column=col).alignment = center

        # Money formatting.
        for col in (9, 10):
            ws.cell(row=row_no, column=col).alignment = right
            ws.cell(row=row_no, column=col).number_format = '$#,##0.00'

        # Wrap long address/reason text.
        ws.cell(row=row_no, column=8).alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
        )
        ws.cell(row=row_no, column=12).alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
        )

        status_cell = ws.cell(row=row_no, column=11)
        status_cell.fill = status_fills.get(
            status,
            PatternFill("solid", fgColor=white),
        )
        status_cell.font = bold

    data_end = data_start + len(rows) - 1
    total_row = data_end + 1 if rows else data_start

    ws.cell(row=total_row, column=1, value="TOTAL").font = bold
    ws.merge_cells(
        start_row=total_row,
        start_column=1,
        end_row=total_row,
        end_column=8,
    )

    for col in range(1, last_col + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = total_fill
        cell.border = border

    # Total only Original COD and Real COD.
    for col in (9, 10):
        if rows:
            letter = get_column_letter(col)
            value = f"=SUM({letter}{data_start}:{letter}{data_end})"
        else:
            value = 0

        total_cell = ws.cell(row=total_row, column=col, value=value)
        total_cell.font = bold
        total_cell.alignment = right
        total_cell.number_format = '$#,##0.00'

    ws.auto_filter.ref = (
        f"A{header_row}:{last_col_letter}"
        f"{data_end if rows else header_row}"
    )
    ws.freeze_panes = f"A{data_start}"
    ws.print_title_rows = f"1:{header_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:{last_col_letter}{total_row}"

    filename_date = timezone.localdate().strftime("%Y%m%d")
    if filters["date_from"] or filters["date_to"]:
        filename_date = (
            f"{filters['date_from'] or 'start'}_to_"
            f"{filters['date_to'] or 'today'}"
        )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = (
        f'attachment; filename="province_cod_report_{filename_date}.xlsx"'
    )
    return response
