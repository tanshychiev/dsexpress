from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def export_shipper_commission_excel(report, date_from="", date_to=""):
    wb = Workbook()
    ws = wb.active
    ws.title = "Commission Report"

    blue_fill = PatternFill("solid", fgColor="DDEBFF")
    head_fill = PatternFill("solid", fgColor="1F4E79")
    total_fill = PatternFill("solid", fgColor="F2F2F2")
    red_fill = PatternFill("solid", fgColor="FFE5E5")

    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    red_font = Font(color="C00000", bold=True)

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row_no = 1

    ws.cell(row=row_no, column=1, value="Shipper Commission Report").font = Font(bold=True, size=16)
    row_no += 1
    ws.cell(row=row_no, column=1, value=f"From: {date_from or '-'}   To: {date_to or '-'}")
    row_no += 1
    ws.cell(row=row_no, column=1, value="Morning = 12:00 AM - 11:59 AM | Afternoon = 12:00 PM - 11:59 PM")
    row_no += 2

    headers = [
        "#",
        "Date",
        "Morning Assign",
        "Afternoon Assign",
        "Done Morning",
        "Done Afternoon",
        "Total Done PC",
        "Commission (KHR)",
    ]

    for grp in report.get("shipper_groups", []):
        ws.cell(row=row_no, column=1, value=grp["shipper_name"]).font = Font(bold=True, size=13)
        for c in range(1, 9):
            ws.cell(row=row_no, column=c).fill = blue_fill
            ws.cell(row=row_no, column=c).border = border
        row_no += 1

        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=row_no, column=i, value=h)
            cell.fill = head_fill
            cell.font = white_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        row_no += 1

        for idx, r in enumerate(grp.get("rows", []), start=1):
            values = [
                idx,
                str(r["date"]),
                r["morning_assign"],
                r["afternoon_assign"],
                r["done_morning"],
                r["done_afternoon"],
                r["total_done_pc"],
                r["commission_khr"],
            ]

            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row_no, column=col, value=val)
                cell.border = border
                cell.alignment = Alignment(horizontal="center" if col in (1, 2) else "right")

            if r.get("is_all_zero"):
                for c in range(1, 9):
                    ws.cell(row=row_no, column=c).fill = red_fill
                    ws.cell(row=row_no, column=c).font = red_font
            else:
                if int(r["morning_assign"] or 0) == 0:
                    ws.cell(row=row_no, column=3).fill = red_fill
                    ws.cell(row=row_no, column=3).font = red_font

                if int(r["afternoon_assign"] or 0) == 0:
                    ws.cell(row=row_no, column=4).fill = red_fill
                    ws.cell(row=row_no, column=4).font = red_font

            row_no += 1

        t = grp["shipper_total"]
        total_values = [
            "-",
            "Total",
            t["morning_assign"],
            t["afternoon_assign"],
            t["done_morning"],
            t["done_afternoon"],
            t["total_done_pc"],
            t["commission_khr"],
        ]
        for col, val in enumerate(total_values, start=1):
            cell = ws.cell(row=row_no, column=col, value=val)
            cell.border = border
            cell.fill = total_fill
            cell.font = bold_font
            cell.alignment = Alignment(horizontal="center" if col in (1, 2) else "right")

        row_no += 2

    ws.cell(row=row_no, column=1, value="Grand Total").font = Font(bold=True, size=13)
    for c in range(1, 8):
        ws.cell(row=row_no, column=c).fill = blue_fill
        ws.cell(row=row_no, column=c).border = border
    row_no += 1

    summary_headers = [
        "Summary",
        "Morning Assign",
        "Afternoon Assign",
        "Done Morning",
        "Done Afternoon",
        "Total Done PC",
        "Commission (KHR)",
    ]
    for i, h in enumerate(summary_headers, start=1):
        cell = ws.cell(row=row_no, column=i, value=h)
        cell.fill = head_fill
        cell.font = white_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    row_no += 1

    gt = report.get("grand_total", {})
    grand_values = [
        "Grand Total",
        gt.get("morning_assign", 0),
        gt.get("afternoon_assign", 0),
        gt.get("done_morning", 0),
        gt.get("done_afternoon", 0),
        gt.get("total_done_pc", 0),
        gt.get("commission_khr", 0),
    ]
    for col, val in enumerate(grand_values, start=1):
        cell = ws.cell(row=row_no, column=col, value=val)
        cell.border = border
        cell.fill = total_fill
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="left" if col == 1 else "right")

    widths = {
        "A": 10,
        "B": 18,
        "C": 18,
        "D": 18,
        "E": 16,
        "F": 16,
        "G": 16,
        "H": 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="shipper_commission_report.xlsx"'
    return response