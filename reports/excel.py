import re
from decimal import Decimal
from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


def _to_decimal(value):
    if value in [None, ""]:
        return Decimal("0.00")
    try:
        return Decimal(str(value))
    except Exception:
        return Decimal("0.00")


def export_delivery_report_xlsx(
    grouped_data,
    report_title,
    classify_row,
    calc_totals,
    d_from=None,
    d_to=None,
    filename_prefix=None,
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Delivery Report"

    blue_fill = PatternFill("solid", fgColor="1F4E79")
    light_fill = PatternFill("solid", fgColor="F8FAFC")
    pending_fill = PatternFill("solid", fgColor="FFF2CC")
    returned_fill = PatternFill("solid", fgColor="C6EFCE")

    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    title_font = Font(bold=True, size=16)
    sub_title_font = Font(size=11)
    shop_font = Font(bold=True, size=12)

    thin = Side(style="thin", color="000000")
    all_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    headers = [
        "No",
        "Code",
        "Shipment ID",
        "Seller",
        "Pickup Date",
        "Delivery Date",
        "Location",
        "Description",
        "Qty",
        "Phone Number",
        "Receiver Name",
        "Price",
        "Delivery Fee",
        "Addition Fee",
        "Total Fee",
        "COD",
        "Status",
        "Shipper Name",
        "Reason",
    ]

    widths = {
        1: 6,
        2: 16,
        3: 22,
        4: 18,
        5: 14,
        6: 14,
        7: 28,
        8: 22,
        9: 6,
        10: 16,
        11: 16,
        12: 10,
        13: 12,
        14: 12,
        15: 12,
        16: 10,
        17: 12,
        18: 16,
        19: 20,
    }

    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    row_no = 1

    for seller_key, rows in grouped_data.items():
        # ===== top title =====
        ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=19)
        c = ws.cell(row=row_no, column=1, value="Delivery Report")
        c.font = title_font
        c.alignment = center
        row_no += 1

        date_from_display = d_from.strftime("%Y-%m-%d") if d_from else ""
        date_to_display = d_to.strftime("%Y-%m-%d") if d_to else ""
        delivered_text = f"Delivered: {date_from_display} → {date_to_display}"

        ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=19)
        c = ws.cell(row=row_no, column=1, value=delivered_text)
        c.font = sub_title_font
        c.alignment = center
        row_no += 1

        ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=19)
        c = ws.cell(row=row_no, column=1, value=f"Shop: {seller_key}")
        c.font = shop_font
        c.alignment = left
        row_no += 1

        # ===== header =====
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_no, column=col, value=header)
            cell.fill = blue_fill
            cell.font = white_font
            cell.alignment = center
            cell.border = all_border
        row_no += 1

        # ===== data rows =====
        for idx, o in enumerate(rows, start=1):
            row_type = classify_row(o)

            if row_type == "done_return":
                fill = returned_fill
                delivery_fee = Decimal("0.00")
                addition_fee = Decimal("0.00")
                total_fee = Decimal("0.00")
                cod = Decimal("0.00")
                report_status = "DONE RETURN"
            elif row_type == "done":
                fill = None
                delivery_fee = _to_decimal(getattr(o, "delivery_fee", 0))
                addition_fee = _to_decimal(getattr(o, "additional_fee", 0))
                total_fee = delivery_fee + addition_fee
                cod = _to_decimal(getattr(o, "cod", 0))
                report_status = "DONE"
            else:
                fill = pending_fill
                delivery_fee = Decimal("0.00")
                addition_fee = Decimal("0.00")
                total_fee = Decimal("0.00")
                cod = Decimal("0.00")
                report_status = "PENDING"

            created_at = getattr(o, "created_at", None)
            done_at = getattr(o, "done_at", None)

            values = [
                idx,
                getattr(o, "seller_order_code", "") or "",
                getattr(o, "tracking_no", "") or "",
                getattr(o, "seller_name", "") or (o.seller.name if getattr(o, "seller", None) else ""),
                created_at.strftime("%Y-%m-%d") if created_at else "",
                done_at.strftime("%Y-%m-%d") if hasattr(done_at, "strftime") else (done_at or ""),
                getattr(o, "receiver_address", "") or "",
                getattr(o, "product_desc", "") or "",
                getattr(o, "quantity", "") or "",
                getattr(o, "receiver_phone", "") or "",
                getattr(o, "receiver_name", "") or "",
                float(_to_decimal(getattr(o, "price", 0))),
                float(delivery_fee),
                float(addition_fee),
                float(total_fee),
                float(cod),
                report_status,
                o.delivery_shipper.name if getattr(o, "delivery_shipper", None) else "",
                getattr(o, "reason", "") or "",
            ]

            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row_no, column=col, value=value)
                cell.border = all_border

                if col in [1, 5, 6, 9, 17]:
                    cell.alignment = center
                elif col in [12, 13, 14, 15, 16]:
                    cell.alignment = right
                else:
                    cell.alignment = left

                if fill:
                    cell.fill = fill

            row_no += 1

        # ===== totals box on far right under shipper/reason =====
        total_cod, total_fee, pay = calc_totals(rows)

        box_start_row = row_no
        box_col_1 = 18  # Shipper Name
        box_col_2 = 19  # Reason

        total_items = [
            ("COD", float(_to_decimal(total_cod))),
            ("FEE", float(_to_decimal(total_fee))),
            ("PAY", float(_to_decimal(pay))),
        ]

        for i, (label, value) in enumerate(total_items):
            r = box_start_row + i

            c1 = ws.cell(row=r, column=box_col_1, value=label)
            c1.font = bold_font
            c1.fill = light_fill
            c1.border = all_border
            c1.alignment = center

            c2 = ws.cell(row=r, column=box_col_2, value=value)
            c2.font = bold_font
            c2.fill = light_fill
            c2.border = all_border
            c2.alignment = right

        row_no += 5

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    shop_name = "All_Shops"
    if grouped_data:
        first_key = list(grouped_data.keys())[0]
        if first_key:
            shop_name = str(first_key).strip()

    base_name = filename_prefix or shop_name or "shop"
    safe_name = re.sub(r'[\\/*?:"<>|]+', "_", str(base_name).strip())

    date_from_txt = d_from.strftime("%Y-%m-%d") if d_from else "no_from"
    date_to_txt = d_to.strftime("%Y-%m-%d") if d_to else "no_to"

    filename = f"{safe_name}_{date_from_txt}_to_{date_to_txt}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    return response