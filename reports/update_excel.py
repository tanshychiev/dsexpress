from datetime import datetime

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADERS = [
    "Tracking No",
    "Shop Code",
    "Shop Name",
    "Seller Name",
    "Seller Order Code",
    "Receiver Name",
    "Receiver Phone",
    "Receiver Address",
    "Product Desc",
    "Qty",
    "Price",
    "Delivery Fee",
    "Additional Fee",
    "COD",
    "Status",
    "Reason",
    "Delivery Shipper",
]


def _shop_code(order):
    seller = getattr(order, "seller", None)
    if seller:
        return getattr(seller, "code", "") or ""
    return ""


def _shop_name(order):
    seller = getattr(order, "seller", None)
    if seller:
        return getattr(seller, "name", "") or ""
    return ""


def _seller_name(order):
    """
    Seller Name here means the order-level seller/staff/name text,
    for example: Ngoc / An / Labut.
    Shop Name means the masterdata Seller shop/business name.
    """
    return getattr(order, "seller_name", "") or ""


def _shipper_name(order):
    shipper = getattr(order, "delivery_shipper", None)
    if shipper:
        return getattr(shipper, "name", "") or ""
    return ""


def export_update_template_xlsx(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Update Orders"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    # Header
    for col, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border

    # Data
    for i, o in enumerate(rows, start=2):
        values = [
            getattr(o, "tracking_no", "") or "",
            _shop_code(o),
            _shop_name(o),
            _seller_name(o),
            getattr(o, "seller_order_code", "") or "",
            getattr(o, "receiver_name", "") or "",
            getattr(o, "receiver_phone", "") or "",
            getattr(o, "receiver_address", "") or "",
            getattr(o, "product_desc", "") or "",
            getattr(o, "quantity", 0) or 0,
            float(getattr(o, "price", 0) or 0),
            float(getattr(o, "delivery_fee", 0) or 0),
            float(getattr(o, "additional_fee", 0) or 0),
            float(getattr(o, "cod", 0) or 0),
            getattr(o, "status", "") or "",
            getattr(o, "reason", "") or "",
            _shipper_name(o),
        ]

        for col, value in enumerate(values, start=1):
            c = ws.cell(row=i, column=col, value=value)
            c.border = border

            if col in [1, 2, 10, 15]:
                c.alignment = center
            elif col in [11, 12, 13, 14]:
                c.alignment = right
            else:
                c.alignment = left

    widths = [
        20,  # Tracking No
        16,  # Shop Code
        26,  # Shop Name
        18,  # Seller Name
        18,  # Seller Order Code
        18,  # Receiver Name
        16,  # Receiver Phone
        30,  # Receiver Address
        22,  # Product Desc
        8,   # Qty
        12,  # Price
        12,  # Delivery Fee
        14,  # Additional Fee
        12,  # COD
        18,  # Status
        24,  # Reason
        20,  # Delivery Shipper
    ]

    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    filename = f"delivery_report_update_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response