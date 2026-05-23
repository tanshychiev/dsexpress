import json
import tempfile
from decimal import Decimal
from pathlib import Path

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render
from django.template.loader import render_to_string
from django.utils import timezone
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

from masterdata.models import Seller
from orders.models import Order
from provinceops.models import ProvinceBatchItem

from .excel import export_delivery_report_xlsx
from .forms import DeliveryReportFilterForm, DeliveryReportUploadForm
from .services import (
    get_done_queryset,
    get_pending_queryset,
    group_by_seller,
    classify_row,
    calc_totals,
    get_shipper_name,
    report_money,
)
from .update_excel import export_update_template_xlsx


def apply_keyword_filter(rows, keyword):
    if not keyword:
        return rows

    keyword = keyword.strip().lower()
    filtered = []

    for o in rows:
        values = [
            getattr(o, "tracking_no", ""),
            getattr(o, "seller_order_code", ""),
            getattr(o, "seller_name", ""),
            getattr(o, "receiver_name", ""),
            getattr(o, "receiver_phone", ""),
            getattr(o, "receiver_address", ""),
            getattr(o, "product_desc", ""),
            getattr(o, "reason", ""),
            getattr(o, "report_shipper_name", ""),
        ]

        seller_obj = getattr(o, "seller", None)
        if seller_obj:
            values.append(getattr(seller_obj, "name", ""))
            values.append(getattr(seller_obj, "code", ""))

        haystack = " ".join([str(v or "") for v in values]).lower()
        if keyword in haystack:
            filtered.append(o)

    return filtered


def apply_status_filter(rows, status_filter):
    if not status_filter:
        return rows

    filtered = []
    for o in rows:
        row_type = classify_row(o)

        if status_filter == "DONE" and row_type == "done":
            filtered.append(o)
        elif status_filter == "PENDING" and row_type == "pending":
            filtered.append(o)
        elif status_filter == "DONE_RETURN" and row_type == "done_return":
            filtered.append(o)

    return filtered


def build_top_summary(rows, seller_count=0):
    total_done = 0
    total_pending = 0
    total_return = 0

    for o in rows:
        row_type = classify_row(o)
        if row_type == "done":
            total_done += 1
        elif row_type == "pending":
            total_pending += 1
        elif row_type == "done_return":
            total_return += 1

    total_cod, total_fee, total_pay = calc_totals(rows)

    return {
        "total_sent": len(rows),
        "total_done": total_done,
        "total_pending": total_pending,
        "total_return": total_return,
        "total_cod": total_cod,
        "total_fee": total_fee,
        "total_pay": total_pay,
        "total_selected_shops": seller_count,
    }


def _build_province_date_map(rows):
    """
    Runtime helper only. Does not save DB.

    Province orders may not have done_at.
    Province delivery date should come from ProvinceBatch.assigned_at.
    """
    order_ids = [o.id for o in rows if getattr(o, "id", None)]
    if not order_ids:
        return {}

    province_items = (
        ProvinceBatchItem.objects
        .select_related("batch")
        .filter(order_id__in=order_ids)
        .order_by("-batch__assigned_at", "-id")
    )

    province_date_map = {}
    for it in province_items:
        if it.order_id not in province_date_map:
            province_date_map[it.order_id] = getattr(it.batch, "assigned_at", None)

    return province_date_map


def enrich_report_rows(rows):
    """
    Add runtime-only helpers for template display.
    Does not save to DB.

    Important fix:
    Existing report HTML may still use o.done_at or o.delivery_date.
    So we set those runtime-only fallback values too.
    """
    from provinceops.models import ProvinceBatch, ProvinceBatchItem

    order_ids = [o.id for o in rows if getattr(o, "id", None)]

    province_date_map = {}
    if order_ids:
        province_items = (
            ProvinceBatchItem.objects
            .select_related("batch")
            .filter(
                order_id__in=order_ids,
                batch__status=ProvinceBatch.STATUS_DONE,
                batch__assigned_at__isnull=False,
            )
            .order_by("-batch__assigned_at", "-id")
        )

        for it in province_items:
            if it.order_id not in province_date_map:
                province_date_map[it.order_id] = getattr(it.batch, "assigned_at", None)

    for o in rows:
        o.report_shipper_name = get_shipper_name(o)

        money = report_money(o)
        o.report_delivery_fee = money.get("delivery_fee", 0)
        o.report_additional_fee = money.get("additional_fee", 0)
        o.report_province_fee = money.get("province_fee", 0)
        o.report_total_fee = money.get("total_fee", 0)
        o.report_cod = money.get("cod", 0)

        province_date = province_date_map.get(getattr(o, "id", None))
        original_done_at = getattr(o, "done_at", None)

        final_delivery_date = original_done_at or province_date

        # New helper fields
        o.report_pickup_date = getattr(o, "created_at", None)
        o.report_delivery_date = final_delivery_date

        # ✅ Important: old HTML compatibility
        # This only changes the Python object in memory. It does NOT save DB.
        if not original_done_at and province_date:
            o.done_at = province_date

        # ✅ If old HTML uses delivery_date, make it work too
        if not getattr(o, "delivery_date", None):
            o.delivery_date = final_delivery_date

        if not getattr(o, "pickup_date", None):
            o.pickup_date = getattr(o, "created_at", None)

    return rows


@login_required
def delivery_report(request):
    now = timezone.localtime()
    today_str = now.strftime("%Y-%m-%d")

    data = request.GET.copy()

    if not data.get("delivery_date_from"):
        data["delivery_date_from"] = f"{today_str}T00:00"

    if not data.get("delivery_date_to"):
        data["delivery_date_to"] = f"{today_str}T23:59"

    if not data.get("pending_date_from"):
        data["pending_date_from"] = "2020-01-01T00:00"

    if not data.get("pending_date_to"):
        data["pending_date_to"] = f"{today_str}T23:59"

    form = DeliveryReportFilterForm(data)

    seller_id = (request.GET.get("seller") or "").strip()
    selected_seller = None
    selected_seller_name = ""

    if seller_id:
        selected_seller = Seller.objects.filter(id=seller_id).first()
        if selected_seller:
            selected_seller_name = selected_seller.name

    seller_options = list(
        Seller.objects.order_by("name").values("id", "name")
    )

    action = request.GET.get("action", "").strip()
    shop_key = request.GET.get("shop_key", "").strip()

    show_results = action in ["show", "export", "export_selected", "download_update"]

    selected_shop_keys_raw = request.GET.get("selected_shop_keys", "")
    try:
        selected_shop_keys = json.loads(selected_shop_keys_raw) if selected_shop_keys_raw else []
    except Exception:
        selected_shop_keys = []

    report_title = "Delivery Report"
    mode = "EMPTY"
    grouped = {}

    d_from = None
    d_to = None
    p_from = None
    p_to = None

    top_summary = {
        "total_sent": 0,
        "total_done": 0,
        "total_pending": 0,
        "total_return": 0,
        "total_cod": 0,
        "total_fee": 0,
        "total_pay": 0,
        "total_selected_shops": 0,
    }

    seller_summaries = []

    if form.is_valid() and show_results:
        seller = form.cleaned_data.get("seller")
        seller_label = seller.name if seller else "All Shops"

        d_from = form.cleaned_data.get("delivery_date_from")
        d_to = form.cleaned_data.get("delivery_date_to")
        p_from = form.cleaned_data.get("pending_date_from")
        p_to = form.cleaned_data.get("pending_date_to")

        keyword = (
            form.cleaned_data.get("keyword")
            or form.cleaned_data.get("search")
            or ""
        )
        status_filter = form.cleaned_data.get("status_filter")

        done_rows = Order.objects.none()
        pending_rows = Order.objects.none()

        if d_from or d_to:
            done_rows = get_done_queryset(Order, form.cleaned_data)

        if p_from or p_to:
            pending_rows = get_pending_queryset(Order, form.cleaned_data)

        if (d_from or d_to) and (p_from or p_to):
            mode = "DONE_PENDING"
            report_title = f"Report to {seller_label}"
        elif d_from or d_to:
            mode = "DONE"
            report_title = f"Report to {seller_label}"
        elif p_from or p_to:
            mode = "PENDING"
            report_title = f"Report to {seller_label}"
        else:
            mode = "EMPTY"

        rows = list(done_rows) + list(pending_rows)
        rows = enrich_report_rows(rows)
        rows = apply_keyword_filter(rows, keyword)
        rows = apply_status_filter(rows, status_filter)

        grouped = group_by_seller(rows)

        if shop_key:
            grouped = {
                k: v
                for k, v in grouped.items()
                if str(k).strip() == shop_key
            }

        if selected_shop_keys:
            selected_set = {str(x).strip() for x in selected_shop_keys}
            grouped = {
                k: v
                for k, v in grouped.items()
                if str(k).strip() in selected_set
            }

        filtered_rows = []
        for _, seller_rows in grouped.items():
            filtered_rows.extend(seller_rows)

        top_summary = build_top_summary(filtered_rows, seller_count=len(grouped))

        if action == "export" and mode != "EMPTY":
            return export_delivery_report_xlsx(
                grouped,
                report_title,
                classify_row,
                calc_totals,
                d_from,
                d_to,
            )

        if action == "export_selected" and mode != "EMPTY":
            return export_delivery_report_xlsx(
                grouped,
                "Delivery Report",
                classify_row,
                calc_totals,
                d_from,
                d_to,
                filename_prefix=f"Selected_{len(grouped)}_Shops",
            )

        if action == "download_update" and mode != "EMPTY":
            return export_update_template_xlsx(filtered_rows)

        for seller_key_name, seller_rows in grouped.items():
            total_cod, total_fee, pay = calc_totals(seller_rows)
            seller_summaries.append({
                "seller_key": seller_key_name,
                "rows": seller_rows,
                "total_cod": total_cod,
                "total_fee": total_fee,
                "pay": pay,
                "total_sent": len(seller_rows),
                "total_done": len([o for o in seller_rows if classify_row(o) == "done"]),
                "total_pending": len([o for o in seller_rows if classify_row(o) == "pending"]),
                "total_return": len([o for o in seller_rows if classify_row(o) == "done_return"]),
            })

    return render(request, "reports/delivery_report.html", {
        "form": form,
        "report_title": report_title,
        "mode": mode,
        "seller_summaries": seller_summaries,
        "delivery_from": d_from,
        "delivery_to": d_to,
        "pending_from": p_from,
        "pending_to": p_to,
        "top_summary": top_summary,
        "show_results": show_results,
        "selected_seller": selected_seller,
        "selected_seller_name": selected_seller_name,
        "seller_options": seller_options,
    })


@login_required
def delivery_report_upload(request):
    form = DeliveryReportUploadForm(request.POST or None, request.FILES or None)
    summary = None

    def _clean_header(value):
        return str(value or "").strip().lower()

    def _get(row_map, *names, default=""):
        for name in names:
            key = _clean_header(name)
            if key in row_map:
                value = row_map.get(key)
                return default if value is None else value
        return default

    def _to_decimal(value):
        try:
            return Decimal(str(value or 0))
        except Exception:
            return Decimal("0")

    def _to_int(value):
        try:
            return int(value or 0)
        except Exception:
            return 0

    if request.method == "POST" and form.is_valid():
        f = form.cleaned_data["file"]
        wb = load_workbook(f)
        ws = wb.active

        total_rows = 0
        updated_rows = 0
        skipped_rows = 0
        not_found_rows = 0
        error_rows = []

        headers = []
        for cell in ws[1]:
            headers.append(_clean_header(cell.value))

        for excel_row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            total_rows += 1

            row_map = {}
            for idx, header in enumerate(headers):
                if header:
                    row_map[header] = row[idx] if idx < len(row) else None

            tracking_no = str(
                _get(row_map, "Tracking No", "tracking_no", default="") or ""
            ).strip()

            if not tracking_no:
                skipped_rows += 1
                error_rows.append(f"Row {excel_row_no}: without tracking number")
                continue

            try:
                order = Order.objects.filter(tracking_no=tracking_no).first()

                if not order:
                    not_found_rows += 1
                    error_rows.append(f"{tracking_no}: not found")
                    continue

                seller_name = str(
                    _get(row_map, "Seller Name", default=getattr(order, "seller_name", "") or "") or ""
                ).strip()

                if hasattr(order, "seller_name"):
                    order.seller_name = seller_name

                order.seller_order_code = str(
                    _get(
                        row_map,
                        "Seller Order Code",
                        default=getattr(order, "seller_order_code", "") or "",
                    ) or ""
                )

                order.receiver_name = str(_get(row_map, "Receiver Name", default="") or "")
                order.receiver_phone = str(_get(row_map, "Receiver Phone", default="") or "")
                order.receiver_address = str(_get(row_map, "Receiver Address", default="") or "")
                order.product_desc = str(_get(row_map, "Product Desc", default="") or "")

                order.quantity = _to_int(_get(row_map, "Qty", default=0))
                order.price = _to_decimal(_get(row_map, "Price", default=0))
                order.delivery_fee = _to_decimal(_get(row_map, "Delivery Fee", default=0))
                order.additional_fee = _to_decimal(_get(row_map, "Additional Fee", default=0))
                order.cod = _to_decimal(_get(row_map, "COD", default=0))

                status_value = str(_get(row_map, "Status", default="") or "").strip().upper()
                if status_value:
                    order.status = status_value

                order.reason = str(_get(row_map, "Reason", default="") or "")

                order.save()
                updated_rows += 1

            except Exception as e:
                skipped_rows += 1
                error_rows.append(f"{tracking_no}: {str(e)}")

        summary = {
            "total_rows": total_rows,
            "updated_rows": updated_rows,
            "skipped_rows": skipped_rows,
            "not_found_rows": not_found_rows,
            "errors": error_rows[:30],
        }

        messages.success(request, f"Upload complete. Updated {updated_rows} rows.")

    return render(request, "reports/delivery_report_upload.html", {
        "form": form,
        "summary": summary,
    })


@login_required
def delivery_report_png(request):
    form = DeliveryReportFilterForm(request.GET.copy())
    shop_key = request.GET.get("shop_key", "").strip()

    show_results = True
    grouped = {}
    seller_summaries = []
    report_title = "Delivery Report"
    mode = "EMPTY"

    d_from = None
    d_to = None
    p_from = None
    p_to = None

    top_summary = {
        "total_sent": 0,
        "total_done": 0,
        "total_pending": 0,
        "total_return": 0,
        "total_cod": 0,
        "total_fee": 0,
        "total_pay": 0,
        "total_selected_shops": 0,
    }

    if form.is_valid():
        cleaned = form.cleaned_data

        account = getattr(request.user, "account", None)
        if account and account.account_type == "seller" and account.seller and account.seller.is_active:
            cleaned["seller"] = account.seller

        seller = cleaned.get("seller")
        seller_label = seller.name if seller else "All Shops"

        d_from = cleaned.get("delivery_date_from")
        d_to = cleaned.get("delivery_date_to")
        p_from = cleaned.get("pending_date_from")
        p_to = cleaned.get("pending_date_to")

        keyword = (
            cleaned.get("keyword")
            or cleaned.get("search")
            or ""
        )
        status_filter = cleaned.get("status_filter")

        done_rows = Order.objects.none()
        pending_rows = Order.objects.none()

        if d_from or d_to:
            done_rows = get_done_queryset(Order, cleaned)

        if p_from or p_to:
            pending_rows = get_pending_queryset(Order, cleaned)

        if (d_from or d_to) and (p_from or p_to):
            mode = "DONE_PENDING"
            report_title = f"Report to {seller_label}"
        elif d_from or d_to:
            mode = "DONE"
            report_title = f"Report to {seller_label}"
        elif p_from or p_to:
            mode = "PENDING"
            report_title = f"Report to {seller_label}"

        rows = list(done_rows) + list(pending_rows)
        rows = enrich_report_rows(rows)
        rows = apply_keyword_filter(rows, keyword)
        rows = apply_status_filter(rows, status_filter)

        grouped = group_by_seller(rows)

        if shop_key:
            grouped = {
                k: v for k, v in grouped.items()
                if str(k).strip() == shop_key
            }

        filtered_rows = []
        for _, seller_rows in grouped.items():
            filtered_rows.extend(seller_rows)

        top_summary = build_top_summary(filtered_rows, seller_count=len(grouped))

        for seller_key_name, seller_rows in grouped.items():
            total_cod, total_fee, pay = calc_totals(seller_rows)
            seller_summaries.append({
                "seller_key": seller_key_name,
                "rows": seller_rows,
                "total_cod": total_cod,
                "total_fee": total_fee,
                "pay": pay,
                "total_sent": len(seller_rows),
                "total_done": len([o for o in seller_rows if classify_row(o) == "done"]),
                "total_pending": len([o for o in seller_rows if classify_row(o) == "pending"]),
                "total_return": len([o for o in seller_rows if classify_row(o) == "done_return"]),
            })

    html = render_to_string("reports/delivery_report_png.html", {
        "form": form,
        "report_title": report_title,
        "mode": mode,
        "seller_summaries": seller_summaries,
        "delivery_from": d_from,
        "delivery_to": d_to,
        "pending_from": p_from,
        "pending_to": p_to,
        "top_summary": top_summary,
        "show_results": show_results,
    }, request=request)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".html", mode="w", encoding="utf-8") as f:
        f.write(html)
        temp_html = f.name

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch()
            page = browser.new_page(viewport={"width": 1600, "height": 2200}, device_scale_factor=2)
            page.goto(Path(temp_html).as_uri(), wait_until="load")
            report = page.locator("#reportCapture")
            png_bytes = report.screenshot(type="png")
            browser.close()
    finally:
        Path(temp_html).unlink(missing_ok=True)

    download_name = request.GET.get("download_name", "").strip() or "report"
    response = HttpResponse(png_bytes, content_type="image/png")
    response["Content-Disposition"] = f'inline; filename="{download_name}.png"'
    return response


@login_required
def delivery_report_pdf(request):
    form = DeliveryReportFilterForm(request.GET.copy())
    shop_key = request.GET.get("shop_key", "").strip()

    show_results = True
    grouped = {}
    seller_summaries = []
    report_title = "Delivery Report"
    mode = "EMPTY"

    d_from = None
    d_to = None
    p_from = None
    p_to = None

    top_summary = {
        "total_sent": 0,
        "total_done": 0,
        "total_pending": 0,
        "total_return": 0,
        "total_cod": 0,
        "total_fee": 0,
        "total_pay": 0,
        "total_selected_shops": 0,
    }

    if form.is_valid():
        cleaned = form.cleaned_data

        account = getattr(request.user, "account", None)
        if account and account.account_type == "seller" and account.seller and account.seller.is_active:
            cleaned["seller"] = account.seller

        seller = cleaned.get("seller")
        seller_label = seller.name if seller else "All Shops"

        d_from = cleaned.get("delivery_date_from")
        d_to = cleaned.get("delivery_date_to")
        p_from = cleaned.get("pending_date_from")
        p_to = cleaned.get("pending_date_to")

        keyword = (
            cleaned.get("keyword")
            or cleaned.get("search")
            or ""
        )
        status_filter = cleaned.get("status_filter")

        done_rows = Order.objects.none()
        pending_rows = Order.objects.none()

        if d_from or d_to:
            done_rows = get_done_queryset(Order, cleaned)

        if p_from or p_to:
            pending_rows = get_pending_queryset(Order, cleaned)

        if (d_from or d_to) and (p_from or p_to):
            mode = "DONE_PENDING"
            report_title = f"Report to {seller_label}"
        elif d_from or d_to:
            mode = "DONE"
            report_title = f"Report to {seller_label}"
        elif p_from or p_to:
            mode = "PENDING"
            report_title = f"Report to {seller_label}"

        rows = list(done_rows) + list(pending_rows)
        rows = enrich_report_rows(rows)
        rows = apply_keyword_filter(rows, keyword)
        rows = apply_status_filter(rows, status_filter)

        grouped = group_by_seller(rows)

        if shop_key:
            grouped = {
                k: v for k, v in grouped.items()
                if str(k).strip() == shop_key
            }

        filtered_rows = []
        for _, seller_rows in grouped.items():
            filtered_rows.extend(seller_rows)

        top_summary = build_top_summary(filtered_rows, seller_count=len(grouped))

        for seller_key_name, seller_rows in grouped.items():
            total_cod, total_fee, pay = calc_totals(seller_rows)
            seller_summaries.append({
                "seller_key": seller_key_name,
                "rows": seller_rows,
                "total_cod": total_cod,
                "total_fee": total_fee,
                "pay": pay,
                "total_sent": len(seller_rows),
                "total_done": len([o for o in seller_rows if classify_row(o) == "done"]),
                "total_pending": len([o for o in seller_rows if classify_row(o) == "pending"]),
                "total_return": len([o for o in seller_rows if classify_row(o) == "done_return"]),
            })

    html = render_to_string("reports/delivery_report_png.html", {
        "form": form,
        "report_title": report_title,
        "mode": mode,
        "seller_summaries": seller_summaries,
        "delivery_from": d_from,
        "delivery_to": d_to,
        "pending_from": p_from,
        "pending_to": p_to,
        "top_summary": top_summary,
        "show_results": show_results,
    }, request=request)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".html", mode="w", encoding="utf-8") as f:
        f.write(html)
        temp_html = f.name

    temp_png = f"{temp_html}.png"
    temp_pdf = f"{temp_html}.pdf"

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch()
            page = browser.new_page(
                viewport={"width": 1600, "height": 2200},
                device_scale_factor=2
            )
            page.goto(Path(temp_html).as_uri(), wait_until="networkidle")
            page.screenshot(path=temp_png, full_page=True)
            browser.close()

        from PIL import Image

        image = Image.open(temp_png)

        if image.mode == "RGBA":
            bg = Image.new("RGB", image.size, (255, 255, 255))
            bg.paste(image, mask=image.split()[3])
            image = bg
        elif image.mode != "RGB":
            image = image.convert("RGB")

        image.save(temp_pdf, "PDF", resolution=100.0)

        with open(temp_pdf, "rb") as f:
            pdf_bytes = f.read()

    finally:
        Path(temp_html).unlink(missing_ok=True)
        Path(temp_png).unlink(missing_ok=True)
        Path(temp_pdf).unlink(missing_ok=True)

    download_name = request.GET.get("download_name", "").strip() or "report"

    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{download_name}.pdf"'
    return response