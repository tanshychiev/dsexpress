from __future__ import annotations

import base64
import io
import json
from datetime import date, datetime, time
from decimal import Decimal

import qrcode
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Count, F, Q
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from masterdata.models import Seller
from .activity import add_order_activity
from .audit import add_audit_log
from .models import (
    AuditLog,
    BulkUpdateBatch,
    BulkUpdateRow,
    ImportBatch,
    Order,
    OrderActivity,
    OrderSetting,
)

PER_PAGE = 50

# ============================================================
# EXCEL HEADERS (MUST MATCH DOWNLOAD FILES 100%)
# ============================================================
ORDER_EXCEL_HEADERS = [
    "ID",
    "Tracking",
    "Created At",
    "Shop",
    "Shop Code",
    "Seller Name",
    "Order Code",
    "Product Desc",
    "Qty",
    "Receiver Address",
    "Receiver Phone",
    "Receiver Name",
    "Price",
    "COD",
    "Delivery Fee",
    "Additional Fee",
    "Province Fee",
    "Remark",
    "Reason",
    "Status",
    "Printed",
]


# ============================================================
# Helpers
# ============================================================
def _cell(v) -> str:
    return "" if v is None else str(v).strip()


def _to_int(v, default=0) -> int:
    try:
        return int(float(str(v).strip()))
    except Exception:
        return default


def _to_decimal(v, default=0) -> Decimal:
    try:
        s = str(v).strip()
        if s == "":
            return Decimal(str(default))
        return Decimal(s)
    except Exception:
        return Decimal(str(default))


def _make_tracking_no(order_id: int) -> str:
    today = timezone.localdate().strftime("%Y%m%d")
    return f"DS{today}{order_id:06d}"


def _parse_date_safe(s: str | None) -> date | None:
    if not s:
        return None
    try:
        return date.fromisoformat(str(s).strip())
    except Exception:
        return None


def _parse_time_safe(s: str | None) -> time | None:
    if not s:
        return None
    try:
        return time.fromisoformat(str(s).strip())
    except Exception:
        return None


def _is_search_clicked(request: HttpRequest) -> bool:
    return (request.GET.get("search") or "") == "1"


def _dt_range_from_request(request: HttpRequest):
    fd = _parse_date_safe(request.GET.get("from_date") or request.GET.get("from"))
    ft = _parse_time_safe(request.GET.get("from_time"))
    td = _parse_date_safe(request.GET.get("to_date") or request.GET.get("to"))
    tt = _parse_time_safe(request.GET.get("to_time"))

    if not fd and not td:
        return None, None

    if fd and not ft:
        ft = time(0, 0)
    if td and not tt:
        tt = time(23, 59, 59)

    if fd and td:
        start_dt = datetime.combine(fd, ft or time(0, 0))
        end_dt = datetime.combine(td, tt or time(23, 59, 59))
    elif fd and not td:
        start_dt = datetime.combine(fd, ft or time(0, 0))
        end_dt = datetime.combine(fd, time(23, 59, 59))
    else:
        start_dt = datetime.combine(td, time(0, 0))
        end_dt = datetime.combine(td, tt or time(23, 59, 59))

    tz = timezone.get_current_timezone()
    return timezone.make_aware(start_dt, tz), timezone.make_aware(end_dt, tz)


def _get_setting() -> OrderSetting:
    setting = OrderSetting.objects.first()
    if not setting:
        setting = OrderSetting.objects.create(usd_to_khr=Decimal("4100"))
    return setting


def _make_qr_data_uri(text: str) -> str:
    text = (text or "").strip()
    if not text:
        return ""

    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=10,
        border=2,
    )
    qr.add_data(text)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    b64 = base64.b64encode(buf.getvalue()).decode("ascii")
    return f"data:image/png;base64,{b64}"


def _fmt_khr_no_decimal(v: Decimal | int) -> str:
    try:
        n = int(Decimal(v))
    except Exception:
        n = 0
    return f"{n:,}".replace(",", ".")


def _build_params(request: HttpRequest, drop: set[str] | None = None) -> str:
    drop = drop or set()
    items = []
    for k, v in request.GET.items():
        if k in drop:
            continue
        items.append(f"{k}={v}")
    return "&".join(items)


def _qs_orders_filtered(request: HttpRequest, require_search_click=True):
    if require_search_click and not _is_search_clicked(request):
        return Order.objects.none()

    qs = Order.objects.filter(is_deleted=False).select_related("seller").order_by("-id")

    start_dt, end_dt = _dt_range_from_request(request)
    if start_dt and end_dt:
        qs = qs.filter(created_at__gte=start_dt, created_at__lte=end_dt)

    tracking = (request.GET.get("tracking") or "").strip()
    order_code = (request.GET.get("order_code") or request.GET.get("seller_order_code") or "").strip()
    seller_id = (request.GET.get("seller") or "").strip()
    receiver_name = (request.GET.get("receiver_name") or "").strip()
    receiver_phone = (request.GET.get("receiver_phone") or "").strip()
    status_group = (request.GET.get("status") or "").strip()
    q = (request.GET.get("q") or "").strip()

    if tracking:
        qs = qs.filter(tracking_no__icontains=tracking)
    if order_code:
        qs = qs.filter(seller_order_code__icontains=order_code)
    if seller_id.isdigit():
        qs = qs.filter(seller_id=int(seller_id))
    if receiver_name:
        qs = qs.filter(receiver_name__icontains=receiver_name)
    if receiver_phone:
        qs = qs.filter(receiver_phone__icontains=receiver_phone)
    if q:
        qs = qs.filter(
            Q(tracking_no__icontains=q)
            | Q(receiver_name__icontains=q)
            | Q(receiver_phone__icontains=q)
            | Q(receiver_address__icontains=q)
            | Q(seller_order_code__icontains=q)
            | Q(seller__code__icontains=q)
            | Q(seller__name__icontains=q)
        )

    if status_group == "DONE":
        qs = qs.filter(status=Order.STATUS_DELIVERED)
    elif status_group == "PENDING":
        qs = qs.exclude(status=Order.STATUS_DELIVERED)

    return qs


def _excel_orders_response(qs, filename: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    ws.append(ORDER_EXCEL_HEADERS)

    def fmt_dt(dt):
        if not dt:
            return ""
        try:
            return timezone.localtime(dt).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            return str(dt)

    for o in qs:
        ws.append(
            [
                o.id,
                o.tracking_no,
                fmt_dt(o.created_at),
                o.seller.name if o.seller_id else "",
                o.seller.code if o.seller_id else "",
                o.seller_name or "",
                o.seller_order_code or "",
                o.product_desc or "",
                o.quantity,
                o.receiver_address or "",
                o.receiver_phone or "",
                o.receiver_name or "",
                float(o.price or 0),
                float(o.cod or 0),
                float(o.delivery_fee or 0),
                float(o.additional_fee or 0),
                float(o.province_fee or 0),
                o.remark or "",
                o.reason or "",
                o.status,
                int(o.print_count or 0),
            ]
        )

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    resp = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _soft_delete_queryset(qs, user):
    now = timezone.now()
    return qs.update(
        is_deleted=True,
        deleted_at=now,
        deleted_by=user,
    )


# ============================================================
# API: Seller autocomplete
# ============================================================
@login_required
def seller_autocomplete(request: HttpRequest):
    q = (request.GET.get("q") or "").strip()
    qs = Seller.objects.filter(is_active=True).order_by("name")
    if q:
        qs = qs.filter(Q(code__icontains=q) | Q(name__icontains=q) | Q(phone__icontains=q))
    qs = qs[:50]
    items = [{"id": s.id, "code": s.code or "", "name": s.name or "", "phone": s.phone or ""} for s in qs]
    return JsonResponse({"items": items})


# ============================================================
# Orders List + Download
# ============================================================
@login_required
def order_list(request: HttpRequest):
    sellers = Seller.objects.filter(is_active=True).order_by("name")
    qs = _qs_orders_filtered(request, require_search_click=True)

    paginator = Paginator(qs, PER_PAGE)
    page_obj = paginator.get_page(request.GET.get("page"))

    today = timezone.localdate().isoformat()
    params = _build_params(request, drop={"page"})

    return render(
        request,
        "orders/order_list.html",
        {
            "sellers": sellers,
            "page_obj": page_obj,
            "search": _is_search_clicked(request),
            "default_from": today,
            "default_to": today,
            "params": params,
        },
    )


@login_required
def download_orders_excel(request: HttpRequest):
    qs = _qs_orders_filtered(request, require_search_click=True)
    return _excel_orders_response(qs, "orders.xlsx")


# ============================================================
# Import
# ============================================================
@login_required
def download_import_sample_excel(request: HttpRequest):
    wb = Workbook()
    ws = wb.active
    ws.title = "Import Sample"

    headers = [
        "Customer CODE",
        "Order Coder",
        "Seller Name",
        "Description",
        "quantity",
        "COD",
        "Receiver name",
        "Phone",
        "Address",
        "Deliver Fee",
        "Additional",
        "Price",
        "Remark",
    ]
    ws.append(headers)
    ws.append(
        [
            "25000020",
            "SO-001",
            "Seller Name Example",
            "Product A",
            1,
            10,
            "Receiver Name",
            "012345678",
            "Phnom Penh...",
            0,
            0,
            10,
            "",
        ]
    )

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    resp = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="import_orders_sample.xlsx"'
    return resp


@login_required
def import_orders(request: HttpRequest):
    if request.method == "POST":
        f = request.FILES.get("file")
        errors: list[str] = []
        success_count = 0

        if not f:
            messages.error(request, "Please choose an Excel file.")
            return redirect("import_orders")

        try:
            wb = load_workbook(filename=f, data_only=True)
            ws = wb.active
        except Exception:
            messages.error(request, "Invalid Excel file.")
            return redirect("import_orders")

        headers = [_cell(c.value) for c in ws[1]]
        header_map = {h.strip().lower(): idx for idx, h in enumerate(headers)}

        required = [
            "customer code",
            "order coder",
            "seller name",
            "description",
            "quantity",
            "cod",
            "receiver name",
            "phone",
            "address",
            "deliver fee",
            "additional",
            "price",
            "remark",
        ]
        for r in required:
            if r not in header_map:
                errors.append(f"Missing column: {r}")

        def get(row, col: str) -> str:
            i = header_map.get(col)
            return _cell(row[i].value) if i is not None else ""

        prepared_rows = []

        if not errors:
            for row_idx in range(2, ws.max_row + 1):
                row = ws[row_idx]

                customer_code = get(row, "customer code")
                order_code = get(row, "order coder")
                seller_name = get(row, "seller name")
                desc = get(row, "description")
                qty_raw = get(row, "quantity")
                cod_raw = get(row, "cod")
                receiver_name = get(row, "receiver name")
                phone = get(row, "phone")
                address = get(row, "address")
                deliver_fee_raw = get(row, "deliver fee")
                additional_raw = get(row, "additional")
                price_raw = get(row, "price")
                remark = get(row, "remark")

                if all(
                    x == ""
                    for x in [
                        customer_code,
                        order_code,
                        seller_name,
                        desc,
                        qty_raw,
                        cod_raw,
                        receiver_name,
                        phone,
                        address,
                        deliver_fee_raw,
                        additional_raw,
                        price_raw,
                        remark,
                    ]
                ):
                    continue

                required_values = {
                    "Customer CODE": customer_code,
                    "Order Coder": order_code,
                    "Seller Name": seller_name,
                    "Description": desc,
                    "quantity": qty_raw,
                    "COD": cod_raw,
                    "Receiver name": receiver_name,
                    "Phone": phone,
                    "Address": address,
                    "Deliver Fee": deliver_fee_raw,
                    "Additional": additional_raw,
                    "Price": price_raw,
                }

                row_has_err = False
                for k, v in required_values.items():
                    if _cell(v) == "":
                        errors.append(f"Row {row_idx}: {k} is required")
                        row_has_err = True

                if row_has_err:
                    continue

                seller = Seller.objects.filter(code=customer_code).first()
                if not seller:
                    errors.append(f"Row {row_idx}: Customer code not existed: {customer_code}")
                    continue

                qty = _to_int(qty_raw, 0)
                if qty <= 0:
                    errors.append(f"Row {row_idx}: quantity must be number greater than 0")
                    continue

                cod = _to_decimal(cod_raw, 0)
                deliver_fee = _to_decimal(deliver_fee_raw, 0)
                additional_fee = _to_decimal(additional_raw, 0)
                price = _to_decimal(price_raw, 0)

                prepared_rows.append(
                    {
                        "seller": seller,
                        "seller_name": seller_name or None,
                        "seller_order_code": order_code or None,
                        "product_desc": desc or None,
                        "quantity": qty,
                        "cod": cod,
                        "delivery_fee": deliver_fee,
                        "additional_fee": additional_fee,
                        "price": price,
                        "receiver_name": receiver_name or None,
                        "receiver_phone": phone or None,
                        "receiver_address": address or None,
                        "remark": remark.strip() if remark.strip() else None,
                    }
                )

        if errors:
            request.session["import_errors"] = errors
            messages.error(request, f"❌ Import failed: {len(errors)} error(s). No data was uploaded.")
            return redirect("import_orders")

        try:
            with transaction.atomic():
                batch = ImportBatch.objects.create(filename=f.name)

                for idx, item in enumerate(prepared_rows, start=2):
                    o = Order.objects.create(
                        tracking_no=f"TEMP-{timezone.now().timestamp()}-{idx}",
                        seller=item["seller"],
                        seller_code=item["seller"].code or "",
                        seller_name=item["seller_name"],
                        seller_order_code=item["seller_order_code"],
                        product_desc=item["product_desc"],
                        quantity=item["quantity"],
                        cod=item["cod"],
                        delivery_fee=item["delivery_fee"],
                        additional_fee=item["additional_fee"],
                        price=item["price"],
                        receiver_name=item["receiver_name"],
                        receiver_phone=item["receiver_phone"],
                        receiver_address=item["receiver_address"],
                        remark=item["remark"],
                        import_batch=batch,
                        status=Order.STATUS_CREATED,
                    )
                    o.tracking_no = _make_tracking_no(o.id)
                    o.save(update_fields=["tracking_no"])
                    success_count += 1

                add_audit_log(
                    module=AuditLog.MODULE_IMPORT,
                    obj=batch,
                    action=AuditLog.ACTION_IMPORT,
                    user=request.user,
                    note=f"Imported {success_count} order(s) from {f.name}",
                )

            messages.success(request, f"✅ Import finished: {success_count} orders created.")
            return redirect("import_batch_detail", batch_id=batch.id)

        except Exception as e:
            messages.error(request, f"❌ Import failed. Nothing was uploaded. {str(e)}")
            return redirect("import_orders")

    show = _is_search_clicked(request)
    batch_qs = ImportBatch.objects.all().order_by("-id")

    if show:
        start_dt, end_dt = _dt_range_from_request(request)
        if start_dt and end_dt:
            batch_qs = batch_qs.filter(created_at__gte=start_dt, created_at__lte=end_dt)

        file_q = (request.GET.get("filename_contains") or request.GET.get("q") or "").strip()
        if file_q:
            batch_qs = batch_qs.filter(filename__icontains=file_q)

        batch_qs = batch_qs.annotate(
            total_orders=Count("orders", filter=Q(orders__is_deleted=False), distinct=True),
            total_shops=Count("orders__seller", filter=Q(orders__is_deleted=False), distinct=True),
        )
    else:
        batch_qs = ImportBatch.objects.none()

    total_batches = batch_qs.count() if show else 0

    if show:
        total_orders_sum = Order.objects.filter(
            import_batch__in=batch_qs.values_list("id", flat=True),
            is_deleted=False,
        ).count()
        total_shops_sum = Seller.objects.filter(
            orders__import_batch__in=batch_qs.values_list("id", flat=True),
            orders__is_deleted=False,
        ).distinct().count()
    else:
        total_orders_sum = 0
        total_shops_sum = 0

    today = timezone.localdate().isoformat()
    params = _build_params(request, drop={"page"})
    errors = request.session.pop("import_errors", [])

    return render(
        request,
        "orders/import_orders.html",
        {
            "errors": errors,
            "page_obj": batch_qs,
            "default_from": today,
            "default_to": today,
            "params": params,
            "search": show,
            "total_batches": total_batches,
            "total_orders_sum": total_orders_sum,
            "total_shops_sum": total_shops_sum,
        },
    )


@login_required
def import_batch_detail(request: HttpRequest, batch_id: int):
    batch = get_object_or_404(ImportBatch, id=batch_id)
    qs = Order.objects.filter(is_deleted=False, import_batch=batch).select_related("seller").order_by("-id")

    q = (request.GET.get("q") or "").strip()
    if q:
        qs = qs.filter(
            Q(tracking_no__icontains=q)
            | Q(receiver_phone__icontains=q)
            | Q(receiver_name__icontains=q)
            | Q(receiver_address__icontains=q)
            | Q(seller_order_code__icontains=q)
            | Q(seller__name__icontains=q)
        )

    total_orders = qs.count()
    total_shops = qs.exclude(seller_id__isnull=True).values("seller_id").distinct().count()
    params = _build_params(request, drop={"page"})

    return render(
        request,
        "orders/import_batch_detail.html",
        {
            "batch": batch,
            "page_obj": qs,
            "params": params,
            "total_orders": total_orders,
            "total_shops": total_shops,
        },
    )


@login_required
def download_import_batch_excel(request: HttpRequest, batch_id: int):
    batch = get_object_or_404(ImportBatch, id=batch_id)
    qs = Order.objects.filter(is_deleted=False, import_batch=batch).select_related("seller").order_by("-id")

    q = (request.GET.get("q") or "").strip()
    if q:
        qs = qs.filter(
            Q(tracking_no__icontains=q)
            | Q(receiver_phone__icontains=q)
            | Q(receiver_name__icontains=q)
            | Q(receiver_address__icontains=q)
            | Q(seller_order_code__icontains=q)
            | Q(seller__name__icontains=q)
        )

    return _excel_orders_response(qs, f"import_batch_{batch.id}.xlsx")


# ============================================================
# BULK UPDATE
# ============================================================
def _safe_json(v):
    if isinstance(v, Decimal):
        return str(v)
    return v


def _snapshot(order: Order) -> dict:
    return {
        "created_at": timezone.localtime(order.created_at).strftime("%Y-%m-%d %H:%M:%S") if order.created_at else "",
        "shop_code": order.seller.code if order.seller_id else "",
        "shop_name": order.seller.name if order.seller_id else "",
        "seller_name": order.seller_name or "",
        "order_code": order.seller_order_code or "",
        "product_desc": order.product_desc or "",
        "qty": order.quantity or 0,
        "price": str(order.price or 0),
        "cod": str(order.cod or 0),
        "delivery_fee": str(order.delivery_fee or 0),
        "additional_fee": str(order.additional_fee or 0),
        "province_fee": str(order.province_fee or 0),
        "receiver_name": order.receiver_name or "",
        "receiver_phone": order.receiver_phone or "",
        "receiver_address": order.receiver_address or "",
        "remark": order.remark or "",
        "reason": order.reason or "",
    }


def _parse_created_at_excel(value):
    if value in (None, ""):
        return None

    if isinstance(value, datetime):
        dt = value
    else:
        s = str(value).strip()
        if not s:
            return None

        dt = None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                dt = datetime.strptime(s, fmt)
                break
            except ValueError:
                continue

        if dt is None:
            return None

    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt, timezone.get_current_timezone())

    return dt


def _cell_str(cell):
    if cell is None or cell.value is None:
        return ""
    return str(cell.value).strip()


@login_required
def download_update_template(request: HttpRequest):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bulk Update"
    ws.append(ORDER_EXCEL_HEADERS)

    ws.append(
        [
            1,
            "DS20260209000001",
            "2026-02-09 12:00:00",
            "Shop name",
            "SHOP001",
            "Seller Name",
            "ORDER-001",
            "Product A",
            1,
            "Phnom Penh",
            "012345678",
            "Receiver Name",
            10,
            10,
            0,
            0,
            0,
            "Remark here",
            "Reason here",
            "CREATED",
            0,
        ]
    )

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    resp = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="bulk_update_template.xlsx"'
    return resp


@login_required
def bulk_update(request: HttpRequest):
    if request.method == "POST":
        excel_file = request.FILES.get("file")

        if not excel_file:
            messages.error(request, "Please choose an Excel file.")
            return redirect("bulk_update")

        try:
            wb = load_workbook(excel_file, data_only=True)
            ws = wb.active
        except Exception:
            messages.error(request, "Invalid Excel file.")
            return redirect("bulk_update")

        header_map = {}
        for idx, cell in enumerate(ws[1]):
            if cell.value:
                header_map[str(cell.value).strip().lower()] = idx

        required_headers = ["id", "tracking"]
        missing_headers = [h for h in required_headers if h not in header_map]
        if missing_headers:
            messages.error(request, f"Missing required columns: {', '.join(missing_headers)}")
            return redirect("bulk_update")

        batch = BulkUpdateBatch.objects.create(filename=excel_file.name)

        updated_count = 0
        skipped_count = 0
        error_count = 0
        errors: list[str] = []

        for row_idx in range(2, ws.max_row + 1):
            row_cells = list(ws[row_idx])

            try:
                id_cell = row_cells[header_map["id"]]
                tracking_cell = row_cells[header_map["tracking"]]

                order_id = int(id_cell.value) if id_cell and id_cell.value not in (None, "") else None
                tracking = _cell_str(tracking_cell)

                if not order_id or not tracking:
                    skipped_count += 1
                    errors.append(f"Row {row_idx}: Missing ID or Tracking")
                    BulkUpdateRow.objects.create(
                        batch=batch,
                        status="SKIPPED",
                    )
                    continue

                try:
                    order = Order.objects.get(id=order_id, tracking_no=tracking)
                except Order.DoesNotExist:
                    error_count += 1
                    errors.append(f"Row {row_idx}: Order not found (ID={order_id}, Tracking={tracking})")
                    BulkUpdateRow.objects.create(
                        batch=batch,
                        status="ERROR",
                    )
                    continue

                before = _snapshot(order)
                changed = False

                created_at_cell = row_cells[header_map["created at"]] if "created at" in header_map else None
                product_desc_cell = row_cells[header_map["product desc"]] if "product desc" in header_map else None
                receiver_name_cell = row_cells[header_map["receiver name"]] if "receiver name" in header_map else None
                receiver_phone_cell = row_cells[header_map["receiver phone"]] if "receiver phone" in header_map else None
                receiver_address_cell = row_cells[header_map["receiver address"]] if "receiver address" in header_map else None
                remark_cell = row_cells[header_map["remark"]] if "remark" in header_map else None
                reason_cell = row_cells[header_map["reason"]] if "reason" in header_map else None
                qty_cell = row_cells[header_map["qty"]] if "qty" in header_map else None
                price_cell = row_cells[header_map["price"]] if "price" in header_map else None
                cod_cell = row_cells[header_map["cod"]] if "cod" in header_map else None
                delivery_fee_cell = row_cells[header_map["delivery fee"]] if "delivery fee" in header_map else None
                additional_fee_cell = row_cells[header_map["additional fee"]] if "additional fee" in header_map else None
                province_fee_cell = row_cells[header_map["province fee"]] if "province fee" in header_map else None

                new_created_at = _parse_created_at_excel(created_at_cell.value if created_at_cell else "")
                if new_created_at is not None:
                    current_created_at = order.created_at
                    if current_created_at and timezone.is_naive(current_created_at):
                        current_created_at = timezone.make_aware(
                            current_created_at,
                            timezone.get_current_timezone(),
                        )
                    if current_created_at != new_created_at:
                        order.created_at = new_created_at
                        changed = True

                if product_desc_cell:
                    new_product_desc = _cell_str(product_desc_cell)
                    if order.product_desc != new_product_desc:
                        order.product_desc = new_product_desc
                        changed = True

                if receiver_name_cell:
                    new_receiver_name = _cell_str(receiver_name_cell)
                    if order.receiver_name != new_receiver_name:
                        order.receiver_name = new_receiver_name
                        changed = True

                if receiver_phone_cell:
                    new_receiver_phone = _cell_str(receiver_phone_cell)
                    if order.receiver_phone != new_receiver_phone:
                        order.receiver_phone = new_receiver_phone
                        changed = True

                if receiver_address_cell:
                    new_receiver_address = _cell_str(receiver_address_cell)
                    if order.receiver_address != new_receiver_address:
                        order.receiver_address = new_receiver_address
                        changed = True

                if remark_cell:
                    new_remark = _cell_str(remark_cell)
                    if (order.remark or "") != new_remark:
                        order.remark = new_remark
                        changed = True

                if reason_cell:
                    new_reason = _cell_str(reason_cell)
                    if (order.reason or "") != new_reason:
                        order.reason = new_reason
                        changed = True

                if qty_cell and qty_cell.value not in (None, ""):
                    new_qty = int(qty_cell.value)
                    if order.quantity != new_qty:
                        order.quantity = new_qty
                        changed = True

                if price_cell and price_cell.value not in (None, ""):
                    new_price = _to_decimal(price_cell.value, order.price)
                    if order.price != new_price:
                        order.price = new_price
                        changed = True

                if cod_cell and cod_cell.value not in (None, ""):
                    new_cod = _to_decimal(cod_cell.value, order.cod)
                    if order.cod != new_cod:
                        order.cod = new_cod
                        changed = True

                if delivery_fee_cell and delivery_fee_cell.value not in (None, ""):
                    new_delivery_fee = _to_decimal(delivery_fee_cell.value, order.delivery_fee)
                    if order.delivery_fee != new_delivery_fee:
                        order.delivery_fee = new_delivery_fee
                        changed = True

                if additional_fee_cell and additional_fee_cell.value not in (None, ""):
                    new_additional_fee = _to_decimal(additional_fee_cell.value, order.additional_fee)
                    if order.additional_fee != new_additional_fee:
                        order.additional_fee = new_additional_fee
                        changed = True

                if province_fee_cell and province_fee_cell.value not in (None, ""):
                    new_province_fee = _to_decimal(province_fee_cell.value, order.province_fee)
                    if order.province_fee != new_province_fee:
                        order.province_fee = new_province_fee
                        changed = True

                if changed:
                    with transaction.atomic():
                        order.save()
                        after = _snapshot(order)

                        BulkUpdateRow.objects.create(
                            batch=batch,
                            order=order,
                            status="UPDATED",
                            before_json=before,
                            after_json=after,
                        )
                    updated_count += 1
                else:
                    skipped_count += 1
                    BulkUpdateRow.objects.create(
                        batch=batch,
                        order=order,
                        status="SKIPPED",
                        before_json=before,
                        after_json=before,
                    )

            except Exception as e:
                error_count += 1
                errors.append(f"Row {row_idx}: {str(e)}")
                BulkUpdateRow.objects.create(
                    batch=batch,
                    status="ERROR",
                )

        request.session["bulk_update_errors"] = errors[:200]

        messages.success(
            request,
            f"Bulk update complete. Updated: {updated_count}, Skipped: {skipped_count}, Errors: {error_count}",
        )
        return redirect("bulk_update_batch_detail", batch_id=batch.id)

    show = _is_search_clicked(request)
    batch_qs = BulkUpdateBatch.objects.all().order_by("-id")

    if show:
        start_dt, end_dt = _dt_range_from_request(request)
        if start_dt and end_dt:
            batch_qs = batch_qs.filter(created_at__gte=start_dt, created_at__lte=end_dt)

        file_q = (request.GET.get("filename_contains") or request.GET.get("q") or "").strip()
        if file_q:
            batch_qs = batch_qs.filter(filename__icontains=file_q)
    else:
        batch_qs = BulkUpdateBatch.objects.none()

    paginator = Paginator(batch_qs, PER_PAGE)
    page_obj = paginator.get_page(request.GET.get("page"))

    total_batches = batch_qs.count() if show else 0
    total_updated_sum = 0

    if show:
        for batch in page_obj:
            batch.total_orders = BulkUpdateRow.objects.filter(batch=batch, status="UPDATED").count()
            total_updated_sum += batch.total_orders

    today = timezone.localdate().isoformat()
    errors = request.session.pop("bulk_update_errors", [])

    return render(
        request,
        "orders/bulk_update.html",
        {
            "errors": errors,
            "page_obj": page_obj,
            "search": show,
            "default_from": today,
            "default_to": today,
            "total_batches": total_batches,
            "total_updated_sum": total_updated_sum,
        },
    )


@login_required
def bulk_update_batch_detail(request: HttpRequest, batch_id: int):
    batch = get_object_or_404(BulkUpdateBatch, id=batch_id)

    rows = BulkUpdateRow.objects.select_related("order").filter(batch=batch).order_by("-id")

    q = (request.GET.get("q") or "").strip()
    if q:
        rows = rows.filter(
            Q(order__tracking_no__icontains=q)
            | Q(order__receiver_phone__icontains=q)
            | Q(order__receiver_name__icontains=q)
            | Q(order__seller_order_code__icontains=q)
        )

    items = [{"row": r, "before": r.before(), "after": r.after()} for r in rows]

    paginator = Paginator(items, PER_PAGE)
    page_obj = paginator.get_page(request.GET.get("page"))
    params = _build_params(request, drop={"page"})

    return render(
        request,
        "orders/bulk_update_batch_detail.html",
        {
            "batch": batch,
            "page_obj": page_obj,
            "params": params,
        },
    )


@login_required
def download_bulk_update_batch_excel(request: HttpRequest, batch_id: int):
    batch = get_object_or_404(BulkUpdateBatch, id=batch_id)
    order_ids = BulkUpdateRow.objects.filter(batch=batch).values_list("order_id", flat=True).distinct()
    qs = Order.objects.filter(is_deleted=False, id__in=order_ids).select_related("seller").order_by("-id")
    return _excel_orders_response(qs, f"bulk_update_batch_{batch.id}.xlsx")


upload_update = bulk_update


# ============================================================
# Create / Edit / Detail / Label / Bulk Action / Print
# ============================================================
@login_required
def create_order(request: HttpRequest):
    errors: list[str] = []
    form_data = {}
    sellers = Seller.objects.filter(is_active=True).order_by("name")

    if request.method == "POST":
        seller_id = (request.POST.get("seller") or "").strip()
        seller_order_code = (request.POST.get("seller_order_code") or "").strip()
        seller_name = (request.POST.get("seller_name") or "").strip()

        receiver_name = (request.POST.get("receiver_name") or "").strip()
        receiver_phone = (request.POST.get("receiver_phone") or "").strip()
        receiver_address = (request.POST.get("receiver_address") or "").strip()

        product_desc = (request.POST.get("product_desc") or "").strip()
        quantity = _to_int(request.POST.get("quantity"), 1)

        price = _to_decimal(request.POST.get("price"), 0)
        cod = _to_decimal(request.POST.get("cod"), 0)
        delivery_fee = _to_decimal(request.POST.get("delivery_fee"), 0)
        additional_fee = _to_decimal(request.POST.get("additional_fee"), 0)
        province_fee = _to_decimal(request.POST.get("province_fee"), 0)

        remark = (request.POST.get("remark") or "").strip()
        reason = (request.POST.get("reason") or "").strip()

        form_data = {
            "seller_order_code": seller_order_code,
            "seller_name": seller_name,
            "receiver_name": receiver_name,
            "receiver_phone": receiver_phone,
            "receiver_address": receiver_address,
            "product_desc": product_desc,
            "quantity": str(quantity),
            "price": str(price),
            "cod": str(cod),
            "delivery_fee": str(delivery_fee),
            "additional_fee": str(additional_fee),
            "province_fee": str(province_fee),
            "remark": remark,
            "reason": reason,
        }

        if not seller_id.isdigit():
            errors.append("Please select shop from dropdown.")
        if not receiver_name:
            errors.append("Receiver name is required.")
        if not receiver_phone:
            errors.append("Receiver phone is required.")
        if not receiver_address:
            errors.append("Receiver address is required.")

        if not errors:
            seller = Seller.objects.get(id=int(seller_id))

            o = Order.objects.create(
                tracking_no=f"TEMP-{timezone.now().timestamp()}",
                seller=seller,
                seller_code=seller.code or "",
                seller_name=seller_name or None,
                seller_order_code=seller_order_code or None,
                product_desc=product_desc or None,
                quantity=quantity,
                price=price,
                cod=cod,
                delivery_fee=delivery_fee,
                additional_fee=additional_fee,
                province_fee=province_fee,
                receiver_name=receiver_name,
                receiver_phone=receiver_phone,
                receiver_address=receiver_address,
                remark=remark or None,
                reason=reason or None,
                status=Order.STATUS_CREATED,
            )

            o.tracking_no = _make_tracking_no(o.id)
            o.save(update_fields=["tracking_no"])

            add_audit_log(
                module=AuditLog.MODULE_ORDER,
                obj=o,
                action=AuditLog.ACTION_CREATE,
                user=request.user,
                note="Created order",
            )

            add_order_activity(
                order=o,
                action=OrderActivity.ACTION_CREATE,
                user=request.user,
                new_status=o.status,
                note="Order created",
            )

            messages.success(request, f"✅ Successfully created order: {o.tracking_no}")
            return redirect("order_created", pk=o.id)

    return render(
        request,
        "orders/create_order.html",
        {
            "errors": errors,
            "form_data": form_data,
            "sellers": sellers,
        },
    )


@login_required
def order_edit(request: HttpRequest, pk: int):
    order = get_object_or_404(Order, pk=pk, is_deleted=False)

    original_id = order.id
    original_tracking = order.tracking_no

    sellers = Seller.objects.filter(is_active=True).order_by("name")
    errors: list[str] = []

    if order.is_locked:
        messages.error(request, "This order is locked and cannot be edited.")
        return redirect("order_detail", pk=order.id)

    is_admin = (
        request.user.is_superuser
        or request.user.groups.filter(name__in=["Admin", "Super Admin"]).exists()
    )

    is_done_or_delivered = (
        order.status == Order.STATUS_DELIVERED
        or bool(order.done_at)
    )

    old_status = order.status

    old_snapshot = {
        "seller_id": order.seller_id,
        "seller_code": order.seller_code,
        "seller_order_code": order.seller_order_code,
        "seller_name": order.seller_name,
        "product_desc": order.product_desc,
        "quantity": order.quantity,
        "price": str(order.price),
        "cod": str(order.cod),
        "delivery_fee": str(order.delivery_fee),
        "additional_fee": str(order.additional_fee),
        "province_fee": str(order.province_fee),
        "receiver_name": order.receiver_name,
        "receiver_phone": order.receiver_phone,
        "receiver_address": order.receiver_address,
        "remark": order.remark,
        "reason": order.reason,
    }

    if request.method == "POST":
        seller_id = (request.POST.get("seller") or "").strip()
        override_password = (request.POST.get("override_password") or "").strip()

        old_cod_decimal = order.cod
        cod_override_used = False

        if seller_id:
            if not seller_id.isdigit():
                errors.append("Invalid shop.")
            else:
                seller = Seller.objects.filter(id=int(seller_id), is_active=True).first()
                if not seller:
                    errors.append("Selected shop not found.")
                else:
                    order.seller = seller
                    order.seller_code = seller.code or ""

        order.seller_order_code = (request.POST.get("seller_order_code") or "").strip() or None
        order.seller_name = (request.POST.get("seller_name") or "").strip() or None
        order.product_desc = (request.POST.get("product_desc") or "").strip() or None
        order.quantity = _to_int(request.POST.get("quantity"), 1)

        order.price = _to_decimal(request.POST.get("price"), order.price)

        posted_cod = _to_decimal(request.POST.get("cod"), order.cod)

        can_change_cod = True
        if is_done_or_delivered and posted_cod != old_cod_decimal:
            if is_admin:
                can_change_cod = True
            elif override_password and override_password == getattr(settings, "ORDER_COD_OVERRIDE_PASSWORD", ""):
                can_change_cod = True
                cod_override_used = True
            else:
                can_change_cod = False

        if can_change_cod:
            order.cod = posted_cod
        else:
            errors.append("COD cannot be updated after delivered/done. Only admin or correct override password can change it.")

        order.delivery_fee = _to_decimal(request.POST.get("delivery_fee"), order.delivery_fee)
        order.additional_fee = _to_decimal(request.POST.get("additional_fee"), order.additional_fee)
        order.province_fee = _to_decimal(request.POST.get("province_fee"), order.province_fee)

        order.receiver_name = (request.POST.get("receiver_name") or "").strip() or None
        order.receiver_phone = (request.POST.get("receiver_phone") or "").strip() or None
        order.receiver_address = (request.POST.get("receiver_address") or "").strip() or None
        order.remark = (request.POST.get("remark") or "").strip() or None
        order.reason = (request.POST.get("reason") or "").strip() or None

        if not order.receiver_name:
            errors.append("Receiver name is required.")
        if not order.receiver_phone:
            errors.append("Receiver phone is required.")
        if not order.receiver_address:
            errors.append("Receiver address is required.")

        if not errors:
            order.id = original_id
            order.tracking_no = original_tracking
            order.updated_at = timezone.now()
            order.updated_by = request.user
            order.save()

            new_snapshot = {
                "seller_id": order.seller_id,
                "seller_code": order.seller_code,
                "seller_order_code": order.seller_order_code,
                "seller_name": order.seller_name,
                "product_desc": order.product_desc,
                "quantity": order.quantity,
                "price": str(order.price),
                "cod": str(order.cod),
                "delivery_fee": str(order.delivery_fee),
                "additional_fee": str(order.additional_fee),
                "province_fee": str(order.province_fee),
                "receiver_name": order.receiver_name,
                "receiver_phone": order.receiver_phone,
                "receiver_address": order.receiver_address,
                "remark": order.remark,
                "reason": order.reason,
            }

            if old_snapshot != new_snapshot:
                audit_note = "Edited order"
                if old_snapshot["cod"] != new_snapshot["cod"]:
                    if cod_override_used:
                        audit_note = "Edited order | COD changed with override password"
                    elif is_admin and is_done_or_delivered:
                        audit_note = "Edited order | COD changed by admin after delivered/done"

                add_audit_log(
                    module=AuditLog.MODULE_ORDER,
                    obj=order,
                    action=AuditLog.ACTION_UPDATE,
                    user=request.user,
                    old_value=json.dumps(old_snapshot, ensure_ascii=False),
                    new_value=json.dumps(new_snapshot, ensure_ascii=False),
                    note=audit_note,
                )

                note_parts = []

                if old_snapshot["seller_id"] != new_snapshot["seller_id"]:
                    note_parts.append("Shop changed")

                if old_snapshot["seller_name"] != new_snapshot["seller_name"]:
                    note_parts.append(f"Seller name: {old_snapshot['seller_name'] or '-'} -> {new_snapshot['seller_name'] or '-'}")

                if old_snapshot["seller_order_code"] != new_snapshot["seller_order_code"]:
                    note_parts.append(f"Order code: {old_snapshot['seller_order_code'] or '-'} -> {new_snapshot['seller_order_code'] or '-'}")

                if old_snapshot["product_desc"] != new_snapshot["product_desc"]:
                    note_parts.append("Product updated")

                if old_snapshot["quantity"] != new_snapshot["quantity"]:
                    note_parts.append(f"Qty: {old_snapshot['quantity']} -> {new_snapshot['quantity']}")

                if old_snapshot["price"] != new_snapshot["price"]:
                    note_parts.append(f"Price: {old_snapshot['price']} -> {new_snapshot['price']}")

                if old_snapshot["cod"] != new_snapshot["cod"]:
                    cod_note = f"COD: {old_snapshot['cod']} -> {new_snapshot['cod']}"
                    if cod_override_used:
                        cod_note += " (override password used)"
                    elif is_admin and is_done_or_delivered:
                        cod_note += " (changed by admin after delivered/done)"
                    note_parts.append(cod_note)

                if old_snapshot["delivery_fee"] != new_snapshot["delivery_fee"]:
                    note_parts.append(f"Delivery fee: {old_snapshot['delivery_fee']} -> {new_snapshot['delivery_fee']}")

                if old_snapshot["additional_fee"] != new_snapshot["additional_fee"]:
                    note_parts.append(f"Additional fee: {old_snapshot['additional_fee']} -> {new_snapshot['additional_fee']}")

                if old_snapshot["province_fee"] != new_snapshot["province_fee"]:
                    note_parts.append(f"Province fee: {old_snapshot['province_fee']} -> {new_snapshot['province_fee']}")

                if old_snapshot["receiver_name"] != new_snapshot["receiver_name"]:
                    note_parts.append(f"Receiver name: {old_snapshot['receiver_name'] or '-'} -> {new_snapshot['receiver_name'] or '-'}")

                if old_snapshot["receiver_phone"] != new_snapshot["receiver_phone"]:
                    note_parts.append(f"Receiver phone: {old_snapshot['receiver_phone'] or '-'} -> {new_snapshot['receiver_phone'] or '-'}")

                if old_snapshot["receiver_address"] != new_snapshot["receiver_address"]:
                    note_parts.append("Receiver address updated")

                if old_snapshot["remark"] != new_snapshot["remark"]:
                    note_parts.append("Remark updated")

                if old_snapshot["reason"] != new_snapshot["reason"]:
                    note_parts.append(f"Reason: {old_snapshot['reason'] or '-'} -> {new_snapshot['reason'] or '-'}")

                timeline_note = " | ".join(note_parts) if note_parts else "Order edited"

                add_order_activity(
                    order=order,
                    action=OrderActivity.ACTION_EDIT,
                    user=request.user,
                    shipper=order.delivery_shipper,
                    old_status=old_status,
                    new_status=order.status,
                    note=timeline_note,
                )

            messages.success(request, "✅ Order updated.")
            return redirect("order_created", pk=order.id)

    return render(
        request,
        "orders/order_edit.html",
        {
            "order": order,
            "errors": errors,
            "sellers": sellers,
            "is_done_or_delivered": is_done_or_delivered,
            "is_admin_user": is_admin,
        },
    )


@login_required
def order_detail(request: HttpRequest, pk: int):
    order = get_object_or_404(Order, pk=pk, is_deleted=False)
    return render(request, "orders/order_detail.html", {"order": order})


@login_required
def order_label(request: HttpRequest, pk: int):
    order = get_object_or_404(Order, pk=pk, is_deleted=False)

    Order.objects.filter(pk=order.pk).update(print_count=F("print_count") + 1)
    order.refresh_from_db()

    add_audit_log(
        module=AuditLog.MODULE_ORDER,
        obj=order,
        action=AuditLog.ACTION_PRINT,
        user=request.user,
        note="Printed label",
    )

    add_order_activity(
        order=order,
        action=OrderActivity.ACTION_PRINT,
        user=request.user,
        shipper=order.delivery_shipper,
        old_status=order.status,
        new_status=order.status,
        note="Printed label",
    )

    setting = _get_setting()
    rate = Decimal(str(setting.usd_to_khr or 4100))
    cod_usd = Decimal(str(order.cod or 0))
    cod_khr = cod_usd * rate

    sn = (order.tracking_no or "").strip()
    qr_src = _make_qr_data_uri(sn)
    print_dt = timezone.localtime(timezone.now())

    return render(
        request,
        "orders/order_label.html",
        {
            "order": order,
            "sn": sn,
            "print_dt": print_dt,
            "rate": rate,
            "cod_usd": cod_usd,
            "cod_khr_text": _fmt_khr_no_decimal(cod_khr),
            "qr_src": qr_src,
        },
    )


@login_required
@transaction.atomic
def order_bulk_action(request: HttpRequest):
    if request.method != "POST":
        return redirect("order_list")

    action = (request.POST.get("action") or "").strip().lower()
    ids = request.POST.getlist("ids")
    clean_ids = [int(x) for x in ids if str(x).isdigit()]

    if not clean_ids:
        messages.error(request, "Please select at least 1 order.")
        return redirect(request.META.get("HTTP_REFERER", "order_list"))

    if action == "delete":
        qs = Order.objects.filter(id__in=clean_ids, is_deleted=False)
        affected = list(qs.values_list("id", "tracking_no"))
        count = _soft_delete_queryset(qs, request.user)

        for oid, tracking_no in affected:
            add_audit_log(
                module=AuditLog.MODULE_ORDER,
                obj=type("TmpObj", (), {"pk": oid, "__str__": lambda self, t=tracking_no: t})(),
                action=AuditLog.ACTION_DELETE,
                user=request.user,
                note="Bulk moved to trash",
            )

        messages.success(request, f"✅ Deleted {count} orders.")
        return redirect(request.META.get("HTTP_REFERER", "order_list"))

    if action == "print":
        Order.objects.filter(id__in=clean_ids).update(print_count=F("print_count") + 1)
        ids_str = ",".join(str(i) for i in clean_ids)
        return redirect(f"/orders/batch-print/?ids={ids_str}")

    messages.error(request, "Unknown action.")
    return redirect(request.META.get("HTTP_REFERER", "order_list"))


@login_required
def order_batch_print(request: HttpRequest):
    ids_str = (request.GET.get("ids") or "").strip()
    id_list = [int(x) for x in ids_str.split(",") if x.strip().isdigit()]

    orders = Order.objects.filter(id__in=id_list, is_deleted=False).select_related("seller").order_by("id")

    setting = _get_setting()
    rate = Decimal(str(setting.usd_to_khr or 4100))
    print_dt = timezone.localtime(timezone.now())

    items = []
    for order in orders:
        sn = (order.tracking_no or "").strip()
        cod_usd = Decimal(str(order.cod or 0))
        cod_khr = cod_usd * rate

        items.append(
            {
                "order": order,
                "sn": sn,
                "print_dt": print_dt,
                "rate": rate,
                "cod_usd": cod_usd,
                "cod_khr_text": _fmt_khr_no_decimal(cod_khr),
                "qr_src": _make_qr_data_uri(sn),
            }
        )

    return render(request, "orders/order_batch_print.html", {"items": items})


@login_required
def order_setting_rate(request: HttpRequest):
    setting = _get_setting()

    if request.method == "POST":
        v = (request.POST.get("usd_to_khr") or "").strip()
        try:
            old_value = str(setting.usd_to_khr)
            setting.usd_to_khr = Decimal(v)
            setting.save()

            add_audit_log(
                module=AuditLog.MODULE_ORDER,
                obj=setting,
                action=AuditLog.ACTION_UPDATE,
                user=request.user,
                field_name="usd_to_khr",
                old_value=old_value,
                new_value=str(setting.usd_to_khr),
                note="Updated exchange rate",
            )

            messages.success(request, "✅ Rate updated successfully.")
            return redirect("order_setting_rate")
        except Exception:
            messages.error(request, "❌ Invalid number. Example: 4100")
            return redirect("order_setting_rate")

    return render(request, "orders/order_setting_rate.html", {"setting": setting})


@login_required
@transaction.atomic
def delete_import_batch(request: HttpRequest, batch_id: int):
    batch = get_object_or_404(ImportBatch, id=batch_id)

    if request.method != "POST":
        return redirect("import_batch_detail", batch_id=batch.id)

    qs = Order.objects.filter(import_batch=batch, is_deleted=False)
    count = _soft_delete_queryset(qs, request.user)

    add_audit_log(
        module=AuditLog.MODULE_IMPORT,
        obj=batch,
        action=AuditLog.ACTION_ROLLBACK_IMPORT,
        user=request.user,
        note=f"Soft deleted {count} order(s) from import batch before deleting batch row",
    )

    batch.delete()

    messages.success(request, "✅ Import batch deleted successfully.")
    return redirect("import_orders")


# ============================================================
# Trash / Restore
# ============================================================
@login_required
def order_delete(request, pk):
    order = get_object_or_404(Order, pk=pk, is_deleted=False)

    if request.method == "POST":
        order.soft_delete(user=request.user)

        add_audit_log(
            module=AuditLog.MODULE_ORDER,
            obj=order,
            action=AuditLog.ACTION_DELETE,
            user=request.user,
            note="Moved to trash",
        )

        messages.success(request, f"Order {order.tracking_no} moved to trash.")
        return redirect("order_list")

    return render(request, "orders/order_confirm_delete.html", {"order": order})


@login_required
def order_restore(request, pk):
    order = get_object_or_404(Order, pk=pk, is_deleted=True)

    if request.method == "POST":
        order.restore()

        add_audit_log(
            module=AuditLog.MODULE_ORDER,
            obj=order,
            action=AuditLog.ACTION_RESTORE,
            user=request.user,
            note="Restored from trash",
        )

        messages.success(request, f"Order {order.tracking_no} restored.")
        return redirect("order_trash")

    return render(request, "orders/order_restore_confirm.html", {"order": order})


@login_required
def order_trash(request):
    show = _is_search_clicked(request)

    rows = Order.objects.filter(is_deleted=True).select_related("deleted_by", "seller").order_by("-deleted_at", "-id")

    if show:
        tracking = (request.GET.get("tracking") or "").strip()
        seller_id = (request.GET.get("seller") or "").strip()
        deleted_by = (request.GET.get("deleted_by") or "").strip()

        start_dt, end_dt = _dt_range_from_request(request)
        if start_dt and end_dt:
            rows = rows.filter(deleted_at__gte=start_dt, deleted_at__lte=end_dt)

        if tracking:
            rows = rows.filter(tracking_no__icontains=tracking)

        if seller_id.isdigit():
            rows = rows.filter(seller_id=int(seller_id))

        if deleted_by:
            rows = rows.filter(
                Q(deleted_by__username__icontains=deleted_by)
                | Q(deleted_by__first_name__icontains=deleted_by)
                | Q(deleted_by__last_name__icontains=deleted_by)
            )
    else:
        rows = Order.objects.none()

    sellers = Seller.objects.filter(is_active=True).order_by("name")
    today = timezone.localdate().isoformat()
    params = _build_params(request, drop={"page"})

    return render(
        request,
        "orders/order_trash.html",
        {
            "rows": rows,
            "sellers": sellers,
            "search": show,
            "default_from": today,
            "default_to": today,
            "params": params,
        },
    )


@login_required
def audit_log_list(request):
    show = _is_search_clicked(request)

    rows = AuditLog.objects.select_related("created_by").order_by("-id")

    if show:
        module = (request.GET.get("module") or "").strip()
        action = (request.GET.get("action") or "").strip()
        user_q = (request.GET.get("user_q") or "").strip()
        object_q = (request.GET.get("object_q") or "").strip()

        start_dt, end_dt = _dt_range_from_request(request)
        if start_dt and end_dt:
            rows = rows.filter(created_at__gte=start_dt, created_at__lte=end_dt)

        if module:
            rows = rows.filter(module=module)

        if action:
            rows = rows.filter(action=action)

        if user_q:
            rows = rows.filter(
                Q(created_by__username__icontains=user_q)
                | Q(created_by__first_name__icontains=user_q)
                | Q(created_by__last_name__icontains=user_q)
            )

        if object_q:
            rows = rows.filter(
                Q(object_repr__icontains=object_q)
                | Q(note__icontains=object_q)
                | Q(old_value__icontains=object_q)
                | Q(new_value__icontains=object_q)
            )
    else:
        rows = AuditLog.objects.none()

    today = timezone.localdate().isoformat()
    params = _build_params(request, drop={"page"})

    return render(
        request,
        "orders/audit_log_list.html",
        {
            "rows": rows,
            "search": show,
            "default_from": today,
            "default_to": today,
            "params": params,
            "module_choices": AuditLog.MODULE_CHOICES,
            "action_choices": AuditLog.ACTION_CHOICES,
        },
    )


@login_required
def order_created(request: HttpRequest, pk: int):
    order = get_object_or_404(Order, pk=pk, is_deleted=False)

    activities = order.activities.select_related("actor", "shipper").all().order_by("-id")
    logs = AuditLog.objects.filter(
        module=AuditLog.MODULE_ORDER,
        object_id=order.id,
    ).select_related("created_by").order_by("-id")[:50]

    return render(
        request,
        "orders/order_created.html",
        {
            "order": order,
            "activities": activities,
            "logs": logs,
        },
    )